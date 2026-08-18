"""
VinCert — certificate OCR / parse desktop app.
"""

from __future__ import annotations

from pathlib import Path
import json
import re
import shutil
import sys
import threading
import webbrowser

import customtkinter
import tkinter as tk
from tkinter import filedialog
from tkinter import font as tkfont

from vincert.models import CertificateFields, ParseResult
from vincert.pipeline import parse_certificate
from vincert.folder_import import find_pdfs_in_folder
from vincert.mas_autofill import (
    AutofillControl,
    AutofillItem,
    close_keepalive_browser,
    load_credentials,
    next_export_path,
    resolve_eams_environment,
    run_mas_autofill,
    save_credentials,
    write_batch_excel,
)
from vincert.parse_metrology import LABEL_ALIASES, parse_fields
from vincert.pdf_preview import PdfPreviewController

customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("blue")

PROJECT_ROOT = Path(__file__).resolve().parent
UI_SETTINGS_PATH = PROJECT_ROOT / "ui_settings.json"
UI_SCALE_NORMAL = 1.0
UI_SCALE_ZOOMED = 1.2
DEFAULT_AUTOFILL_STEP_DELAY_SEC = 1.0
AUTOFILL_EXIT_WARN_MS = 10_000


FAILED_ITEMS_SUBDIR = "failed_items"

# Settings → 解析规则：user can add label aliases for these fields.
PARSE_RULE_FIELDS = [
    ("name", "计量器具名称"),
    ("serial_num", "计量器具编号"),
    ("manufacturer", "制造厂"),
    ("model", "型号/规格"),
    ("certificate_no", "证书编号"),
    ("client_name", "客户名称"),
    ("measurement_date", "本次检测日期"),
    ("due_date", "本次检测有效期至"),
    ("issue_date", "发布日期"),
    ("measurement_unit", "检测机构"),
    ("measurement_type", "检验方式"),
]
PARSE_RULE_FIELD_LABELS = {key: label for key, label in PARSE_RULE_FIELDS}


def load_ui_settings(path: Path | None = None) -> dict:
    """Load persisted UI prefs from ui_settings.json."""
    settings_path = Path(path or UI_SETTINGS_PATH)
    defaults = {
        "ui_zoomed": True,
        "ocr_enabled": False,
        "buttons_bold": True,
        "content_centering": True,
        "status_dots": True,
        "doc_list_scale_fonts": True,
        "hide_scrollbars": False,
        "compact_min_window": False,
        "auto_window_snap": True,
        "pdf_preview_enabled": True,
        # Default OFF for the UI toggle "使用提取结果":
        # - off => system auto-determination
        "valid_till_from_system": True,
        "feature_extensions": False,
        "testing_mode": False,
        "demo_folder_enabled": False,
        "demo_folder": "",
        "autofill_step_delay_sec": DEFAULT_AUTOFILL_STEP_DELAY_SEC,
        "parse_rules": {},
    }
    if not settings_path.exists():
        return dict(defaults)
    try:
        data = json.loads(settings_path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return dict(defaults)
    except Exception:  # noqa: BLE001
        return dict(defaults)
    out = dict(defaults)
    if "ui_zoomed" in data:
        out["ui_zoomed"] = bool(data["ui_zoomed"])
    if "ocr_enabled" in data:
        out["ocr_enabled"] = bool(data["ocr_enabled"])
    if "buttons_bold" in data:
        out["buttons_bold"] = bool(data["buttons_bold"])
    if "content_centering" in data:
        out["content_centering"] = bool(data["content_centering"])
    if "status_dots" in data:
        out["status_dots"] = bool(data["status_dots"])
    if "doc_list_scale_fonts" in data:
        out["doc_list_scale_fonts"] = bool(data["doc_list_scale_fonts"])
    if "hide_scrollbars" in data:
        out["hide_scrollbars"] = bool(data["hide_scrollbars"])
    if "compact_min_window" in data:
        out["compact_min_window"] = bool(data["compact_min_window"])
    if "auto_window_snap" in data:
        out["auto_window_snap"] = bool(data["auto_window_snap"])
    if "pdf_preview_enabled" in data:
        out["pdf_preview_enabled"] = bool(data["pdf_preview_enabled"])
    if "valid_till_from_system" in data:
        out["valid_till_from_system"] = bool(data["valid_till_from_system"])
    if "feature_extensions" in data:
        out["feature_extensions"] = bool(data["feature_extensions"])
    if "testing_mode" in data:
        out["testing_mode"] = bool(data["testing_mode"])
    if "demo_folder_enabled" in data:
        out["demo_folder_enabled"] = bool(data["demo_folder_enabled"])
    if "demo_folder" in data and data["demo_folder"]:
        out["demo_folder"] = str(data["demo_folder"])
    if "autofill_step_delay_sec" in data:
        try:
            delay = float(data["autofill_step_delay_sec"])
            out["autofill_step_delay_sec"] = max(0.0, min(30.0, delay))
        except (TypeError, ValueError):
            pass
    if "parse_rules" in data:
        out["parse_rules"] = _normalize_parse_rules(data.get("parse_rules"))
    return out


def _normalize_parse_rules(raw) -> dict[str, list[str]]:
    """Sanitize persisted label-alias rules: field key → list of aliases."""
    allowed = {key for key, _label in PARSE_RULE_FIELDS} | set(LABEL_ALIASES.keys())
    if not isinstance(raw, dict):
        return {}
    out: dict[str, list[str]] = {}
    for key, values in raw.items():
        field = str(key or "").strip()
        if field not in allowed:
            continue
        aliases: list[str] = []
        seen: set[str] = set()
        if isinstance(values, (list, tuple)):
            seq = values
        elif values:
            seq = [values]
        else:
            seq = []
        for item in seq:
            alias = str(item or "").strip()
            if not alias or alias in seen:
                continue
            seen.add(alias)
            aliases.append(alias)
        if aliases:
            out[field] = aliases
    return out


def load_parse_rules(path: Path | None = None) -> dict[str, list[str]]:
    """User-defined label aliases for certificate field parsing."""
    return _normalize_parse_rules(load_ui_settings(path).get("parse_rules"))


def save_parse_rules(rules: dict[str, list[str]], path: Path | None = None) -> dict[str, list[str]]:
    cleaned = _normalize_parse_rules(rules)
    save_ui_settings(parse_rules=cleaned)
    return cleaned


def save_ui_settings(**updates) -> Path:
    settings_path = Path(UI_SETTINGS_PATH)
    data = load_ui_settings(settings_path)
    data.update(updates)
    settings_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return settings_path


def load_ui_zoomed(path: Path | None = None) -> bool:
    """Return True when large UI scale is enabled (default: zoomed)."""
    return bool(load_ui_settings(path).get("ui_zoomed", True))


def save_ui_zoomed(zoomed: bool, path: Path | None = None) -> Path:
    return save_ui_settings(ui_zoomed=bool(zoomed))


def load_ocr_enabled(path: Path | None = None) -> bool:
    """Return True when OCR extraction is enabled (default: off)."""
    return bool(load_ui_settings(path).get("ocr_enabled", False))


def save_ocr_enabled(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(ocr_enabled=bool(enabled))


def load_buttons_bold(path: Path | None = None) -> bool:
    """Return True when button labels use bold weight (default: on)."""
    return bool(load_ui_settings(path).get("buttons_bold", True))


def save_buttons_bold(bold: bool, path: Path | None = None) -> Path:
    return save_ui_settings(buttons_bold=bool(bold))


def load_content_centering(path: Path | None = None) -> bool:
    """Return True when page content is vertically centered when it fits."""
    return bool(load_ui_settings(path).get("content_centering", True))


def save_content_centering(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(content_centering=bool(enabled))


def load_status_dots(path: Path | None = None) -> bool:
    """Return True when approved/removed status uses red/green dots."""
    return bool(load_ui_settings(path).get("status_dots", True))


def save_status_dots(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(status_dots=bool(enabled))


def load_doc_list_scale_fonts(path: Path | None = None) -> bool:
    """Return True when document-list canvas text follows UI zoom scaling."""
    return bool(load_ui_settings(path).get("doc_list_scale_fonts", True))


def save_doc_list_scale_fonts(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(doc_list_scale_fonts=bool(enabled))


def load_hide_scrollbars(path: Path | None = None) -> bool:
    """Return True when native CTk scrollbars are hidden (wheel scroll still works)."""
    return bool(load_ui_settings(path).get("hide_scrollbars", False))


def save_hide_scrollbars(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(hide_scrollbars=bool(enabled))


def load_compact_min_window(path: Path | None = None) -> bool:
    """Return True when a smaller window minsize is allowed."""
    return bool(load_ui_settings(path).get("compact_min_window", False))


def save_compact_min_window(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(compact_min_window=bool(enabled))


def load_auto_window_snap(path: Path | None = None) -> bool:
    """Return True when VinCert auto-snaps beside the browser / PDF preview."""
    return bool(load_ui_settings(path).get("auto_window_snap", True))


def save_auto_window_snap(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(auto_window_snap=bool(enabled))


def load_pdf_preview_enabled(path: Path | None = None) -> bool:
    """Return True when selecting a certificate opens the Chromium PDF preview."""
    return bool(load_ui_settings(path).get("pdf_preview_enabled", True))


def save_pdf_preview_enabled(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(pdf_preview_enabled=bool(enabled))


def load_feature_extensions(path: Path | None = None) -> bool:
    """Return True when extended automation tools (e.g. custom-dir) are shown."""
    return bool(load_ui_settings(path).get("feature_extensions", False))


def save_feature_extensions(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(feature_extensions=bool(enabled))


def load_testing_mode(path: Path | None = None) -> bool:
    """Return True when EAMS UAT testing environment is selected."""
    return bool(load_ui_settings(path).get("testing_mode", False))


def save_testing_mode(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(testing_mode=bool(enabled))


def load_demo_folder(path: Path | None = None) -> str:
    """Return the configured startup certificate folder path (may be empty)."""
    raw = load_ui_settings(path).get("demo_folder") or ""
    return str(raw).strip()


def save_demo_folder(folder: str | Path | None, path: Path | None = None) -> str:
    value = "" if folder is None else str(Path(folder).expanduser())
    if value:
        try:
            value = str(Path(value).resolve())
        except Exception:  # noqa: BLE001
            pass
    save_ui_settings(demo_folder=value)
    return value


def load_demo_folder_enabled(path: Path | None = None) -> bool:
    """Return True when startup should auto-load the configured certificate folder."""
    return bool(load_ui_settings(path).get("demo_folder_enabled", False))


def save_demo_folder_enabled(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(demo_folder_enabled=bool(enabled))


def load_autofill_step_delay_sec(path: Path | None = None) -> float:
    """Seconds to pause between autofill steps (default 1.0)."""
    raw = load_ui_settings(path).get(
        "autofill_step_delay_sec", DEFAULT_AUTOFILL_STEP_DELAY_SEC
    )
    try:
        return max(0.0, min(30.0, float(raw)))
    except (TypeError, ValueError):
        return DEFAULT_AUTOFILL_STEP_DELAY_SEC


def save_autofill_step_delay_sec(seconds: float, path: Path | None = None) -> float:
    value = max(0.0, min(30.0, float(seconds)))
    save_ui_settings(autofill_step_delay_sec=value)
    return value


def load_valid_till_from_system(path: Path | None = None) -> bool:
    """When True, let EAMS system auto-determination set 有效期至."""
    return bool(load_ui_settings(path).get("valid_till_from_system", False))


def save_valid_till_from_system(enabled: bool, path: Path | None = None) -> Path:
    return save_ui_settings(valid_till_from_system=bool(enabled))


def apply_ui_scale(zoomed: bool) -> float:
    scale = UI_SCALE_ZOOMED if zoomed else UI_SCALE_NORMAL
    customtkinter.set_widget_scaling(scale)
    # Keep window scaling fixed. Changing it with zoom (esp. zoom-out in
    # fullscreen) shrinks the drawable area and leaves black letterbox bars.
    customtkinter.set_window_scaling(1.0)
    return scale


# Apply saved scale before widgets are constructed.
apply_ui_scale(load_ui_zoomed())

DOC_SIDEBAR_WIDTH = 400
DOC_SIDEBAR_MARGIN = 10
MAIN_MIN_WIDTH = 480
MAIN_MAX_WIDTH = int(DOC_SIDEBAR_WIDTH * 2.5)  # center view cap
CONTENT_WRAP = MAIN_MAX_WIDTH - 48
DOC_WRAP = DOC_SIDEBAR_WIDTH - 32
STEP_TILE_HEIGHT = 72
BRAND_TILE_HEIGHT = 48
NAV_BADGE_SIZE = 36
SETTINGS_BTN_HEIGHT = 40
SMALL_BTN_HEIGHT = 36
ENTRY_HEIGHT = 44
PRIMARY_ACTION_BTN_HEIGHT = 45  # 45×1.2 = 54px — avoids CTk odd-height text bias when zoomed
UI_RADIUS = 12  # shared corner radius for panels + buttons
BUILD_VERSION = "v1.0.1"
BUILD_DATE = "18/08/2026"
RELEASES_URL = "https://github.com/AketfONG/VinCert/releases"

# Typography — sizes chosen for readability at both 1.0× and 1.2× UI scale.
FONT_BRAND = 22
FONT_TITLE = 18
FONT_SECTION = 15
FONT_BODY = 14
FONT_LABEL = 13
FONT_ENTRY = 15
FONT_META = 13
FONT_CAPTION = 12
FONT_BUTTON = 14
FONT_STEP = 15
FONT_BADGE = 18

STEPS = [
    ("extract", "提取核对", "1"),
    ("automate", "自动化", "2"),
]

_theme = customtkinter.ThemeManager.theme
APP_BG_COLOR = _theme["CTk"]["fg_color"]
_theme_frame = _theme["CTkFrame"]
_frame_fg = _theme_frame["fg_color"]
_frame_top = _theme_frame["top_fg_color"]
_dark_frame_fg = _frame_fg[1] if isinstance(_frame_fg, (tuple, list)) else _frame_fg
_dark_frame_top = _frame_top[1] if isinstance(_frame_top, (tuple, list)) else _frame_top
_theme_textbox = _theme.get("CTkTextbox", {})
_textbox_fg = _theme_textbox.get("fg_color", ("#F9F9FA", "#1D1E1E"))
if isinstance(_textbox_fg, (tuple, list)):
    _dark_field_fg = _textbox_fg[1]
else:
    _dark_field_fg = _textbox_fg
# Light mode: white surfaces for sidebar modules, doc panel, and fields.
EMBED_BG_COLOR = ("#ffffff", _dark_frame_fg)
ACTIVE_OUTLINE = ("#3b8ed0", "#1f6aa5")
TILE_BG_NORMAL = ("#ffffff", _dark_frame_fg)
TILE_BG_HOVER = ("#eef1f4", _dark_frame_top)
TILE_BG_ACTIVE = ("#e3f0fa", "#343638")  # light blue wash / softer dark fill
SECONDARY_BTN_FG = ("#c9d1d9", "#3d444b")
SECONDARY_BTN_HOVER = ("#dde4eb", "#556068")
SECONDARY_BTN_TEXT = ("gray10", "gray90")
OUTLINE_BTN_HOVER = ("gray70", "gray30")
PRIMARY_BTN_FG = ("#3B8ED0", "#1F6AA5")
PRIMARY_BTN_HOVER = ("#5BA3DB", "#2E83C4")
PRIMARY_BTN_TEXT = ("#DCE4EE", "#DCE4EE")
SUCCESS_BTN_FG = ("#2d8a4e", "#2d8a4e")
SUCCESS_BTN_HOVER = ("#38a460", "#38a460")
SUCCESS_BTN_TEXT = ("#ffffff", "#ffffff")
DANGER_BTN_FG = ("#c0392b", "#c0392b")
DANGER_BTN_HOVER = ("#e74c3c", "#e74c3c")
DANGER_BTN_TEXT = ("#ffffff", "#ffffff")
CUSTOM_AUTOFILL_BTN_FG = ("#7c3aed", "#6d28d9")
CUSTOM_AUTOFILL_BTN_HOVER = ("#8b5cf6", "#7c3aed")
CUSTOM_AUTOFILL_BTN_TEXT = ("#ffffff", "#ffffff")
FEATURE_ACCENT_OUTLINE = CUSTOM_AUTOFILL_BTN_FG
FEATURE_ACCENT_TILE_ACTIVE = ("#ede9fe", "#343638")
FEATURE_DOC_ROW_ACTIVE = CUSTOM_AUTOFILL_BTN_FG
FEATURE_DOC_MARK_NUMBER_ACTIVE = ("#c4b5fd", "#a78bfa")
TOAST_BG = ("#ffffff", "#1a1a1a")
TOAST_TITLE_COLOR = ("gray10", "#ffffff")
TOAST_MESSAGE_COLOR = ("gray30", "#f0f0f0")
TOAST_PROGRESS_TRACK = ("#e5e5e5", "#2a2a2a")
TOAST_WIDTH = 360  # compact fixed toast; do not stretch to ops column
TOAST_PAD = 12
TOAST_BTN_HEIGHT = 40
TOAST_BORDER_WIDTH = 2  # match active tile / settings outline
TOAST_RADIUS = UI_RADIUS + 2  # 2px rounder than shared UI radius
TOAST_MIN_MS = 10000
TOAST_DEFAULT_MS = 10000
TOAST_SUCCESS_MS = 10000
TOAST_TICK_MS = 50
TOAST_STACK_MAX = 8
TOAST_STACK_GAP = 8
AUTOFILL_LOG_WIDTH = TOAST_WIDTH
AUTOFILL_LOG_PAD = TOAST_PAD
AUTOFILL_LOG_FINISH_MS = 12000
AUTOFILL_LOG_BUBBLE = 44
# Extra gap between the shrunk terminal pill and the toast stack.
AUTOFILL_LOG_BUBBLE_TOAST_GAP = 20
DOC_ROW_ACTIVE = ("#3b8ed0", "#1f6aa5")
DOC_ROW_ACTIVE_TEXT = ("#ffffff", "#ffffff")
# Index numbers at ~50% opacity (emoji marks stay full strength).
DOC_MARK_NUMBER_COLOR = ("gray50", "gray50")
DOC_MARK_NUMBER_ACTIVE = ("#9dc6e7", "#9fb5d2")  # white blended ~50% onto selected blue
DOC_STATUS_DOT = "●"
DOC_STATUS_DOT_OK = ("#2d8a4e", "#38a460")
DOC_STATUS_DOT_BAD = ("#c0392b", "#e74c3c")
DOC_STATUS_DOT_SIZE = FONT_BODY + 2
DOC_ROW_HEIGHT = 36
DOC_MARK_COL_WIDTH = 40
DOC_NAME_TIP_PADX = 8
DOC_NAME_TIP_PADY = 4
RESULT_INFO_HEIGHT = 128  # ~4 taller entry rows
FIELD_FG_COLOR = ("#ffffff", _dark_field_fg)
FIELD_TEXT_COLOR = _theme_textbox.get("text_color", ("gray10", "#DCE4EE"))
FIELD_FG_COLOR_DISABLED = ("gray90", "gray22")
FIELD_TEXT_COLOR_DISABLED = ("gray55", "gray55")
# Autofill UI lock: mute chrome that `state=disabled` alone does not cover.
UI_LOCK_BTN_FG = ("#d5dbe0", "#3a3a3a")
UI_LOCK_BTN_TEXT = ("#8b9298", "#777777")
UI_LOCK_TILE_FG = ("#eceff1", "#2c2c2c")
UI_LOCK_BADGE_FG = ("#b0b6bc", "#555555")
UI_LOCK_LABEL = ("gray55", "gray55")
UI_LOCK_DOC_ROW = ("#dfe3e7", "#2f2f2f")
UI_LOCK_DOC_TEXT = ("gray55", "gray55")
# Match fill so CTkEntry's 1px border stays invisible (avoids a hairline under fields).
FIELD_BORDER_WIDTH = 1

MATCH_FIELDS = [
    ("name", "计量器具名称"),
    ("serial_num", "计量器具编号"),
    ("manufacturer", "制造厂"),
]

METROLOGY_FIELDS = [
    ("measurement_type", "检验方式"),
    ("measurement_date", "本次检测日期"),
    ("measurement_unit", "检测机构"),
]

# Review-only user fields (not OCR-extracted).
REVIEW_USER_FIELDS = [
    ("result_info", "计量结果信息"),
]

REVIEW_FILL_FIELDS = METROLOGY_FIELDS + REVIEW_USER_FIELDS
EXTRACT_DISPLAY_FIELDS = MATCH_FIELDS + METROLOGY_FIELDS
REVIEW_DISPLAY_FIELDS = MATCH_FIELDS + REVIEW_FILL_FIELDS
DEFAULT_RESULT_INFO = "合格"
EXPORT_COLUMNS = [
    ("name", "名称"),
    ("serial_num", "编号"),
    ("model", "型号"),
    ("measurement_unit", "计量单位"),
    ("measurement_date", "计量日期"),
    ("measurement_type", "计量类型"),
]

# Sidebar 设置 stays highlighted on these nested pages.
SETTINGS_SUBPAGES = {"settings", "parse_rules", "more_options"}


WINDOW_MIN_HEIGHT_NORMAL = 830  # design units × 1.0 scale
WINDOW_MIN_HEIGHT_ZOOMED = 820  # design units × 1.2 scale — tune separately from normal
WINDOW_MIN_WIDTH = DOC_SIDEBAR_WIDTH + MAIN_MIN_WIDTH
COMPACT_DOC_SIDEBAR_WIDTH = 320
COMPACT_MAIN_COL_MIN = 360
WINDOW_MIN_WIDTH_COMPACT = COMPACT_DOC_SIDEBAR_WIDTH + COMPACT_MAIN_COL_MIN
WINDOW_MIN_HEIGHT_COMPACT = 520


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("VinCert")
        _boot_min_h = (
            WINDOW_MIN_HEIGHT_ZOOMED
            if load_ui_zoomed()
            else WINDOW_MIN_HEIGHT_NORMAL
        )
        self.geometry(
            f"{WINDOW_MIN_WIDTH + 40}x{max(900, _boot_min_h + 40)}"
        )

        self._imported_files: list[str] = []
        self._parse_results: dict[str, ParseResult] = {}
        self._autofill_queue: list[str] = []
        self._removed_paths: set[str] = set()
        self._current_cert_index = 0
        self._current_step = "extract"
        # Workflow gate: field editing unlocks after quarantine.
        # 自动化 unlocks only after at least one certificate is approved.
        self._workflow_phase = "extract"  # "extract" | "review"
        self._step_before_settings = "extract"

        self._source_folder: str | None = None
        self._selected_path: str | None = None
        self._doc_rows: dict[str, dict] = {}
        self._doc_hover_path: str | None = None
        self._doc_name_tip: customtkinter.CTkFrame | None = None
        self._doc_name_tip_path: str | None = None
        self._extract_busy = False
        self._autofill_busy = False
        self._autofill_control: AutofillControl | None = None
        self._autofill_exit_confirming = False
        self._autofill_was_paused_before_exit = False
        self._autofill_disabled_widgets: list = []
        self._autofill_ui_chrome_locked = False
        self._toast_host: customtkinter.CTkFrame | None = None
        self._toasts: list[dict] = []
        self._toast_seq = 0
        self._toast_tick_after_id: str | None = None
        self._autofill_log_frame: customtkinter.CTkFrame | None = None
        self._autofill_log_text: customtkinter.CTkTextbox | None = None
        self._autofill_log_status: customtkinter.CTkLabel | None = None
        self._autofill_log_header: customtkinter.CTkFrame | None = None
        self._autofill_log_bubble: customtkinter.CTkButton | None = None
        self._autofill_log_collapsed = False
        self._autofill_log_finish_after_id: str | None = None
        self._autofill_log_accent = SUCCESS_BTN_FG
        self._pending_quarantine_paths: list[str] = []
        self._ui_zoomed = load_ui_zoomed()
        self._ocr_enabled = load_ocr_enabled()
        self._buttons_bold = load_buttons_bold()
        self._content_centering = load_content_centering()
        self._status_dots = load_status_dots()
        self._doc_list_scale_fonts = load_doc_list_scale_fonts()
        self._hide_scrollbars = load_hide_scrollbars()
        self._compact_min_window = load_compact_min_window()
        self._auto_window_snap = load_auto_window_snap()
        self._pdf_preview_enabled = load_pdf_preview_enabled()
        self._feature_extensions = load_feature_extensions()
        self._accent_buttons: list[customtkinter.CTkButton] = []
        self._testing_mode = load_testing_mode()
        self._demo_folder_enabled = load_demo_folder_enabled()
        self._demo_folder = load_demo_folder()
        self._autofill_step_delay_sec = load_autofill_step_delay_sec()
        self._valid_till_from_system = load_valid_till_from_system()
        self._parse_rules: dict[str, list[str]] = load_parse_rules()
        self._parse_rules_field_key = PARSE_RULE_FIELDS[0][0]
        self._content_wrap_labels: list[customtkinter.CTkLabel] = []
        self._pdf_preview_layout_active = False
        # Ignore the next preview-closed callback (intentional close before re-snap).
        self._pdf_preview_suppress_restore = False
        self._pdf_preview = PdfPreviewController(
            on_closed=lambda: self.after(0, self._on_pdf_preview_closed),
            status=lambda msg: self.after(0, lambda m=msg: self.set_status(m)),
        )

        self.grid_rowconfigure(0, weight=1)
        self._apply_layout_column_minsizes()

        self._build_doc_sidebar()
        self._build_controls_panel()
        self._apply_min_window_size()
        self._apply_doc_sidebar_width()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.show_step("extract")
        self._apply_feature_extension_theme()
        self._update_extract_ocr_ui()
        self.after_idle(self._bootstrap_window_layout)
        if self._hide_scrollbars:
            self.after_idle(self._refresh_scrollbar_visibility)

    def _widget_scaling_factor(self) -> float:
        if hasattr(self, "controls_inner"):
            return max(
                customtkinter.ScalingTracker.get_widget_scaling(self.controls_inner),
                0.01,
            )
        return max(customtkinter.ScalingTracker.widget_scaling, 0.01)

    def _doc_sidebar_width(self) -> int:
        if getattr(self, "_compact_min_window", False):
            return COMPACT_DOC_SIDEBAR_WIDTH
        return DOC_SIDEBAR_WIDTH

    def _apply_doc_sidebar_width(self):
        width = self._doc_sidebar_width()
        if hasattr(self, "doc_sidebar"):
            self.doc_sidebar.configure(width=width)
            self.doc_sidebar.grid(row=0, column=0, sticky="nsew")
        if hasattr(self, "folder_label"):
            self.folder_label.configure(wraplength=max(80, (width - 32) // 2))
        if hasattr(self, "doc_list_frame"):
            self._schedule_doc_list_scrollbar_sync()
            self._refresh_doc_list_fonts()

    def _apply_layout_column_minsizes(self):
        # Tk grid minsize is in pixels; CTk widget widths are design units × scale.
        scale = self._widget_scaling_factor()
        sidebar_min = self._doc_sidebar_width()
        main_min = (
            COMPACT_MAIN_COL_MIN
            if getattr(self, "_compact_min_window", False)
            else MAIN_MIN_WIDTH
        )
        self.grid_columnconfigure(
            0, weight=0, minsize=int(sidebar_min * scale)
        )
        self.grid_columnconfigure(
            1, weight=1, minsize=int(main_min * scale)
        )

    def _bootstrap_window_layout(self):
        """Fullscreen on launch; demo autoload may then snap for a PDF preview."""
        self._apply_window_fullscreen()
        self._maybe_autoload_demo_folder()

    def _screen_work_area(self) -> tuple[int, int, int, int]:
        """Return (x, y, width, height) for the usable desktop area.

        On Windows this uses the monitor work rect (excludes the taskbar /
        bottom navigation bar). Falls back to full screen metrics elsewhere.
        """
        if sys.platform == "win32":
            try:
                import ctypes
                from ctypes import wintypes

                class RECT(ctypes.Structure):
                    _fields_ = [
                        ("left", wintypes.LONG),
                        ("top", wintypes.LONG),
                        ("right", wintypes.LONG),
                        ("bottom", wintypes.LONG),
                    ]

                class MONITORINFO(ctypes.Structure):
                    _fields_ = [
                        ("cbSize", wintypes.DWORD),
                        ("rcMonitor", RECT),
                        ("rcWork", RECT),
                        ("dwFlags", wintypes.DWORD),
                    ]

                user32 = ctypes.windll.user32
                # Prefer the monitor that contains this window (multi-monitor safe).
                try:
                    hwnd = self._win_toplevel_hwnd() or int(self.winfo_id())
                    MONITOR_DEFAULTTONEAREST = 2
                    monitor = user32.MonitorFromWindow(hwnd, MONITOR_DEFAULTTONEAREST)
                    info = MONITORINFO()
                    info.cbSize = ctypes.sizeof(MONITORINFO)
                    if monitor and user32.GetMonitorInfoW(monitor, ctypes.byref(info)):
                        work = info.rcWork
                        return (
                            int(work.left),
                            int(work.top),
                            int(work.right - work.left),
                            int(work.bottom - work.top),
                        )
                except Exception:  # noqa: BLE001
                    pass

                rect = RECT()
                SPI_GETWORKAREA = 0x0030
                if user32.SystemParametersInfoW(
                    SPI_GETWORKAREA, 0, ctypes.byref(rect), 0
                ):
                    return (
                        int(rect.left),
                        int(rect.top),
                        int(rect.right - rect.left),
                        int(rect.bottom - rect.top),
                    )
            except Exception:  # noqa: BLE001
                pass

        self.update_idletasks()
        # macOS / fallback: full screen size (Dock may still overlap slightly).
        return 0, 0, int(self.winfo_screenwidth()), int(self.winfo_screenheight())

    def _win_toplevel_hwnd(self) -> int | None:
        """Return the Win32 HWND for this Tk toplevel (outer frame), if available."""
        if sys.platform != "win32":
            return None
        try:
            import ctypes

            user32 = ctypes.windll.user32
            GA_ROOT = 2
            # Prefer the OS root window; winfo_id() alone is often an inner child.
            try:
                child = int(self.winfo_id())
                root = int(user32.GetAncestor(child, GA_ROOT) or 0)
                if root:
                    return root
            except Exception:  # noqa: BLE001
                pass
            # wm_frame() may be "0x...." — use base 0 so the prefix parses.
            try:
                frame = str(self.wm_frame()).strip()
                if frame:
                    return int(frame, 0)
            except Exception:  # noqa: BLE001
                pass
            return int(self.winfo_id())
        except Exception:  # noqa: BLE001
            return None

    def _frame_chrome_size(self) -> tuple[int, int]:
        """Return (extra_width, extra_height) of the OS frame beyond the Tk client.

        Playwright bounds are outer-window sizes. Tk ``geometry`` width/height
        are the client area — using work height as client height pushes the
        title bar into the taskbar. Subtract this chrome so VinCert's outer
        height matches the Playwright window.
        """
        if sys.platform == "win32":
            hwnd = self._win_toplevel_hwnd()
            if hwnd:
                try:
                    import ctypes
                    from ctypes import wintypes

                    class RECT(ctypes.Structure):
                        _fields_ = [
                            ("left", wintypes.LONG),
                            ("top", wintypes.LONG),
                            ("right", wintypes.LONG),
                            ("bottom", wintypes.LONG),
                        ]

                    user32 = ctypes.windll.user32
                    outer = RECT()
                    client = RECT()
                    if user32.GetWindowRect(
                        wintypes.HWND(hwnd), ctypes.byref(outer)
                    ) and user32.GetClientRect(
                        wintypes.HWND(hwnd), ctypes.byref(client)
                    ):
                        ow = int(outer.right) - int(outer.left)
                        oh = int(outer.bottom) - int(outer.top)
                        cw = int(client.right) - int(client.left)
                        ch = int(client.bottom) - int(client.top)
                        if ow > cw >= 0 and oh > ch >= 0:
                            return ow - cw, oh - ch
                except Exception:  # noqa: BLE001
                    pass
            return 16, 39
        if sys.platform == "darwin":
            return 0, 28
        return 0, 0

    def _set_app_bounds(self, left: int, top: int, width: int, height: int) -> None:
        """Place VinCert via Tk geometry so *outer* size matches work/Playwright."""
        left, top = int(left), int(top)
        width, height = max(400, int(width)), max(400, int(height))
        try:
            if sys.platform == "darwin":
                try:
                    self.attributes("-fullscreen", False)
                except Exception:  # noqa: BLE001
                    pass
            self.state("normal")
        except Exception:  # noqa: BLE001
            pass
        self.update_idletasks()
        chrome_w, chrome_h = self._frame_chrome_size()
        client_w = max(400, width - chrome_w)
        client_h = max(400, height - chrome_h)
        self.geometry(f"{client_w}x{client_h}+{left}+{top}")
        self.update_idletasks()

    def _raise_workspace_windows(self) -> None:
        """Keep VinCert + Chromium above other apps during layout (not permanent).

        Other windows often steal Z-order while we resize/split. Briefly raise
        both; do not leave always-on-top so the user can still switch apps.
        """
        try:
            self.lift()
            self.attributes("-topmost", True)
            self.update_idletasks()
            self.attributes("-topmost", False)
        except Exception:  # noqa: BLE001
            pass

        chrome_open = bool(getattr(self, "_pdf_preview", None) and self._pdf_preview.is_open)
        if chrome_open:
            try:
                self._pdf_preview.raise_window()
            except Exception:  # noqa: BLE001
                pass

        if sys.platform != "win32":
            return
        try:
            from vincert.win_snap import find_chrome_hwnd, raise_hwnds_above_others

            app_hwnd = self._win_toplevel_hwnd()
            chrome_hwnd = find_chrome_hwnd() if chrome_open else None
            # App first so it ends focused; both stay above other apps.
            raise_hwnds_above_others(app_hwnd, chrome_hwnd)
        except Exception:  # noqa: BLE001
            pass

    def _apply_window_fullscreen(self):
        """Maximize / fill the work area when no side browser layout is needed."""
        if not getattr(self, "_auto_window_snap", True):
            self._pdf_preview_layout_active = False
            return
        self._pdf_preview_layout_active = False
        try:
            if sys.platform == "darwin":
                try:
                    self.attributes("-fullscreen", False)
                except Exception:  # noqa: BLE001
                    pass
            self.state("normal")
        except Exception:  # noqa: BLE001
            pass
        self.update_idletasks()
        if sys.platform == "win32":
            try:
                self.state("zoomed")
                self._raise_workspace_windows()
                return
            except Exception:  # noqa: BLE001
                pass
        x, y, w, h = self._screen_work_area()
        self._set_app_bounds(x, y, w, h)
        if sys.platform == "darwin":
            try:
                self.state("zoomed")
            except Exception:  # noqa: BLE001
                pass
        self._raise_workspace_windows()

    def _snap_app_left_half(self):
        """Place VinCert on the left half of the work area via Tk geometry."""
        if not getattr(self, "_auto_window_snap", True):
            return
        x, y, w, h = self._screen_work_area()
        half = max(640, w // 2)
        self._set_app_bounds(x, y, half, h)
        self._pdf_preview_layout_active = True
        self._raise_workspace_windows()

    def _browser_profile_dir(self) -> Path:
        """Persistent Chromium profile shared by PDF preview + EAMS tabs."""
        return resolve_eams_environment(testing=bool(self._testing_mode)).user_data_dir

    def _default_browser_bounds(self) -> tuple[int, int, int, int]:
        """Right half of the work area — same outer height as VinCert."""
        work_x, work_y, work_w, work_h = self._screen_work_area()
        half = max(400, work_w // 2)
        _, chrome_h = self._frame_chrome_size()
        app_h = max(int(self.winfo_height()), 1) + chrome_h
        shared_h = max(400, min(work_h, app_h if app_h >= 400 else work_h))
        return (work_x + work_w - half, work_y, half, shared_h)

    def _prepare_side_browser_layout(self) -> tuple[int, int, int, int]:
        """Optionally snap VinCert left and return bounds for the shared browser."""
        # Never close PDF / EAMS tabs here — they share one Chromium window.
        if self._auto_window_snap:
            self._snap_app_left_half()
            self.update_idletasks()
            return self._pdf_preview_bounds_remaining()
        self.update_idletasks()
        return self._default_browser_bounds()

    def _pdf_preview_bounds_remaining(self) -> tuple[int, int, int, int]:
        """Size the browser to the unused work-area space beside VinCert.

        Height matches VinCert's outer frame so both sit on the same bottom
        edge (above the taskbar / home bar).
        """
        self.update_idletasks()
        work_x, work_y, work_w, work_h = self._screen_work_area()
        work_right = work_x + work_w
        work_bottom = work_y + work_h

        app_left = int(self.winfo_rootx())
        app_top = int(self.winfo_rooty())
        app_width = max(int(self.winfo_width()), 1)
        app_height = max(int(self.winfo_height()), 1)
        chrome_w, chrome_h = self._frame_chrome_size()
        app_right = app_left + app_width
        app_outer_h = app_height + chrome_h
        shared_h = max(400, min(work_h, app_outer_h if app_outer_h >= 400 else work_h))

        # Prefer the strip to the right of the app within the work area.
        right_width = work_right - app_right
        if right_width >= 400:
            left = max(app_right, work_x)
            top = work_y
            width = work_right - left
            return (left, top, max(400, width), shared_h)

        # Prefer the strip to the left of the app.
        left_width = app_left - work_x
        if left_width >= 400:
            return (work_x, work_y, max(400, left_width), shared_h)

        # Prefer space below the app (unusual, but better than overlapping).
        app_bottom = app_top + app_height
        below = work_bottom - app_bottom
        if below >= 300:
            return (
                work_x,
                max(app_bottom, work_y),
                max(400, work_w),
                max(300, below),
            )

        # Last resort: right half of the work area.
        half = max(400, work_w // 2)
        return (work_x + work_w - half, work_y, half, shared_h)

    def _sync_window_layout_to_browser(self) -> None:
        """Fullscreen when no Playwright window; otherwise keep split."""
        if not self._auto_window_snap:
            return
        if self._pdf_preview.is_open:
            if not self._pdf_preview_layout_active:
                self._snap_app_left_half()
            return
        self._apply_window_fullscreen()

    def _sync_pdf_preview(self):
        """Open/update PDF tab when a file is selected; close PDF tab otherwise."""
        if not self._pdf_preview_enabled:
            if self._pdf_preview.has_pdf:
                self._pdf_preview.close()
            elif self._auto_window_snap and not self._pdf_preview.is_open:
                # No preview and no shared browser — restore fullscreen if we were split.
                if self._pdf_preview_layout_active:
                    self._apply_window_fullscreen()
            return

        path = self._selected_path
        profile = self._browser_profile_dir()
        if path and Path(path).is_file():
            if self._autofill_busy:
                return
            if self._auto_window_snap and not self._pdf_preview_layout_active:
                self._snap_app_left_half()
            self.update_idletasks()
            if self._auto_window_snap:
                bounds = (
                    self._pdf_preview_bounds_remaining()
                    if self._pdf_preview.is_open
                    else self._prepare_side_browser_layout()
                )
            else:
                bounds = (
                    self._pdf_preview_bounds_remaining()
                    if self._pdf_preview.is_open
                    else self._default_browser_bounds()
                )
            self._pdf_preview.show(path, bounds, profile_dir=profile)
            self._pdf_preview.focus()
            self._raise_workspace_windows()
            return
        # No document selected — close PDF tab only; keep EAMS if present.
        if self._pdf_preview.has_pdf:
            self._pdf_preview.close()
        self._sync_window_layout_to_browser()

    def _on_pdf_preview_closed(self):
        """Called when the shared Chromium window fully closes (not PDF-tab-only)."""
        try:
            if not self.winfo_exists():
                return
        except Exception:  # noqa: BLE001
            return
        if self._pdf_preview_suppress_restore:
            self._pdf_preview_suppress_restore = False
            return
        if self._autofill_busy:
            return
        if self._pdf_preview.is_open:
            return
        if not self._auto_window_snap:
            self._pdf_preview_layout_active = False
            return
        self._apply_window_fullscreen()

    def _track_content_wrap(self, label: customtkinter.CTkLabel) -> customtkinter.CTkLabel:
        self._content_wrap_labels.append(label)
        return label

    def _update_content_wraplengths(self, wrap: int):
        for label in self._content_wrap_labels:
            try:
                label.configure(wraplength=wrap)
            except Exception:  # noqa: BLE001
                pass

    def _button_font(self, size: int = FONT_BUTTON) -> customtkinter.CTkFont:
        """Shared button typeface; weight follows the 按钮加粗 setting."""
        return customtkinter.CTkFont(
            size=size,
            weight="bold" if self._buttons_bold else "normal",
        )

    def _fix_ctk_button_text_vcenter(self, button: customtkinter.CTkButton) -> None:
        """Re-balance CTkButton label grid so bold text stays centered under UI scale."""
        label = getattr(button, "_text_label", None)
        if label is None:
            return
        try:
            label.configure(anchor="center")
            base = max(int(button._border_width) + 1, int(getattr(button, "_border_spacing", 2) or 2))
            # Zoomed bold glyphs sit slightly low; give the bottom row +1 design-unit.
            extra = 1 if self._ui_zoomed and self._buttons_bold else 0
            top = button._apply_widget_scaling(base)
            bottom = button._apply_widget_scaling(base + extra)
            button.grid_rowconfigure(0, weight=1000, minsize=top)
            button.grid_rowconfigure(4, weight=1000, minsize=bottom)
            button.grid_rowconfigure((1, 3), weight=1)
        except Exception:  # noqa: BLE001
            pass

    @staticmethod
    def _fix_ctk_entry_bottom_line(entry: customtkinter.CTkEntry) -> None:
        """CTkEntry pads the inner widget with (bw, bw+1); that extra bottom px shows as a line."""

        def _create_grid():
            entry._canvas.grid(column=0, row=0, sticky="nswe")
            bw = entry._apply_widget_scaling(entry._border_width)
            if entry._corner_radius >= entry._minimum_x_padding:
                padx = min(
                    entry._apply_widget_scaling(entry._corner_radius),
                    round(entry._apply_widget_scaling(entry._current_height / 2)),
                )
            else:
                padx = entry._apply_widget_scaling(entry._minimum_x_padding)
            entry._entry.grid(
                column=0, row=0, sticky="nswe", padx=padx, pady=(bw, bw)
            )

        entry._create_grid = _create_grid
        entry._create_grid()

    def _make_field_entry(
        self,
        parent,
        *,
        placeholder: str = "",
        height: int = ENTRY_HEIGHT,
        show: str | None = None,
    ) -> customtkinter.CTkEntry:
        kwargs = {}
        if show is not None:
            kwargs["show"] = show
        entry = customtkinter.CTkEntry(
            parent,
            height=height,
            corner_radius=UI_RADIUS,
            placeholder_text=placeholder,
            border_width=FIELD_BORDER_WIDTH,
            border_color=FIELD_FG_COLOR,
            fg_color=FIELD_FG_COLOR,
            text_color=FIELD_TEXT_COLOR,
            font=customtkinter.CTkFont(size=FONT_ENTRY),
            **kwargs,
        )
        self._fix_ctk_entry_bottom_line(entry)
        return entry

    def _style_primary_action_button(self, button: customtkinter.CTkButton) -> customtkinter.CTkButton:
        self._fix_ctk_button_text_vcenter(button)
        return button

    def _feature_accent_active(self) -> bool:
        return bool(getattr(self, "_feature_extensions", False))

    def _accent_outline(self):
        return FEATURE_ACCENT_OUTLINE if self._feature_accent_active() else ACTIVE_OUTLINE

    def _accent_tile_active_bg(self):
        return FEATURE_ACCENT_TILE_ACTIVE if self._feature_accent_active() else TILE_BG_ACTIVE

    def _accent_doc_row_active(self):
        return FEATURE_DOC_ROW_ACTIVE if self._feature_accent_active() else DOC_ROW_ACTIVE

    def _accent_doc_mark_number_active(self):
        return (
            FEATURE_DOC_MARK_NUMBER_ACTIVE
            if self._feature_accent_active()
            else DOC_MARK_NUMBER_ACTIVE
        )

    def _accent_primary_btn_colors(self) -> tuple:
        if self._feature_accent_active():
            return CUSTOM_AUTOFILL_BTN_FG, CUSTOM_AUTOFILL_BTN_HOVER, CUSTOM_AUTOFILL_BTN_TEXT
        return PRIMARY_BTN_FG, PRIMARY_BTN_HOVER, PRIMARY_BTN_TEXT

    def _make_accent_button(self, parent, **kwargs) -> customtkinter.CTkButton:
        fg, hover, text = self._accent_primary_btn_colors()
        kwargs.setdefault("fg_color", fg)
        kwargs.setdefault("hover_color", hover)
        kwargs.setdefault("text_color", text)
        btn = customtkinter.CTkButton(parent, **kwargs)
        self._accent_buttons.append(btn)
        return btn

    def _apply_feature_extension_theme(self) -> None:
        """Purple accent chrome when 功能拓展 is on; autofill CTA stays blue."""
        fg, hover, text = self._accent_primary_btn_colors()
        for btn in getattr(self, "_accent_buttons", []):
            try:
                btn.configure(fg_color=fg, hover_color=hover, text_color=text)
            except Exception:  # noqa: BLE001
                pass
        if hasattr(self, "autofill_button"):
            try:
                self.autofill_button.configure(
                    fg_color=PRIMARY_BTN_FG,
                    hover_color=PRIMARY_BTN_HOVER,
                    text_color=PRIMARY_BTN_TEXT,
                )
            except Exception:  # noqa: BLE001
                pass
        if hasattr(self, "step_tiles"):
            self._update_step_tiles(getattr(self, "_current_step", "extract"))
        if hasattr(self, "settings_button"):
            self._update_settings_button(
                getattr(self, "_current_step", None) in SETTINGS_SUBPAGES
            )
        if hasattr(self, "_doc_rows") and self._doc_rows:
            self._highlight_selected_doc()
        if getattr(self, "_doc_name_tip", None) is not None:
            self._hide_doc_name_tip()

    def _restyle_primary_action_buttons(self):
        for name in (
            "ocr_extract_button",
            "remove_failed_button",
            "approve_toggle_button",
            "remove_toggle_button",
            "export_excel_button",
            "autofill_button",
            "custom_dir_autofill_button",
            "autofill_pause_button",
            "autofill_exit_button",
        ):
            btn = getattr(self, name, None)
            if btn is not None:
                self._fix_ctk_button_text_vcenter(btn)

    # ------------------------------------------------------------------ layout
    def _create_step_tile(
        self,
        parent: customtkinter.CTkFrame,
        key: str,
        label: str,
        number: str,
    ) -> customtkinter.CTkFrame:
        tile = customtkinter.CTkFrame(
            parent,
            height=STEP_TILE_HEIGHT,
            corner_radius=UI_RADIUS,
            fg_color=TILE_BG_NORMAL,
            border_width=0,
            border_color=self._accent_outline(),
        )
        tile.grid_propagate(False)
        tile.grid_columnconfigure(1, weight=1)
        tile.grid_rowconfigure(0, weight=1)

        badge = customtkinter.CTkFrame(
            tile,
            width=NAV_BADGE_SIZE,
            height=NAV_BADGE_SIZE,
            corner_radius=NAV_BADGE_SIZE // 2,
            fg_color=self._accent_outline(),
        )
        badge.grid(row=0, column=0, padx=(12, 8), pady=12, sticky="w")
        badge.grid_propagate(False)
        customtkinter.CTkLabel(
            badge,
            text=number,
            font=customtkinter.CTkFont(
                family="SF Pro Display",
                size=FONT_BADGE,
                weight="bold",
            ),
            text_color=("#ffffff", "#ffffff"),
        ).place(relx=0.5, rely=0.5, anchor="center")

        customtkinter.CTkLabel(
            tile,
            text=label,
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_STEP, weight="bold"),
        ).grid(row=0, column=1, padx=(0, 12), sticky="ew")

        self._bind_step_tile_hover(tile, key)
        self._bind_step_tile_click(tile, key)
        tile.configure(cursor="hand2")
        return tile

    def _bind_step_tile_click(self, widget, key: str):
        widget.bind("<Button-1>", lambda _e, k=key: self._on_step_tile_click(k), add="+")
        for child in widget.winfo_children():
            self._bind_step_tile_click(child, key)

    def _on_step_tile_click(self, key: str):
        if self._autofill_busy:
            self.set_status("自动填写进行中，请先暂停或退出…")
            return
        if self._step_tile_is_locked(key):
            if key == "automate":
                self.set_status("请先批准至少一份证书")
            return
        if key == self._current_step:
            return
        self.show_step(key)

    def _bind_step_tile_hover(self, widget, key: str):
        widget.bind("<Enter>", lambda _e, k=key: self._hover_step_tile(k, True))
        widget.bind("<Leave>", lambda e, k=key: self._on_tile_leave(e, k))
        for child in widget.winfo_children():
            self._bind_step_tile_hover(child, key)

    def _on_tile_leave(self, event, key: str):
        tile = self.step_tiles[key]
        widget_under = tile.winfo_containing(event.x_root, event.y_root)
        if widget_under is None or not self._is_descendant(widget_under, tile):
            self._hover_step_tile(key, False)

    def _is_descendant(self, widget, ancestor) -> bool:
        while widget:
            if widget == ancestor:
                return True
            widget = widget.master
        return False

    def _hover_step_tile(self, key: str, entering: bool):
        if self._autofill_busy:
            return
        if key == self._current_step:
            return
        if self._step_tile_is_locked(key):
            return
        self._apply_tile_style(key, "hover" if entering else "normal")

    def _automate_step_unlocked(self) -> bool:
        """自动化 is reachable after an approve, or when 功能拓展 custom import is on."""
        if self._autofill_queue:
            return True
        return bool(getattr(self, "_feature_extensions", False))

    def _step_tile_is_locked(self, key: str) -> bool:
        """True when a workflow tile should appear greyed / inactive."""
        if key == "automate" and not self._automate_step_unlocked():
            return True
        return False

    def _apply_tile_style(self, key: str, style: str):
        if self._autofill_busy or self._autofill_ui_chrome_locked:
            return
        tile = self.step_tiles[key]
        if style == "locked" or (style != "active" and self._step_tile_is_locked(key)):
            tile.configure(
                fg_color=UI_LOCK_TILE_FG,
                border_color=UI_LOCK_TILE_FG,
                border_width=0,
            )
            for child in tile.winfo_children():
                try:
                    if isinstance(child, customtkinter.CTkFrame):
                        child.configure(fg_color=UI_LOCK_BADGE_FG)
                    elif isinstance(child, customtkinter.CTkLabel):
                        child.configure(text_color=UI_LOCK_LABEL)
                except Exception:  # noqa: BLE001
                    pass
            return
        styles = {
            "normal": {
                "tile_fg": TILE_BG_NORMAL,
                "tile_border": TILE_BG_NORMAL,
                "border_width": 0,
            },
            "hover": {
                "tile_fg": TILE_BG_HOVER,
                "tile_border": TILE_BG_HOVER,
                "border_width": 0,
            },
            "active": {
                "tile_fg": self._accent_tile_active_bg(),
                "tile_border": self._accent_outline(),
                "border_width": 2,
            },
        }
        colors = styles[style]
        tile.configure(
            fg_color=colors["tile_fg"],
            border_color=colors["tile_border"],
            border_width=colors["border_width"],
        )
        # Always restore badge/label after a locked (grey) state.
        for child in tile.winfo_children():
            try:
                if isinstance(child, customtkinter.CTkFrame):
                    child.configure(fg_color=self._accent_outline())
                    for sub in child.winfo_children():
                        if isinstance(sub, customtkinter.CTkLabel):
                            sub.configure(text_color=("#ffffff", "#ffffff"))
                elif isinstance(child, customtkinter.CTkLabel):
                    child.configure(text_color=SECONDARY_BTN_TEXT)
            except Exception:  # noqa: BLE001
                pass

    def _update_step_tiles(self, active_key: str):
        for key in self.step_tiles:
            if self._step_tile_is_locked(key):
                self._apply_tile_style(key, "locked")
            elif key == active_key:
                self._apply_tile_style(key, "active")
            else:
                self._apply_tile_style(key, "normal")

    def _reset_workflow_to_extract(self):
        """Back to extract phase (自动化 locked) — used on clear / new folder."""
        self._workflow_phase = "extract"
        self._update_review_fields_state()
        self._update_remove_failed_button()
        self._update_approve_toggle_button()
        self._update_remove_toggle_button()
        if self._current_step not in ("extract", *SETTINGS_SUBPAGES):
            self.show_step("extract")
        else:
            self._update_step_tiles(self._current_step)

    def _advance_to_review(self):
        """After failed certs are cleared — unlock field editing and 自动化."""
        if not self._imported_files:
            self.set_status("没有可核对的证书")
            self._reset_workflow_to_extract()
            return
        self._workflow_phase = "review"
        self._load_approve_fields_for_current()
        self._update_review_fields_state()
        self._update_remove_failed_button()
        self._update_step_tiles(self._current_step)
        self.set_status("已可核对填写")

    def _update_settings_button(self, active: bool):
        if not hasattr(self, "settings_button"):
            return
        if self._autofill_busy or self._autofill_ui_chrome_locked:
            return
        if active:
            self.settings_button.configure(
                border_width=2,
                border_color=self._accent_outline(),
                fg_color=self._accent_tile_active_bg(),
                hover_color=self._accent_tile_active_bg(),
                text_color=SECONDARY_BTN_TEXT,
            )
        else:
            self.settings_button.configure(
                border_width=0,
                border_color=SECONDARY_BTN_FG,
                fg_color=SECONDARY_BTN_FG,
                hover_color=SECONDARY_BTN_HOVER,
                text_color=SECONDARY_BTN_TEXT,
            )

    # ------------------------------------------------------------- doc sidebar
    def _build_doc_sidebar(self):
        self.doc_sidebar = customtkinter.CTkFrame(
            self,
            width=self._doc_sidebar_width(),
            corner_radius=0,
            fg_color=APP_BG_COLOR,
        )
        self.doc_sidebar.grid(row=0, column=0, sticky="nsew")
        self.doc_sidebar.grid_propagate(False)
        self.doc_sidebar.grid_columnconfigure(0, weight=1)
        self.doc_sidebar.grid_rowconfigure(1, weight=1)

        # Nav strip: brand + step buttons
        nav = customtkinter.CTkFrame(self.doc_sidebar, fg_color="transparent")
        nav.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=DOC_SIDEBAR_MARGIN,
            pady=(DOC_SIDEBAR_MARGIN, 0),
        )
        nav.grid_columnconfigure(0, weight=1)

        brand_tile = customtkinter.CTkFrame(
            nav,
            height=BRAND_TILE_HEIGHT,
            corner_radius=UI_RADIUS,
            fg_color="transparent",
        )
        brand_tile.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        brand_tile.grid_propagate(False)
        brand_tile.grid_columnconfigure(0, weight=1)
        brand_tile.grid_rowconfigure(0, weight=1)

        brand_inner = customtkinter.CTkFrame(brand_tile, fg_color="transparent")
        brand_inner.grid(row=0, column=0, sticky="ew", padx=4, pady=0)
        brand_inner.grid_columnconfigure(0, weight=1)

        customtkinter.CTkLabel(
            brand_inner,
            text="VinCert",
            font=customtkinter.CTkFont(size=FONT_BRAND, weight="bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")

        version_row = customtkinter.CTkFrame(brand_inner, fg_color="transparent")
        version_row.grid(row=0, column=1, sticky="e")
        version_row.grid_columnconfigure(0, weight=0)

        self.build_version_label = customtkinter.CTkLabel(
            version_row,
            text=f"{BUILD_VERSION} · {BUILD_DATE}",
            font=customtkinter.CTkFont(size=FONT_META),
            text_color="gray60",
            anchor="e",
        )
        self.build_version_label.grid(row=0, column=0, sticky="e", padx=(0, 6))

        customtkinter.CTkButton(
            version_row,
            corner_radius=UI_RADIUS,
            text="GitHub",
            width=64,
            height=28,
            fg_color=SECONDARY_BTN_FG,
            hover_color=SECONDARY_BTN_HOVER,
            text_color=SECONDARY_BTN_TEXT,
            font=self._button_font(FONT_CAPTION),
            command=self._open_releases_page,
        ).grid(row=0, column=1, sticky="e")

        steps_row = customtkinter.CTkFrame(nav, fg_color="transparent")
        steps_row.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        steps_row.grid_columnconfigure((0, 1), weight=1, uniform="steps")

        self.step_tiles: dict[str, customtkinter.CTkFrame] = {}
        for col, (key, label, number) in enumerate(STEPS):
            tile = self._create_step_tile(steps_row, key, label, number)
            tile.grid(
                row=0,
                column=col,
                sticky="ew",
                padx=(0 if col == 0 else 4, 0 if col == len(STEPS) - 1 else 4),
            )
            self.step_tiles[key] = tile

        self.doc_panel = customtkinter.CTkFrame(
            self.doc_sidebar,
            corner_radius=UI_RADIUS,
            fg_color=EMBED_BG_COLOR,
        )
        self.doc_panel.grid(
            row=1,
            column=0,
            sticky="nsew",
            padx=DOC_SIDEBAR_MARGIN,
            pady=(0, 8),
        )
        self.doc_panel.grid_columnconfigure(0, weight=1)
        self.doc_panel.grid_rowconfigure(3, weight=1)

        title_row = customtkinter.CTkFrame(self.doc_panel, fg_color="transparent")
        title_row.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 8))
        title_row.grid_columnconfigure(1, weight=1)

        customtkinter.CTkLabel(
            title_row,
            text="文档列表",
            font=customtkinter.CTkFont(size=FONT_TITLE, weight="bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=(0, 8))

        self.folder_label = customtkinter.CTkLabel(
            title_row,
            text="未选择文件夹",
            font=customtkinter.CTkFont(size=FONT_META),
            text_color="gray60",
            anchor="e",
            wraplength=max(80, (self._doc_sidebar_width() - 32) // 2),
            justify="right",
        )
        self.folder_label.grid(row=0, column=1, sticky="e")

        btn_row = customtkinter.CTkFrame(self.doc_panel, fg_color="transparent")
        btn_row.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 8))
        btn_row.grid_columnconfigure((0, 1), weight=1)

        self._make_accent_button(
            btn_row,
            text="选择文件夹…",
            height=SMALL_BTN_HEIGHT,
            corner_radius=UI_RADIUS,
            font=self._button_font(FONT_BUTTON),
            command=self._pick_folder,
        ).grid(row=0, column=0, padx=(0, 4), sticky="ew")

        customtkinter.CTkButton(
            btn_row,
            text="清空",
            height=SMALL_BTN_HEIGHT,
            corner_radius=UI_RADIUS,
            fg_color=(SECONDARY_BTN_HOVER[0], SECONDARY_BTN_FG[1]),
            hover_color=(SECONDARY_BTN_FG[0], SECONDARY_BTN_HOVER[1]),
            text_color=SECONDARY_BTN_TEXT,
            font=self._button_font(FONT_BUTTON),
            command=self._clear_extract,
        ).grid(row=0, column=1, padx=(4, 0), sticky="ew")

        self.doc_list_frame = customtkinter.CTkScrollableFrame(
            self.doc_panel,
            fg_color="transparent",
            corner_radius=0,
        )
        self.doc_list_frame.grid(row=3, column=0, sticky="nsew", padx=16, pady=(0, 14))
        self.doc_list_frame.grid_columnconfigure(0, weight=1)
        self._doc_list_scroll_after: str | None = None
        # Keep scrollregion / scrollbar in sync without fighting CTk every paint.
        self.doc_list_frame._parent_canvas.bind(
            "<Configure>", self._schedule_doc_list_scrollbar_sync, add="+"
        )
        self._bind_scrollable_mousewheel(self.doc_list_frame, self.doc_list_frame)
        self._bind_scrollable_mousewheel(
            self.doc_list_frame, self.doc_list_frame._parent_canvas
        )

        self.doc_list_empty = customtkinter.CTkLabel(
            self.doc_list_frame,
            text="选择文件夹后，\nPDF 将显示在这里",
            text_color="gray50",
            justify="center",
        )
        self.doc_list_empty.grid(row=0, column=0, pady=24)

        self.settings_button = customtkinter.CTkButton(
            self.doc_sidebar,
            text="设置",
            height=SETTINGS_BTN_HEIGHT,
            corner_radius=UI_RADIUS,
            fg_color=SECONDARY_BTN_FG,
            hover_color=SECONDARY_BTN_HOVER,
            text_color=SECONDARY_BTN_TEXT,
            font=self._button_font(FONT_BUTTON),
            border_width=0,
            command=self._on_open_settings,
        )
        self.settings_button.grid(
            row=2,
            column=0,
            sticky="ew",
            padx=DOC_SIDEBAR_MARGIN,
            pady=(0, DOC_SIDEBAR_MARGIN),
        )

    # ------------------------------------------------------------- main panel
    def _build_controls_panel(self):
        self.controls_panel = customtkinter.CTkFrame(
            self, corner_radius=0, fg_color=APP_BG_COLOR
        )
        self.controls_panel.grid(row=0, column=1, sticky="nsew", padx=0)
        self.controls_panel.grid_rowconfigure(0, weight=1)
        # Equal side spacers center content when panel is wider than MAIN_MAX_WIDTH.
        self.controls_panel.grid_columnconfigure(0, weight=1)
        self.controls_panel.grid_columnconfigure(1, weight=0)
        self.controls_panel.grid_columnconfigure(2, weight=1)

        self._controls_inner_width = MAIN_MIN_WIDTH
        self._controls_inner_height = -1
        self.controls_inner = customtkinter.CTkFrame(
            self.controls_panel,
            width=MAIN_MIN_WIDTH,
            corner_radius=0,
            fg_color="transparent",
        )
        # Fill panel height so short pages (extract) don't float centered
        # against the sidebar; field vcenter then uses the real content band.
        self.controls_inner.grid(row=0, column=1, sticky="nsew")
        self.controls_inner.grid_propagate(False)
        self.controls_inner.grid_rowconfigure(1, weight=1)
        self.controls_inner.grid_columnconfigure(0, weight=1)
        self.bind("<Configure>", self._on_app_configure_for_controls, add="+")
        self.after_idle(self._sync_controls_inner_width)

        # Match VinCert brand row: same top inset, height, and type size.
        self.controls_header_wrap = customtkinter.CTkFrame(
            self.controls_inner,
            height=BRAND_TILE_HEIGHT,
            fg_color="transparent",
        )
        self.controls_header_wrap.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=24,
            pady=(DOC_SIDEBAR_MARGIN, 8),
        )
        self.controls_header_wrap.grid_propagate(False)
        self.controls_header_wrap.grid_columnconfigure(0, weight=1)
        self.controls_header_wrap.grid_columnconfigure(1, weight=0)
        self.controls_header_wrap.grid_rowconfigure(0, weight=1)

        self.controls_header = customtkinter.CTkLabel(
            self.controls_header_wrap,
            text="",
            font=customtkinter.CTkFont(size=FONT_BRAND, weight="bold"),
            anchor="w",
        )
        self.controls_header.grid(row=0, column=0, sticky="w")

        self.controls_header_right = customtkinter.CTkFrame(
            self.controls_header_wrap,
            fg_color="transparent",
            width=1,
            height=1,
        )
        self.controls_header_right.grid(row=0, column=1, sticky="e")
        self.controls_header_right.grid_remove()
        self.controls_header_right.grid_columnconfigure(0, weight=0)
        self.controls_header_right.grid_columnconfigure(1, weight=0)

        self.automate_env_label = customtkinter.CTkLabel(
            self.controls_header_right,
            text="",
            font=customtkinter.CTkFont(size=FONT_META),
            text_color="gray60",
            anchor="e",
        )
        self.automate_env_label.grid(row=0, column=0, sticky="e", padx=(0, 6))

        self.automate_testing_mode_switch = customtkinter.CTkSwitch(
            self.controls_header_right,
            text="",
            width=36,
            height=24,
            switch_width=36,
            switch_height=18,
            font=customtkinter.CTkFont(size=FONT_META),
            command=self._on_automate_testing_mode_toggle,
        )
        if self._testing_mode:
            self.automate_testing_mode_switch.select()
        else:
            self.automate_testing_mode_switch.deselect()
        self.automate_testing_mode_switch.grid(row=0, column=1, sticky="e")

        self.controls_header_back_btn = customtkinter.CTkButton(
            self.controls_header_wrap,
            corner_radius=UI_RADIUS,
            text="返回设置",
            width=96,
            height=32,
            fg_color=SECONDARY_BTN_FG,
            hover_color=SECONDARY_BTN_HOVER,
            text_color=SECONDARY_BTN_TEXT,
            font=self._button_font(FONT_META),
            command=self._on_open_settings,
        )
        self.controls_header_back_btn.grid(row=0, column=1, sticky="e")
        self.controls_header_back_btn.grid_remove()

        self.controls_body = customtkinter.CTkFrame(self.controls_inner, fg_color="transparent")
        self.controls_body.grid(row=1, column=0, sticky="nsew", padx=24, pady=(0, 16))
        self.controls_body.grid_rowconfigure(0, weight=1)
        self.controls_body.grid_columnconfigure(0, weight=1)

        self.step_views: dict[str, customtkinter.CTkFrame] = {}
        for key, builder in [
            ("extract", self._build_extract_controls),
            ("automate", self._build_automate_controls),
            ("settings", self._build_settings_controls),
            ("parse_rules", self._build_parse_rules_controls),
            ("more_options", self._build_more_options_controls),
        ]:
            frame = customtkinter.CTkFrame(self.controls_body, fg_color="transparent")
            frame.grid(row=0, column=0, sticky="nsew")
            frame.grid_columnconfigure(0, weight=1)
            # Row weights are owned by each page builder / pinned layout.
            builder(frame)
            self.step_views[key] = frame

    def _on_app_configure_for_controls(self, event):
        if event.widget is not self:
            return
        self.after_idle(self._sync_controls_inner_width)
        self.after_idle(self._sync_active_page_vcenter)

    def _sync_controls_inner_width(self):
        if not hasattr(self, "controls_panel") or not hasattr(self, "controls_inner"):
            return
        panel_w_px = int(self.controls_panel.winfo_width())
        panel_h_px = int(self.controls_panel.winfo_height())
        if panel_w_px <= 1 or panel_h_px <= 1:
            return
        # winfo_* is scaled pixels; configure(width/height=) expects design units.
        scale = self._widget_scaling_factor()
        panel_w = panel_w_px / scale
        width = int(round(min(panel_w, float(MAIN_MAX_WIDTH))))
        height = int(round(panel_h_px / scale))
        if width < 1 or height < 1:
            return
        if (
            width == self._controls_inner_width
            and height == getattr(self, "_controls_inner_height", -1)
        ):
            return
        self._controls_inner_width = width
        self._controls_inner_height = height
        self.controls_inner.configure(width=width, height=height)
        self._update_content_wraplengths(max(120, width - 48))
        self.after_idle(self._sync_active_page_vcenter)

    def show_step(self, key: str):
        if self._autofill_busy and key != self._current_step:
            self.set_status("自动填写进行中，请先暂停或退出…")
            return
        if (
            key == "automate"
            and not self._automate_step_unlocked()
            and key != self._current_step
        ):
            self.set_status("请先批准至少一份证书")
            return
        self._current_step = key

        self._update_step_tiles(key)
        self._update_settings_button(key in SETTINGS_SUBPAGES)

        titles = {
            "extract": "提取核对",
            "automate": "自动化",
            "settings": "设置",
            "parse_rules": "解析规则",
            "more_options": "更多选项",
        }
        self.controls_header.configure(text=titles.get(key, key))
        self._update_controls_header_extras()

        for name, frame in self.step_views.items():
            if name == key:
                frame.tkraise()

        if key == "extract":
            self._load_approve_fields_for_current()
            self._schedule_active_page_vcenter(force=True)
        elif key == "automate":
            self._cancel_pending_quarantine()
            self._update_autofill_button()
            if not self._autofill_busy:
                path = self._selected_path or self._current_cert_path()
                self._load_automate_fields(path)
            self._schedule_active_page_vcenter(force=True)
        elif key == "settings":
            self._schedule_active_page_vcenter(force=True)
        elif key == "parse_rules":
            self._refresh_parse_rules_panel()
            self._schedule_active_page_vcenter(force=True)
        elif key == "more_options":
            self._schedule_active_page_vcenter(force=True)

    def _schedule_active_page_vcenter(self, *, force: bool = False):
        if force:
            self._invalidate_page_vcenters()
        self.after_idle(self._sync_active_page_vcenter)
        self.after(40, self._sync_active_page_vcenter)
        self.after(120, self._sync_active_page_vcenter)
        self.after(280, self._sync_active_page_vcenter)

    def _sync_visible_pinned_vcenters(self):
        """Backward-compatible alias: only the raised step is laid out."""
        self._sync_active_page_vcenter()

    def _sync_active_page_vcenter(self):
        """Center page content when it fits; top-align when scrolling / overflowing."""
        key = getattr(self, "_current_step", None)
        if not key or not hasattr(self, "step_views"):
            return
        frame = self.step_views.get(key)
        if frame is None:
            return
        self._sync_pinned_page_vcenter(frame)

    def _measure_scrollable_inner_height(self, scrollable) -> int:
        """Pixel height of scrollable inner content (window offset ignored)."""
        try:
            # reqheight is independent of canvas y-offset (avoids measure flicker).
            need = int(scrollable.winfo_reqheight())
            if need > 1:
                return need
        except Exception:  # noqa: BLE001
            pass
        try:
            canvas = scrollable._parent_canvas
            window_id = getattr(scrollable, "_create_window_id", None)
            if window_id is not None:
                canvas.coords(window_id, 0, 0)
            scrollable.update_idletasks()
            bbox = canvas.bbox("all")
            if bbox is None:
                return 1
            return max(1, int(bbox[3] - bbox[1]))
        except Exception:  # noqa: BLE001
            return 1

    def _settings_content_fits(self, avail: int, need: int) -> bool:
        return need <= avail + 1

    def _apply_settings_fits_chrome(
        self,
        content,
        *,
        center: bool,
        avail: int,
        need: int,
    ) -> None:
        """Hide scrollbar + lock scrollregion when settings content fits.

        CTkScrollableFrame's Configure handler sets scrollregion=bbox("all").
        After small→fullscreen, a leftover y-offset bbox makes the thumb sit in
        the middle of the content. Use a canvas-sized scrollregion instead.
        """
        canvas = content._parent_canvas
        scrollbar = content._scrollbar
        window_id = getattr(content, "_create_window_id", None)
        try:
            scrollbar.grid_remove()
        except Exception:  # noqa: BLE001
            pass
        y_off = max(0, (avail - need) // 2) if center else 0
        if window_id is not None:
            try:
                cur = canvas.coords(window_id)
            except Exception:  # noqa: BLE001
                cur = None
            want = [0.0, float(y_off)]
            if cur is None or len(cur) < 2 or abs(cur[0] - want[0]) > 0.5 or abs(cur[1] - want[1]) > 0.5:
                canvas.coords(window_id, 0, y_off)
        content.update_idletasks()
        cw = max(int(canvas.winfo_width()), 1)
        ch = max(int(canvas.winfo_height()), avail, 1)
        try:
            current_region = canvas.cget("scrollregion")
        except Exception:  # noqa: BLE001
            current_region = ""
        desired = f"0 0 {cw} {ch}"
        # Avoid redundant configure calls that retrigger CTk Configure.
        if str(current_region).replace(",", " ") != desired:
            canvas.configure(scrollregion=(0, 0, cw, ch))
        canvas.yview_moveto(0)

    def _repair_settings_scroll_after_resize(self, _event=None):
        """Idle repair for settings after window grow / fullscreen (no tight loop)."""
        if getattr(self, "_current_step", None) != "settings":
            return
        frame = getattr(self, "step_views", {}).get("settings")
        if frame is None:
            return
        meta = getattr(frame, "_vcenter_meta", None)
        if not meta or not meta.get("scrollable"):
            return
        content = meta.get("content")
        middle = meta.get("middle")
        if content is None or middle is None:
            return
        try:
            if not frame.winfo_ismapped():
                return
            avail = max(int(middle.winfo_height()), 1)
            if avail < 48:
                return
            need = self._measure_scrollable_inner_height(content)
            if not self._settings_content_fits(avail, need):
                return
            center = bool(self._content_centering)
            meta["centered"] = center
            self._apply_settings_fits_chrome(
                content, center=center, avail=avail, need=need
            )
        except Exception:  # noqa: BLE001
            pass

    def _schedule_settings_scroll_repair(self, _event=None):
        after_id = getattr(self, "_settings_scroll_repair_after", None)
        if after_id is not None:
            try:
                self.after_cancel(after_id)
            except Exception:  # noqa: BLE001
                pass
        # Debounced: catches CTk's scrollregion=bbox overwrite after fullscreen.
        self._settings_scroll_repair_after = self.after(
            80, self._repair_settings_scroll_after_resize
        )

    def _sync_settings_scroll_vcenter(self, parent: customtkinter.CTkFrame):
        """Settings: fill the page; center via canvas offset when no bar needed.

        Placing a CTkScrollableFrame collapses its canvas (invisible layout).
        Always pack-fill, then after size settles show/hide the scrollbar.
        """
        meta = getattr(parent, "_vcenter_meta", None)
        if not meta:
            return
        content = meta["content"]
        middle = meta["middle"]
        try:
            if not parent.winfo_ismapped():
                return
            scale = self._widget_scaling_factor()
            # Keep scrollable filling the middle band (never place it).
            if str(content.winfo_manager()) != "pack":
                content.place_forget()
                content.grid_forget()
            avail = max(int(middle.winfo_height()), 1)
            content.configure(height=max(1, int(round(avail / scale))))
            content.pack(side="top", fill="both", expand=True)
            parent.update_idletasks()

            avail = int(middle.winfo_height())
            if avail < 48:
                meta["centered"] = None
                meta.pop("layout_sig", None)
                return

            need = self._measure_scrollable_inner_height(content)
            fits = self._settings_content_fits(avail, need)
            center = bool(self._content_centering) and fits
            sig = (avail, need, center, bool(self._content_centering))
            if sig == meta.get("layout_sig"):
                # Same layout math, but small→fullscreen can leave a mapped bar
                # or a corrupted scrollregion — repair fits chrome only.
                if fits:
                    self._apply_settings_fits_chrome(
                        content, center=center, avail=avail, need=need
                    )
                return
            meta["layout_sig"] = sig
            meta["centered"] = center

            canvas = content._parent_canvas
            window_id = getattr(content, "_create_window_id", None)

            if fits:
                self._apply_settings_fits_chrome(
                    content, center=center, avail=avail, need=need
                )
                self._schedule_settings_scroll_repair()
            else:
                # Overflow: top-align; show scrollbar unless user hid bars.
                if window_id is not None:
                    canvas.coords(window_id, 0, 0)
                content.update_idletasks()
                bbox = canvas.bbox("all")
                if bbox is not None:
                    _x1, _y1, x2, y2 = bbox
                    canvas.configure(
                        scrollregion=(0, 0, max(0, x2), max(0, y2 - _y1))
                    )
                if getattr(self, "_hide_scrollbars", False):
                    self._hide_scrollable_bar(content)
                else:
                    content._create_grid()
                canvas.yview_moveto(0)
        except Exception:  # noqa: BLE001
            pass

    def _sync_pinned_page_vcenter(self, parent: customtkinter.CTkFrame):
        """Center field content between fixed header/footer when it fits."""
        meta = getattr(parent, "_vcenter_meta", None)
        if not meta:
            return
        if meta.get("scrollable"):
            self._sync_settings_scroll_vcenter(parent)
            return
        try:
            if not parent.winfo_ismapped():
                return
            parent.update_idletasks()
            middle = meta["middle"]
            content = meta["content"]
            top = meta["top"]
            bottom = meta["bottom"]
            avail = int(middle.winfo_height())
            need = max(int(content.winfo_reqheight()), 1)
            if avail < 48:
                meta["centered"] = None
                meta.pop("layout_sig", None)
                return
            center = bool(self._content_centering) and need <= avail + 1
            sig = (avail, need, center, bool(self._content_centering))
            if sig == meta.get("layout_sig"):
                return
            meta["layout_sig"] = sig
            meta["centered"] = center

            # Drop any prior manager so place/pack can take over cleanly.
            for spacer in (top, content, bottom):
                spacer.pack_forget()
                spacer.place_forget()
                spacer.grid_forget()

            if center:
                # Place in the middle band (header/footer stay docked outside).
                content.place(relx=0.5, rely=0.5, anchor="center", relwidth=1.0)
            else:
                content.pack(side="top", fill="both", expand=True)
        except Exception:  # noqa: BLE001
            pass

    def _invalidate_page_vcenters(self):
        for frame in getattr(self, "step_views", {}).values():
            meta = getattr(frame, "_vcenter_meta", None)
            if meta is not None:
                meta["centered"] = None
                meta.pop("layout_sig", None)

    def _make_pinned_footer_layout(
        self,
        parent: customtkinter.CTkFrame,
        *,
        with_header: bool = True,
        with_footer: bool = True,
    ) -> tuple[customtkinter.CTkFrame, customtkinter.CTkFrame, customtkinter.CTkFrame]:
        """Fixed optional header + centerable fields + optional pinned footer.

        Returns (header, content, footer). Only ``content`` is vertically
        centered when it fits; header/footer stay docked when present.
        """
        parent.grid_columnconfigure(0, weight=1)

        # CTkFrame defaults to 200×200; empty headers (e.g. 自动化 with
        # 功能拓展 off) would steal space and break content centering.
        header = customtkinter.CTkFrame(
            parent, fg_color="transparent", width=1, height=1
        )
        header.grid_columnconfigure(0, weight=1)

        middle = customtkinter.CTkFrame(
            parent, fg_color="transparent", width=1, height=1
        )
        middle.grid_columnconfigure(0, weight=1)

        top = customtkinter.CTkFrame(
            middle, fg_color="transparent", width=1, height=0
        )
        content = customtkinter.CTkFrame(
            middle, fg_color="transparent", width=1, height=1
        )
        content.grid_columnconfigure(0, weight=1)
        bottom = customtkinter.CTkFrame(
            middle, fg_color="transparent", width=1, height=0
        )

        footer = customtkinter.CTkFrame(
            parent, fg_color="transparent", width=1, height=1
        )
        footer.grid_columnconfigure(0, weight=1)

        if with_header and with_footer:
            parent.grid_rowconfigure(0, weight=0)
            parent.grid_rowconfigure(1, weight=1)
            parent.grid_rowconfigure(2, weight=0)
            header.grid(row=0, column=0, sticky="ew")
            middle.grid(row=1, column=0, sticky="nsew")
            footer.grid(row=2, column=0, sticky="ew", pady=(16, 0))
        elif with_header and not with_footer:
            parent.grid_rowconfigure(0, weight=0)
            parent.grid_rowconfigure(1, weight=1)
            header.grid(row=0, column=0, sticky="ew")
            middle.grid(row=1, column=0, sticky="nsew")
        elif not with_header and with_footer:
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_rowconfigure(1, weight=0)
            middle.grid(row=0, column=0, sticky="nsew")
            footer.grid(row=1, column=0, sticky="ew", pady=(16, 0))
        else:
            # Settings: full-page middle, same place/pack sync as step 2.
            parent.grid_rowconfigure(0, weight=1)
            middle.grid(row=0, column=0, sticky="nsew")

        # Default centered; sync may switch to top-align on overflow.
        content.place(relx=0.5, rely=0.5, anchor="center", relwidth=1.0)

        parent._vcenter_meta = {
            "top": top,
            "bottom": bottom,
            "middle": middle,
            "content": content,
            "header": header,
            "footer": footer,
            "centered": None,
            "with_header": with_header,
            "with_footer": with_footer,
        }
        parent.bind(
            "<Configure>",
            lambda _e, p=parent: self.after_idle(
                lambda: self._sync_pinned_page_vcenter(p)
            ),
            add="+",
        )
        middle.bind(
            "<Configure>",
            lambda _e, p=parent: self.after_idle(
                lambda: self._sync_pinned_page_vcenter(p)
            ),
            add="+",
        )
        self.after_idle(lambda p=parent: self._sync_pinned_page_vcenter(p))
        return header, content, footer

    def _build_settings_controls(self, parent: customtkinter.CTkFrame):
        # Same place/pack vcenter as step 2; scrollable content shows a bar only
        # after layout settles and content overflows.
        _header, old_content, _footer = self._make_pinned_footer_layout(
            parent, with_header=False, with_footer=False
        )
        meta = parent._vcenter_meta
        middle = meta["middle"]
        old_content.destroy()

        self.settings_scroll = customtkinter.CTkScrollableFrame(
            middle,
            fg_color="transparent",
            corner_radius=0,
        )
        self.settings_scroll.grid_columnconfigure(0, weight=1)
        meta["content"] = self.settings_scroll
        meta["scrollable"] = True
        # After small→fullscreen, CTk overwrites scrollregion with an offset bbox.
        # Debounced repair only — not a tight Configure loop.
        self.settings_scroll.bind(
            "<Configure>", self._schedule_settings_scroll_repair, add="+"
        )
        self.settings_scroll._parent_canvas.bind(
            "<Configure>", self._schedule_settings_scroll_repair, add="+"
        )

        content = self.settings_scroll
        self._bind_scrollable_mousewheel(self.settings_scroll, self.settings_scroll)
        self._bind_scrollable_mousewheel(
            self.settings_scroll, self.settings_scroll._parent_canvas
        )

        customtkinter.CTkLabel(
            content,
            text="EAMS 登录",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=0, column=0, sticky="ew", pady=(0, 8))

        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="填写账号密码并保存，自动填写时会代为登录。",
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=1, column=0, sticky="ew", pady=(0, 12))

        saved_user, saved_pass = load_credentials()

        cred_row = customtkinter.CTkFrame(content, fg_color="transparent")
        cred_row.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        cred_row.grid_columnconfigure((0, 1), weight=1, uniform="eams_creds")

        user_col = customtkinter.CTkFrame(cred_row, fg_color="transparent")
        user_col.grid(row=0, column=0, sticky="ew", padx=(0, 4))
        user_col.grid_columnconfigure(0, weight=1)
        customtkinter.CTkLabel(user_col, text="用户名", anchor="w").grid(
            row=0, column=0, sticky="ew", pady=(0, 4)
        )
        self.eams_username_entry = self._make_field_entry(
            user_col, placeholder="EAMS 用户名"
        )
        self.eams_username_entry.grid(row=1, column=0, sticky="ew")
        if saved_user:
            self.eams_username_entry.insert(0, saved_user)

        pass_col = customtkinter.CTkFrame(cred_row, fg_color="transparent")
        pass_col.grid(row=0, column=1, sticky="ew", padx=(4, 0))
        pass_col.grid_columnconfigure(0, weight=1)
        customtkinter.CTkLabel(pass_col, text="密码", anchor="w").grid(
            row=0, column=0, sticky="ew", pady=(0, 4)
        )
        self.eams_password_entry = self._make_field_entry(
            pass_col, placeholder="EAMS 密码", show="•"
        )
        self.eams_password_entry.grid(row=1, column=0, sticky="ew")
        if saved_pass:
            self.eams_password_entry.insert(0, saved_pass)

        self._make_accent_button(
            content,
            corner_radius=UI_RADIUS,
            text="保存登录信息",
            height=40,
            font=self._button_font(FONT_BUTTON),
            command=self._save_eams_login_info,
        ).grid(row=3, column=0, sticky="ew")

        customtkinter.CTkLabel(
            content,
            text="自动填写",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=4, column=0, sticky="ew", pady=(24, 8))

        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="每个自动化步骤之间的等待时间（秒）。默认 1 秒。",
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=5, column=0, sticky="ew", pady=(0, 8))

        delay_row = customtkinter.CTkFrame(content, fg_color="transparent")
        delay_row.grid(row=6, column=0, sticky="ew", pady=(0, 4))
        delay_row.grid_columnconfigure(0, weight=1)
        customtkinter.CTkLabel(delay_row, text="步骤间隔（秒）", anchor="w").grid(
            row=0, column=0, sticky="w", pady=(0, 4)
        )
        self.autofill_step_delay_entry = self._make_field_entry(
            delay_row, placeholder="1.0"
        )
        self.autofill_step_delay_entry.grid(row=1, column=0, sticky="ew", padx=(0, 8))
        self.autofill_step_delay_entry.insert(0, f"{self._autofill_step_delay_sec:g}")
        self._make_accent_button(
            delay_row,
            corner_radius=UI_RADIUS,
            text="保存",
            width=88,
            height=40,
            font=self._button_font(FONT_BUTTON),
            command=self._save_autofill_step_delay,
        ).grid(row=1, column=1, sticky="e")

        customtkinter.CTkLabel(
            content,
            text="界面与功能",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=7, column=0, sticky="ew", pady=(24, 8))

        switches_row = customtkinter.CTkFrame(content, fg_color="transparent")
        switches_row.grid(row=8, column=0, sticky="ew", pady=(0, 16))
        switches_row.grid_columnconfigure((0, 1), weight=1, uniform="settings_switches")

        self.ui_zoom_switch = customtkinter.CTkSwitch(
            switches_row,
            text="放大界面",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._ui_zoomed:
            self.ui_zoom_switch.select()
        else:
            self.ui_zoom_switch.deselect()
        self.ui_zoom_switch.configure(command=self._on_ui_zoom_toggle)
        self.ui_zoom_switch.grid(row=0, column=0, sticky="w", padx=(0, 4))

        self.buttons_bold_switch = customtkinter.CTkSwitch(
            switches_row,
            text="按钮加粗",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._buttons_bold:
            self.buttons_bold_switch.select()
        else:
            self.buttons_bold_switch.deselect()
        self.buttons_bold_switch.configure(command=self._on_buttons_bold_toggle)
        self.buttons_bold_switch.grid(row=0, column=1, sticky="w", padx=(4, 0))

        self.content_centering_switch = customtkinter.CTkSwitch(
            switches_row,
            text="内容居中",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._content_centering:
            self.content_centering_switch.select()
        else:
            self.content_centering_switch.deselect()
        self.content_centering_switch.configure(command=self._on_content_centering_toggle)
        self.content_centering_switch.grid(
            row=1, column=0, sticky="w", padx=(0, 4), pady=(10, 0)
        )

        self.status_dots_switch = customtkinter.CTkSwitch(
            switches_row,
            text="图标样式",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._status_dots:
            self.status_dots_switch.select()
        else:
            self.status_dots_switch.deselect()
        self.status_dots_switch.configure(command=self._on_status_dots_toggle)
        self.status_dots_switch.grid(
            row=1, column=1, sticky="w", padx=(4, 0), pady=(10, 0)
        )

        self.doc_list_scale_fonts_switch = customtkinter.CTkSwitch(
            switches_row,
            text="列表文字缩放",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._doc_list_scale_fonts:
            self.doc_list_scale_fonts_switch.select()
        else:
            self.doc_list_scale_fonts_switch.deselect()
        self.doc_list_scale_fonts_switch.configure(
            command=self._on_doc_list_scale_fonts_toggle
        )
        self.doc_list_scale_fonts_switch.grid(
            row=2, column=0, sticky="w", padx=(0, 4), pady=(10, 0)
        )

        self.hide_scrollbars_switch = customtkinter.CTkSwitch(
            switches_row,
            text="隐藏滚动条",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._hide_scrollbars:
            self.hide_scrollbars_switch.select()
        else:
            self.hide_scrollbars_switch.deselect()
        self.hide_scrollbars_switch.configure(command=self._on_hide_scrollbars_toggle)
        self.hide_scrollbars_switch.grid(
            row=2, column=1, sticky="w", padx=(4, 0), pady=(10, 0)
        )

        self.compact_min_window_switch = customtkinter.CTkSwitch(
            switches_row,
            text="更小窗口",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._compact_min_window:
            self.compact_min_window_switch.select()
        else:
            self.compact_min_window_switch.deselect()
        self.compact_min_window_switch.configure(
            command=self._on_compact_min_window_toggle
        )
        self.compact_min_window_switch.grid(
            row=3, column=0, sticky="w", padx=(0, 4), pady=(10, 0)
        )

        customtkinter.CTkButton(
            content,
            corner_radius=UI_RADIUS,
            text="更多选项…",
            height=40,
            fg_color=SECONDARY_BTN_FG,
            hover_color=SECONDARY_BTN_HOVER,
            text_color=SECONDARY_BTN_TEXT,
            font=self._button_font(FONT_BUTTON),
            command=self._open_more_options,
        ).grid(row=9, column=0, sticky="ew", pady=(16, 16))

        customtkinter.CTkLabel(
            content,
            text="功能拓展",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=12, column=0, sticky="ew", pady=(8, 8))

        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text=(
                    "将其他工具的能力并入当前流程。开启后，自动化页显示"
                    "「自定义 Excel 导入」，可选择一份已生成格式的 Excel，"
                    "按编号匹配当前文件夹中的 PDF 后导入 EAMS。"
                ),
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=13, column=0, sticky="ew", pady=(0, 6))
        self.feature_extensions_switch = customtkinter.CTkSwitch(
            content,
            text="启用功能拓展",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._feature_extensions:
            self.feature_extensions_switch.select()
        else:
            self.feature_extensions_switch.deselect()
        self.feature_extensions_switch.configure(
            command=self._on_feature_extensions_toggle
        )
        self.feature_extensions_switch.grid(row=14, column=0, sticky="w", pady=(0, 12))

        customtkinter.CTkLabel(
            content,
            text="测试模式",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=15, column=0, sticky="ew", pady=(8, 8))

        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="开启后使用 EAMS 测试环境（UAT）自动填写。",
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=16, column=0, sticky="ew", pady=(0, 6))
        self.testing_mode_switch = customtkinter.CTkSwitch(
            content,
            text="启用测试模式（EAMS UAT）",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._testing_mode:
            self.testing_mode_switch.select()
        else:
            self.testing_mode_switch.deselect()
        self.testing_mode_switch.configure(command=self._on_testing_mode_toggle)
        self.testing_mode_switch.grid(row=17, column=0, sticky="w", pady=(0, 12))

        customtkinter.CTkLabel(
            content,
            text="启动加载文件夹",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=18, column=0, sticky="ew", pady=(8, 8))

        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text=(
                    "开启后，应用启动时自动加载所选证书文件夹（正式 / 测试模式均可）。"
                ),
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=19, column=0, sticky="ew", pady=(0, 6))
        self.demo_folder_enabled_switch = customtkinter.CTkSwitch(
            content,
            text="启用启动加载",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._demo_folder_enabled:
            self.demo_folder_enabled_switch.select()
        else:
            self.demo_folder_enabled_switch.deselect()
        self.demo_folder_enabled_switch.configure(
            command=self._on_demo_folder_enabled_toggle
        )
        self.demo_folder_enabled_switch.grid(row=20, column=0, sticky="w", pady=(0, 8))

        self.demo_folder_label = self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text=self._demo_folder_display(),
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_META),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        )
        self.demo_folder_label.grid(row=21, column=0, sticky="ew", pady=(0, 8))

        demo_dir_row = customtkinter.CTkFrame(content, fg_color="transparent")
        demo_dir_row.grid(row=22, column=0, sticky="ew", pady=(0, 16))
        demo_dir_row.grid_columnconfigure((0, 1), weight=1)

        self._make_accent_button(
            demo_dir_row,
            corner_radius=UI_RADIUS,
            text="选择文件夹…",
            height=40,
            font=self._button_font(FONT_BUTTON),
            command=self._pick_demo_folder,
        ).grid(row=0, column=0, padx=(0, 4), sticky="ew")

        customtkinter.CTkButton(
            demo_dir_row,
            corner_radius=UI_RADIUS,
            text="清除",
            height=40,
            fg_color=SECONDARY_BTN_FG,
            hover_color=SECONDARY_BTN_HOVER,
            text_color=SECONDARY_BTN_TEXT,
            font=self._button_font(FONT_BUTTON),
            command=self._clear_demo_folder,
        ).grid(row=0, column=1, padx=(4, 0), sticky="ew")

        customtkinter.CTkLabel(
            content,
            text="解析规则",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=23, column=0, sticky="ew", pady=(8, 8))

        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text=(
                    "为解析字段增加标签别名（例如新证书上的「设备编号」）。"
                    "配置页会显示当前证书的原始解析文本，便于对照添加规则。"
                ),
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=24, column=0, sticky="ew", pady=(0, 12))

        self._make_accent_button(
            content,
            corner_radius=UI_RADIUS,
            text="配置解析规则…",
            height=40,
            font=self._button_font(FONT_BUTTON),
            command=self._open_parse_rules_config,
        ).grid(row=25, column=0, sticky="ew", pady=(0, 16))

        self._bind_scrollable_mousewheel(self.settings_scroll, self.settings_scroll)
        self._schedule_active_page_vcenter(force=True)

    def _build_more_options_controls(self, parent: customtkinter.CTkFrame):
        """Settings sub-page: OCR, split-screen, and PDF preview."""
        _header, old_content, _footer = self._make_pinned_footer_layout(
            parent, with_header=False, with_footer=False
        )
        meta = parent._vcenter_meta
        middle = meta["middle"]
        old_content.destroy()

        self.more_options_scroll = customtkinter.CTkScrollableFrame(
            middle,
            fg_color="transparent",
            corner_radius=0,
        )
        self.more_options_scroll.grid_columnconfigure(0, weight=1)
        meta["content"] = self.more_options_scroll
        meta["scrollable"] = True

        content = self.more_options_scroll
        self._bind_scrollable_mousewheel(self.more_options_scroll, self.more_options_scroll)
        self._bind_scrollable_mousewheel(
            self.more_options_scroll, self.more_options_scroll._parent_canvas
        )

        customtkinter.CTkLabel(
            content,
            text="OCR",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=0, column=0, sticky="ew", pady=(0, 8))
        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="开启后，提取页显示 OCR 按钮，用于扫描件或封面无嵌入文本的证书。",
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=1, column=0, sticky="ew", pady=(0, 6))
        self.ocr_enabled_switch = customtkinter.CTkSwitch(
            content,
            text="启用 OCR",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._ocr_enabled:
            self.ocr_enabled_switch.select()
        else:
            self.ocr_enabled_switch.deselect()
        self.ocr_enabled_switch.configure(command=self._on_ocr_enabled_toggle)
        self.ocr_enabled_switch.grid(row=2, column=0, sticky="w", pady=(0, 16))

        customtkinter.CTkLabel(
            content,
            text="窗口分屏",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=3, column=0, sticky="ew", pady=(8, 8))
        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="开启后，PDF 预览或自动填写时将 VinCert 与浏览器左右分屏。",
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=4, column=0, sticky="ew", pady=(0, 6))
        self.auto_window_snap_switch = customtkinter.CTkSwitch(
            content,
            text="自动窗口分屏",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._auto_window_snap:
            self.auto_window_snap_switch.select()
        else:
            self.auto_window_snap_switch.deselect()
        self.auto_window_snap_switch.configure(command=self._on_auto_window_snap_toggle)
        self.auto_window_snap_switch.grid(row=5, column=0, sticky="w", pady=(0, 16))

        customtkinter.CTkLabel(
            content,
            text="PDF 预览",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=6, column=0, sticky="ew", pady=(8, 8))
        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="开启后，选中证书时在右侧浏览器中预览 PDF。",
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=7, column=0, sticky="ew", pady=(0, 6))
        self.pdf_preview_enabled_switch = customtkinter.CTkSwitch(
            content,
            text="启用 PDF 预览",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        if self._pdf_preview_enabled:
            self.pdf_preview_enabled_switch.select()
        else:
            self.pdf_preview_enabled_switch.deselect()
        self.pdf_preview_enabled_switch.configure(
            command=self._on_pdf_preview_enabled_toggle
        )
        self.pdf_preview_enabled_switch.grid(row=8, column=0, sticky="w", pady=(0, 16))

        customtkinter.CTkLabel(
            content,
            text="有效期至",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=9, column=0, sticky="ew", pady=(8, 8))
        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="开启后使用已解析的「本次检测有效期至」字段；关闭则清空该输入，让系统自动判定。",
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=10, column=0, sticky="ew", pady=(0, 6))

        self.valid_till_use_extracted_switch = customtkinter.CTkSwitch(
            content,
            text="使用提取结果（开启时写入）",
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        # ON => use extracted due_date; OFF => clear and let system auto-determinate.
        if not self._valid_till_from_system:
            self.valid_till_use_extracted_switch.select()
        else:
            self.valid_till_use_extracted_switch.deselect()
        self.valid_till_use_extracted_switch.configure(
            command=self._on_valid_till_use_extracted_toggle
        )
        self.valid_till_use_extracted_switch.grid(
            row=11, column=0, sticky="w", pady=(0, 16)
        )

        self._bind_scrollable_mousewheel(self.more_options_scroll, self.more_options_scroll)
        self._schedule_active_page_vcenter(force=True)

    def _build_parse_rules_controls(self, parent: customtkinter.CTkFrame):
        """Settings sub-page: edit label aliases with raw parse text visible."""
        _header, old_content, _footer = self._make_pinned_footer_layout(
            parent, with_header=False, with_footer=False
        )
        meta = parent._vcenter_meta
        middle = meta["middle"]
        old_content.destroy()

        self.parse_rules_scroll = customtkinter.CTkScrollableFrame(
            middle,
            fg_color="transparent",
            corner_radius=0,
        )
        self.parse_rules_scroll.grid_columnconfigure(0, weight=1)
        meta["content"] = self.parse_rules_scroll
        meta["scrollable"] = True

        content = self.parse_rules_scroll
        self._bind_scrollable_mousewheel(self.parse_rules_scroll, self.parse_rules_scroll)
        self._bind_scrollable_mousewheel(
            self.parse_rules_scroll, self.parse_rules_scroll._parent_canvas
        )

        self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text=(
                    "对照下方原始解析文本，为字段添加证书上出现的标签别名。"
                    "保存后重新提取即可生效。"
                ),
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_BODY),
                text_color="gray60",
                wraplength=CONTENT_WRAP,
                justify="left",
            )
        ).grid(row=0, column=0, sticky="ew", pady=(0, 12))

        customtkinter.CTkLabel(
            content,
            text="原始解析文本",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=1, column=0, sticky="ew", pady=(0, 6))

        self.parse_rules_raw_source = customtkinter.CTkLabel(
            content,
            text="未选择已解析证书",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_CAPTION),
            text_color="gray60",
        )
        self.parse_rules_raw_source.grid(row=2, column=0, sticky="ew", pady=(0, 6))

        self.parse_rules_raw_text = customtkinter.CTkTextbox(
            content,
            height=180,
            corner_radius=UI_RADIUS,
            font=customtkinter.CTkFont(size=FONT_BODY),
            fg_color=FIELD_FG_COLOR,
            text_color=FIELD_TEXT_COLOR,
            wrap="word",
        )
        self.parse_rules_raw_text.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        self.parse_rules_raw_text.insert("1.0", "请先在左侧选择已提取的证书，以查看原始文本。")
        self.parse_rules_raw_text.configure(state="disabled")
        self._bind_scrollable_mousewheel(self.parse_rules_scroll, self.parse_rules_raw_text)

        customtkinter.CTkLabel(
            content,
            text="字段标签别名",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=4, column=0, sticky="ew", pady=(4, 6))

        field_row = customtkinter.CTkFrame(content, fg_color="transparent")
        field_row.grid(row=5, column=0, sticky="ew", pady=(0, 8))
        field_row.grid_columnconfigure(1, weight=1)
        customtkinter.CTkLabel(field_row, text="目标字段", anchor="w").grid(
            row=0, column=0, sticky="w", padx=(0, 8)
        )
        self.parse_rules_field_menu = customtkinter.CTkOptionMenu(
            field_row,
            values=[label for _key, label in PARSE_RULE_FIELDS],
            command=self._on_parse_rules_field_changed,
            font=customtkinter.CTkFont(size=FONT_BODY),
        )
        self.parse_rules_field_menu.grid(row=0, column=1, sticky="ew")
        self.parse_rules_field_menu.set(PARSE_RULE_FIELDS[0][1])

        self.parse_rules_builtin_label = customtkinter.CTkLabel(
            content,
            text="",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_CAPTION),
            text_color="gray60",
            wraplength=CONTENT_WRAP,
            justify="left",
        )
        self._track_content_wrap(self.parse_rules_builtin_label)
        self.parse_rules_builtin_label.grid(row=6, column=0, sticky="ew", pady=(0, 8))

        add_row = customtkinter.CTkFrame(content, fg_color="transparent")
        add_row.grid(row=7, column=0, sticky="ew", pady=(0, 8))
        add_row.grid_columnconfigure(0, weight=1)
        self.parse_rules_alias_entry = self._make_field_entry(
            add_row, placeholder="新标签别名，例如：设备编号"
        )
        self.parse_rules_alias_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self._make_accent_button(
            add_row,
            corner_radius=UI_RADIUS,
            text="添加",
            width=88,
            height=36,
            font=self._button_font(FONT_BUTTON),
            command=self._on_add_parse_rule_alias,
        ).grid(row=0, column=1, sticky="e")

        self.parse_rules_alias_list = customtkinter.CTkFrame(
            content, fg_color="transparent"
        )
        self.parse_rules_alias_list.grid(row=8, column=0, sticky="ew", pady=(0, 12))
        self.parse_rules_alias_list.grid_columnconfigure(0, weight=1)

        preview_row = customtkinter.CTkFrame(content, fg_color="transparent")
        preview_row.grid(row=9, column=0, sticky="ew", pady=(0, 8))
        preview_row.grid_columnconfigure(1, weight=1)
        self._make_accent_button(
            preview_row,
            corner_radius=UI_RADIUS,
            text="试解析当前原文",
            height=36,
            font=self._button_font(FONT_BUTTON),
            command=self._on_try_parse_rules,
        ).grid(row=0, column=0, sticky="w")

        customtkinter.CTkLabel(
            content,
            text="试解析结果",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
        ).grid(row=10, column=0, sticky="ew", pady=(4, 6))

        self.parse_rules_preview_text = customtkinter.CTkTextbox(
            content,
            height=140,
            corner_radius=UI_RADIUS,
            font=customtkinter.CTkFont(size=FONT_BODY),
            fg_color=FIELD_FG_COLOR,
            text_color=FIELD_TEXT_COLOR,
            wrap="word",
        )
        self.parse_rules_preview_text.grid(row=11, column=0, sticky="ew", pady=(0, 16))
        self.parse_rules_preview_text.insert("1.0", "点击「试解析当前原文」预览规则效果。")
        self.parse_rules_preview_text.configure(state="disabled")
        self._bind_scrollable_mousewheel(
            self.parse_rules_scroll, self.parse_rules_preview_text
        )

        self._refresh_parse_rules_alias_list()
        self._update_parse_rules_builtin_hint()
        self._schedule_active_page_vcenter(force=True)

    def _automate_env_label_text(self) -> str:
        if self._testing_mode:
            return "测试环境 · EAMS UAT"
        return "正式环境 · EAMS"

    def _sync_testing_mode_switches(self) -> None:
        enabled = bool(self._testing_mode)
        for name in ("testing_mode_switch", "automate_testing_mode_switch"):
            switch = getattr(self, name, None)
            if switch is None:
                continue
            try:
                if enabled:
                    switch.select()
                else:
                    switch.deselect()
            except Exception:  # noqa: BLE001
                pass

    def _update_controls_header_extras(self) -> None:
        """Right-side title chrome: automate env + switch, or settings back."""
        step = getattr(self, "_current_step", None)
        show_back = step in {"parse_rules", "more_options"}
        show_automate = step == "automate" and not show_back

        if hasattr(self, "controls_header_back_btn"):
            if show_back:
                self.controls_header_back_btn.grid()
            else:
                self.controls_header_back_btn.grid_remove()

        right = getattr(self, "controls_header_right", None)
        if right is None:
            return
        if show_automate:
            label = getattr(self, "automate_env_label", None)
            if label is not None:
                label.configure(text=self._automate_env_label_text())
            self._sync_testing_mode_switches()
            right.grid()
        else:
            right.grid_remove()

    def _update_controls_header_back(self, visible: bool) -> None:
        """Show/hide the title-bar「返回设置」control on nested settings pages."""
        del visible  # Kept for call-site compat; extras handler owns layout.
        self._update_controls_header_extras()

    def _open_parse_rules_config(self):
        if self._autofill_busy:
            self.set_status("自动填写进行中，请先暂停或退出…")
            return
        if self._current_step not in SETTINGS_SUBPAGES:
            self._step_before_settings = self._current_step
        self.show_step("parse_rules")

    def _open_more_options(self):
        if self._autofill_busy:
            self.set_status("自动填写进行中，请先暂停或退出…")
            return
        if self._current_step not in SETTINGS_SUBPAGES:
            self._step_before_settings = self._current_step
        self.show_step("more_options")

    def _open_releases_page(self):
        """Open the public GitHub Releases page in the default browser."""
        try:
            webbrowser.open(RELEASES_URL, new=2)
            self.set_status("已打开 GitHub Releases")
        except Exception as exc:  # noqa: BLE001
            self.set_status(f"无法打开链接：{exc}")
            self.show_toast(
                f"无法打开：{RELEASES_URL}\n{exc}",
                title="GitHub Releases",
            )

    def _parse_rule_field_key_from_label(self, label: str) -> str:
        for key, name in PARSE_RULE_FIELDS:
            if name == label:
                return key
        return PARSE_RULE_FIELDS[0][0]

    def _on_parse_rules_field_changed(self, label: str):
        self._parse_rules_field_key = self._parse_rule_field_key_from_label(label)
        self._refresh_parse_rules_alias_list()
        self._update_parse_rules_builtin_hint()

    def _update_parse_rules_builtin_hint(self):
        if not hasattr(self, "parse_rules_builtin_label"):
            return
        key = self._parse_rules_field_key
        builtin = LABEL_ALIASES.get(key, ())
        if builtin:
            text = "内置标签：" + "、".join(builtin)
        else:
            text = "该字段主要靠启发式识别；仍可添加标签别名辅助定位。"
        self.parse_rules_builtin_label.configure(text=text)

    def _refresh_parse_rules_alias_list(self):
        if not hasattr(self, "parse_rules_alias_list"):
            return
        for child in self.parse_rules_alias_list.winfo_children():
            child.destroy()
        key = self._parse_rules_field_key
        aliases = list(self._parse_rules.get(key, []))
        if not aliases:
            customtkinter.CTkLabel(
                self.parse_rules_alias_list,
                text="暂无自定义别名",
                anchor="w",
                text_color="gray60",
                font=customtkinter.CTkFont(size=FONT_CAPTION),
            ).grid(row=0, column=0, sticky="ew")
            return
        for i, alias in enumerate(aliases):
            row = customtkinter.CTkFrame(self.parse_rules_alias_list, fg_color="transparent")
            row.grid(row=i, column=0, sticky="ew", pady=(0, 4))
            row.grid_columnconfigure(0, weight=1)
            customtkinter.CTkLabel(row, text=alias, anchor="w").grid(
                row=0, column=0, sticky="ew", padx=(0, 8)
            )
            customtkinter.CTkButton(
                row,
                corner_radius=UI_RADIUS,
                text="删除",
                width=72,
                height=30,
                fg_color=SECONDARY_BTN_FG,
                hover_color=SECONDARY_BTN_HOVER,
                text_color=SECONDARY_BTN_TEXT,
                font=self._button_font(FONT_CAPTION),
                command=lambda a=alias: self._on_remove_parse_rule_alias(a),
            ).grid(row=0, column=1, sticky="e")

    def _persist_parse_rules(self):
        self._parse_rules = save_parse_rules(self._parse_rules)

    def _on_add_parse_rule_alias(self):
        alias = ""
        if hasattr(self, "parse_rules_alias_entry"):
            alias = self.parse_rules_alias_entry.get().strip()
        if not alias:
            self.set_status("请先输入要添加的标签别名")
            return
        key = self._parse_rules_field_key
        current = list(self._parse_rules.get(key, []))
        if alias in current:
            self.set_status(f"别名已存在：{alias}")
            return
        builtin = LABEL_ALIASES.get(key, ())
        if alias in builtin:
            self.set_status(f"「{alias}」已是内置标签，无需添加")
            return
        current.append(alias)
        self._parse_rules[key] = current
        self._persist_parse_rules()
        if hasattr(self, "parse_rules_alias_entry"):
            self.parse_rules_alias_entry.delete(0, "end")
        self._refresh_parse_rules_alias_list()
        self.set_status(f"已添加解析别名：{PARSE_RULE_FIELD_LABELS.get(key, key)} ← {alias}")
        self.show_success_toast(
            f"已添加「{alias}」→ {PARSE_RULE_FIELD_LABELS.get(key, key)}",
            title="解析规则",
        )

    def _on_remove_parse_rule_alias(self, alias: str):
        key = self._parse_rules_field_key
        current = [a for a in self._parse_rules.get(key, []) if a != alias]
        if current:
            self._parse_rules[key] = current
        else:
            self._parse_rules.pop(key, None)
        self._persist_parse_rules()
        self._refresh_parse_rules_alias_list()
        self.set_status(f"已删除别名：{alias}")

    def _current_raw_parse_text(self) -> str:
        path = self._selected_path
        if not path:
            return ""
        result = self._parse_results.get(path)
        if result is None:
            return ""
        raw = (result.raw_text or "").strip()
        if raw:
            return raw
        if result.lines:
            return "\n".join(result.lines)
        return ""

    def _set_parse_rules_textbox(self, widget, text: str) -> None:
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", text)
        widget.configure(state="disabled")

    def _refresh_parse_rules_panel(self):
        if not hasattr(self, "parse_rules_raw_text"):
            return
        path = self._selected_path
        raw = self._current_raw_parse_text()
        if path and raw:
            source = f"来源：{Path(path).name}"
            body = raw
        elif path and path in self._parse_results:
            source = f"来源：{Path(path).name}（无原始文本）"
            body = "（当前证书没有可用的原始解析文本）"
        elif path:
            source = f"来源：{Path(path).name}（尚未提取）"
            body = "请先在「提取核对」中解析该证书，再回到此处查看原文。"
        else:
            source = "未选择已解析证书"
            body = "请先在左侧选择已提取的证书，以查看原始文本。"
        if hasattr(self, "parse_rules_raw_source"):
            self.parse_rules_raw_source.configure(text=source)
        self._set_parse_rules_textbox(self.parse_rules_raw_text, body)
        self._refresh_parse_rules_alias_list()
        self._update_parse_rules_builtin_hint()

    def _on_try_parse_rules(self):
        raw = self._current_raw_parse_text()
        if not raw.strip():
            self.set_status("没有可用的原始解析文本")
            self.show_toast(
                "请先选择并提取一份证书，以便对照原文试解析。",
                title="解析规则",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return
        try:
            fields = parse_fields(raw, extra_label_aliases=self._parse_rules)
        except Exception as exc:  # noqa: BLE001
            self.set_status(f"试解析失败：{exc}")
            return
        lines = [
            f"{label}：{getattr(fields, key, '') or '—'}"
            for key, label in PARSE_RULE_FIELDS
        ]
        preview = "\n".join(lines)
        self._set_parse_rules_textbox(self.parse_rules_preview_text, preview)
        self.set_status("试解析完成（未写回证书结果；重新提取后规则才会应用到列表）")

    def _extra_label_aliases(self) -> dict[str, list[str]]:
        return dict(self._parse_rules or {})

    def _demo_folder_display(self) -> str:
        return self._demo_folder or "未选择启动加载文件夹"

    def _update_demo_folder_label(self):
        if hasattr(self, "demo_folder_label"):
            self.demo_folder_label.configure(text=self._demo_folder_display())

    def _on_demo_folder_enabled_toggle(self):
        self._apply_demo_folder_enabled(bool(self.demo_folder_enabled_switch.get()))

    def _apply_demo_folder_enabled(self, enabled: bool):
        self._demo_folder_enabled = enabled
        save_demo_folder_enabled(enabled)
        if enabled:
            if self._demo_folder and Path(self._demo_folder).is_dir():
                self.set_status(f"启动加载已开启 · {self._demo_folder}")
                self.show_success_toast(
                    f"启动时将自动加载：\n{self._demo_folder}",
                    title="启动加载文件夹",
                )
            else:
                self.set_status("启动加载已开启 · 请先选择文件夹")
                self.show_toast(
                    "已开启启动加载，请先选择证书文件夹。",
                    title="启动加载文件夹",
                    duration_ms=TOAST_SUCCESS_MS,
                )
        else:
            self.set_status("启动加载已关闭")
            self.show_success_toast("启动时不再自动加载文件夹。", title="启动加载文件夹")

    def _apply_testing_mode(self, enabled: bool) -> None:
        self._testing_mode = bool(enabled)
        save_testing_mode(self._testing_mode)
        self._sync_testing_mode_switches()
        self._update_controls_header_extras()
        if self._testing_mode:
            self.set_status("测试模式已开启 · EAMS 测试环境（auth.masuat.apps.ocpuat）")
            self.show_success_toast(
                "已切换到 EAMS 测试环境。",
                title="测试模式",
            )
        else:
            self.set_status("测试模式已关闭 · 使用 EAMS 正式环境")
            self.show_success_toast(
                "已切换到 EAMS 正式环境。",
                title="测试模式",
            )

    def _on_testing_mode_toggle(self):
        self._apply_testing_mode(bool(self.testing_mode_switch.get()))

    def _on_automate_testing_mode_toggle(self):
        self._apply_testing_mode(bool(self.automate_testing_mode_switch.get()))

    def _pick_demo_folder(self):
        initial = self._demo_folder if self._demo_folder and Path(self._demo_folder).is_dir() else None
        path = filedialog.askdirectory(
            parent=self,
            title="选择启动加载证书文件夹",
            initialdir=initial,
        )
        if not path:
            self.set_status("未更改启动加载文件夹")
            return
        self._demo_folder = save_demo_folder(path)
        self._update_demo_folder_label()
        msg = f"启动加载文件夹已更新：{self._demo_folder}"
        self.set_status(msg)
        self.show_success_toast(msg, title="启动加载文件夹")

    def _clear_demo_folder(self):
        """Ask for confirmation before clearing the startup folder path."""
        current = (self._demo_folder or "").strip()
        if not current:
            self.set_status("未配置启动加载文件夹")
            return
        self.show_toast(
            f"将清除启动加载文件夹路径：\n{current}\n"
            "下次启动将不再自动加载该文件夹。",
            title="清除启动加载文件夹",
            action_text="确认清除",
            undo_text="取消",
            on_complete=self._confirm_clear_demo_folder,
            complete_on_timeout=False,
            duration_ms=TOAST_DEFAULT_MS,
            style="danger",
        )
        self.set_status("确认清除启动加载文件夹？")

    def _confirm_clear_demo_folder(self):
        self._demo_folder = save_demo_folder("")
        self._update_demo_folder_label()
        self.set_status("已清除启动加载文件夹")
        self.show_success_toast("已清除启动加载文件夹。", title="启动加载文件夹")

    def _maybe_autoload_demo_folder(self):
        """On launch, optionally load the configured certificate folder."""
        if not getattr(self, "_demo_folder_enabled", False):
            return
        folder = (self._demo_folder or "").strip()
        if not folder:
            self.set_status("启动加载已开启，但未配置文件夹")
            self.show_toast(
                "启动加载已开启，但未配置证书文件夹。",
                title="启动加载文件夹",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return
        path = Path(folder)
        if not path.is_dir():
            self.set_status(f"启动加载文件夹不存在：{folder}")
            self.show_toast(
                f"启动加载文件夹不存在：\n{folder}",
                title="启动加载文件夹",
            )
            return
        self.set_status("正在加载启动文件夹…")
        self._load_folder(str(path))

    def _on_ui_zoom_toggle(self):
        self._apply_ui_zoom(bool(self.ui_zoom_switch.get()))

    def _on_buttons_bold_toggle(self):
        self._apply_buttons_bold(bool(self.buttons_bold_switch.get()))

    def _on_content_centering_toggle(self):
        self._apply_content_centering(bool(self.content_centering_switch.get()))

    def _on_status_dots_toggle(self):
        self._apply_status_dots(bool(self.status_dots_switch.get()))

    def _on_doc_list_scale_fonts_toggle(self):
        self._apply_doc_list_scale_fonts(bool(self.doc_list_scale_fonts_switch.get()))

    def _on_hide_scrollbars_toggle(self):
        self._apply_hide_scrollbars(bool(self.hide_scrollbars_switch.get()))

    def _on_compact_min_window_toggle(self):
        self._apply_compact_min_window(bool(self.compact_min_window_switch.get()))

    def _on_auto_window_snap_toggle(self):
        self._apply_auto_window_snap(bool(self.auto_window_snap_switch.get()))

    def _on_pdf_preview_enabled_toggle(self):
        self._apply_pdf_preview_enabled(bool(self.pdf_preview_enabled_switch.get()))

    def _on_valid_till_use_extracted_toggle(self):
        # UI switch means "use extracted due_date when ON".
        use_extracted = bool(self.valid_till_use_extracted_switch.get())
        self._apply_valid_till_from_system(not use_extracted)

    def _on_feature_extensions_toggle(self):
        self._apply_feature_extensions(bool(self.feature_extensions_switch.get()))

    def _apply_content_centering(self, enabled: bool):
        self._content_centering = enabled
        save_content_centering(enabled)
        self._schedule_active_page_vcenter(force=True)
        self.set_status("已开启内容居中" if enabled else "已关闭内容居中")

    def _apply_status_dots(self, enabled: bool):
        self._status_dots = enabled
        save_status_dots(enabled)
        self._refresh_doc_list_marks()
        self.set_status("已使用圆点图标" if enabled else "已使用表情图标")

    def _apply_doc_list_scale_fonts(self, enabled: bool):
        self._doc_list_scale_fonts = enabled
        save_doc_list_scale_fonts(enabled)
        self._refresh_doc_list_fonts()
        self.set_status(
            "已开启列表文字缩放" if enabled else "已关闭列表文字缩放（固定字号）"
        )

    def _apply_hide_scrollbars(self, enabled: bool):
        self._hide_scrollbars = enabled
        save_hide_scrollbars(enabled)
        self._refresh_scrollbar_visibility()
        self.set_status(
            "已隐藏滚动条（仍可用滚轮滑动）"
            if enabled
            else "已显示滚动条"
        )

    def _apply_compact_min_window(self, enabled: bool):
        self._compact_min_window = enabled
        save_compact_min_window(enabled)
        self._apply_layout_column_minsizes()
        self._apply_min_window_size()
        self._apply_doc_sidebar_width()
        self._controls_inner_width = -1
        self._sync_controls_inner_width()
        self._schedule_active_page_vcenter(force=True)
        self.set_status("已允许更小窗口" if enabled else "已恢复默认最小窗口")

    def _apply_auto_window_snap(self, enabled: bool):
        self._auto_window_snap = enabled
        save_auto_window_snap(enabled)
        if enabled:
            self._sync_window_layout_to_browser()
            if self._pdf_preview_enabled:
                self._sync_pdf_preview()
            self.set_status("已开启自动窗口分屏")
        else:
            self._pdf_preview_layout_active = False
            self.set_status("已关闭自动窗口分屏")

    def _apply_pdf_preview_enabled(self, enabled: bool):
        self._pdf_preview_enabled = enabled
        save_pdf_preview_enabled(enabled)
        if enabled:
            self._sync_pdf_preview()
            self.set_status("已开启 PDF 预览")
        else:
            if self._pdf_preview.has_pdf:
                self._pdf_preview.close()
            self.set_status("已关闭 PDF 预览")

    def _apply_valid_till_from_system(self, enabled: bool):
        self._valid_till_from_system = enabled
        save_valid_till_from_system(enabled)
        self.set_status(
            "有效期至：已使用系统自动判定"
            if enabled
            else "有效期至：将使用提取结果"
        )

    def _apply_feature_extensions(self, enabled: bool):
        self._feature_extensions = enabled
        save_feature_extensions(enabled)
        self._apply_feature_extension_theme()
        self._update_custom_dir_autofill_button()
        self._update_autofill_button()
        self.set_status("已开启功能拓展" if enabled else "已关闭功能拓展")

    def _update_custom_dir_autofill_button(self):
        """Show the purple custom Excel button at top when 功能拓展 is on."""
        btn = getattr(self, "custom_dir_autofill_button", None)
        header = getattr(self, "automate_page_header", None)
        if btn is None:
            return
        show = (
            bool(getattr(self, "_feature_extensions", False))
            and not self._autofill_busy
        )
        if show:
            try:
                btn.configure(
                    state="normal",
                    fg_color=CUSTOM_AUTOFILL_BTN_FG,
                    hover_color=CUSTOM_AUTOFILL_BTN_HOVER,
                    text_color=CUSTOM_AUTOFILL_BTN_TEXT,
                )
                btn.grid(row=0, column=0, sticky="ew", pady=(0, 12))
            except Exception:  # noqa: BLE001
                pass
            if header is not None:
                try:
                    header.grid(row=0, column=0, sticky="ew")
                except Exception:  # noqa: BLE001
                    pass
        else:
            try:
                btn.grid_remove()
            except Exception:  # noqa: BLE001
                pass
            if header is not None:
                try:
                    header.grid_remove()
                except Exception:  # noqa: BLE001
                    pass
        self._schedule_active_page_vcenter(force=True)

    def _iter_app_scrollables(self):
        for name in (
            "doc_list_frame",
            "settings_scroll",
            "parse_rules_scroll",
            "more_options_scroll",
        ):
            frame = getattr(self, name, None)
            if frame is not None:
                yield frame

    def _hide_scrollable_bar(self, scrollable) -> None:
        try:
            scrollable._scrollbar.grid_remove()
        except Exception:  # noqa: BLE001
            pass

    def _refresh_scrollbar_visibility(self) -> None:
        """Re-apply show/hide for all known scrollable frames."""
        if self._hide_scrollbars:
            for frame in self._iter_app_scrollables():
                self._hide_scrollable_bar(frame)
        self._schedule_doc_list_scrollbar_sync()
        self._schedule_active_page_vcenter(force=True)

    def _apply_buttons_bold(self, bold: bool):
        self._buttons_bold = bold
        save_buttons_bold(bold)
        weight = "bold" if bold else "normal"

        def walk(widget):
            if isinstance(widget, customtkinter.CTkButton):
                try:
                    font = widget.cget("font")
                    size = FONT_BUTTON
                    if isinstance(font, customtkinter.CTkFont):
                        size = int(font.cget("size"))
                    widget.configure(
                        font=customtkinter.CTkFont(size=size, weight=weight)
                    )
                    self._fix_ctk_button_text_vcenter(widget)
                except Exception:  # noqa: BLE001
                    pass
            for child in widget.winfo_children():
                walk(child)

        walk(self)
        self.set_status("已开启按钮加粗" if bold else "已关闭按钮加粗")

    def _on_ocr_enabled_toggle(self):
        enabled = bool(self.ocr_enabled_switch.get())
        if self._extract_busy and enabled != self._ocr_enabled:
            # Revert while OCR is running.
            if self._ocr_enabled:
                self.ocr_enabled_switch.select()
            else:
                self.ocr_enabled_switch.deselect()
            self.set_status("正在处理中，请稍后再切换 OCR")
            self.show_toast(
                "正在处理中，请稍后再切换 OCR。",
                title="OCR 提取",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return
        self._apply_ocr_enabled(enabled)

    def _apply_ocr_enabled(self, enabled: bool):
        self._ocr_enabled = enabled
        save_ocr_enabled(enabled)
        self._update_extract_ocr_ui()
        self._rebuild_doc_list()
        self.set_status("已启用 OCR" if enabled else "已关闭 OCR")

    def _update_extract_ocr_ui(self):
        """Show/hide OCR progress + OCR button (移出失败证书 is always on the form)."""
        if not hasattr(self, "ocr_extract_button"):
            return
        if self._ocr_enabled:
            if hasattr(self, "ocr_progress"):
                self.ocr_progress.grid(row=0, column=0, sticky="ew", pady=(0, 8))
            if hasattr(self, "ocr_progress_label"):
                self.ocr_progress_label.grid(row=1, column=0, sticky="ew", pady=(0, 8))
            self.ocr_extract_button.grid(row=2, column=0, sticky="ew", pady=(0, 10))
            self.ocr_extract_button.configure(
                text="OCR提取",
                fg_color=SUCCESS_BTN_FG,
                hover_color=SUCCESS_BTN_HOVER,
                text_color=SUCCESS_BTN_TEXT,
                command=self._on_ocr_extract,
            )
        else:
            if hasattr(self, "ocr_progress"):
                self.ocr_progress.grid_remove()
            if hasattr(self, "ocr_progress_label"):
                self.ocr_progress_label.grid_remove()
            self.ocr_extract_button.grid_remove()

    def _apply_ui_zoom(self, zoomed: bool):
        self._ui_zoomed = zoomed
        apply_ui_scale(zoomed)
        save_ui_zoomed(zoomed)
        # Zoom-out (esp. fullscreen) can leave an unfilled black edge unless we
        # re-assert fixed widths and force the window to relayout.
        self._refresh_layout_after_scale()
        self.set_status("已切换为放大界面" if zoomed else "已切换为标准界面")

    def _refresh_layout_after_scale(self):
        """Recompute layout after widget/window scaling changes."""
        self._controls_inner_width = -1
        self._controls_inner_height = -1
        self._apply_layout_column_minsizes()
        self._apply_min_window_size()
        self._invalidate_page_vcenters()

        if hasattr(self, "doc_sidebar"):
            # Re-apply design width so scaled pixel size updates on zoom-out.
            self._apply_doc_sidebar_width()
        if hasattr(self, "controls_panel"):
            self.controls_panel.grid(row=0, column=1, sticky="nsew", padx=0)

        self.update_idletasks()
        self._nudge_window_geometry_after_scale()
        self.update_idletasks()
        self._sync_controls_inner_width()
        # Second pass after CTk finishes propagating the new scale.
        self.after_idle(self._sync_controls_inner_width)
        self.after(40, self._sync_controls_inner_width)
        self.after(40, self._restyle_primary_action_buttons)
        if hasattr(self, "doc_list_frame"):
            self.after(40, self._schedule_doc_list_scrollbar_sync)
            self.after(40, self._refresh_doc_list_fonts)
        self._schedule_active_page_vcenter(force=True)

    def _nudge_window_geometry_after_scale(self):
        """Force Tk/CTk to refill the window after scale-down (clears black bars)."""
        try:
            if self._auto_window_snap and self._pdf_preview_layout_active:
                self._snap_app_left_half()
                return
            if self._auto_window_snap and sys.platform == "win32":
                self._apply_window_fullscreen()
                return
            width = int(self.winfo_width())
            height = int(self.winfo_height())
            if width <= 1 or height <= 1:
                return
            self.geometry(f"{width}x{height + 1}")
            self.update_idletasks()
            self.geometry(f"{width}x{height}")
        except Exception:  # noqa: BLE001
            pass

    def _eams_credentials(self) -> tuple[str, str]:
        username = ""
        password = ""
        if hasattr(self, "eams_username_entry"):
            username = self.eams_username_entry.get().strip()
        if hasattr(self, "eams_password_entry"):
            password = self.eams_password_entry.get()
        return username, password

    def set_status(self, message: str):
        print(f"[VinCert] {message}")

    def _toast_target_width(self) -> int:
        """Fixed design-unit width for toasts."""
        return TOAST_WIDTH

    def _ensure_toast_host(self) -> customtkinter.CTkFrame:
        if self._toast_host is None:
            host = customtkinter.CTkFrame(self, fg_color="transparent")
            self._toast_host = host
            self._place_toast_host()
        return self._toast_host

    def _toast_host_x(self) -> int:
        """Right inset for the toast stack; shift left when autofill log is open."""
        if self._autofill_log_frame is not None or self._autofill_log_bubble is not None:
            if self._autofill_log_collapsed:
                return -(
                    AUTOFILL_LOG_PAD
                    + AUTOFILL_LOG_BUBBLE
                    + AUTOFILL_LOG_BUBBLE_TOAST_GAP
                    + TOAST_PAD
                )
            return -(AUTOFILL_LOG_PAD + AUTOFILL_LOG_WIDTH + TOAST_PAD)
        return -TOAST_PAD

    def _place_toast_host(self):
        host = self._toast_host
        if host is None:
            return
        # Compact toasts anchor to the top-right (shift left when terminal is open).
        host.place(
            relx=1.0,
            rely=0.0,
            x=self._toast_host_x(),
            y=TOAST_PAD,
            anchor="ne",
        )
        try:
            host.lift()
        except Exception:  # noqa: BLE001
            pass

    def _relayout_toast_stack(self):
        host = self._toast_host
        if host is None:
            return
        for entry in self._toasts:
            try:
                entry["frame"].pack_forget()
            except Exception:  # noqa: BLE001
                pass
        for i, entry in enumerate(self._toasts):
            gap = TOAST_STACK_GAP if i < len(self._toasts) - 1 else 0
            try:
                entry["frame"].pack(side="top", anchor="e", pady=(0, gap))
            except Exception:  # noqa: BLE001
                pass
        self._place_toast_host()

    def _autofill_log_height(self) -> int:
        """Terminal occupies roughly the top half of the window."""
        scale = self._widget_scaling_factor()
        try:
            win_h = int(round(self.winfo_height() / scale))
        except Exception:  # noqa: BLE001
            win_h = 720
        half = max(1, (win_h - 2 * AUTOFILL_LOG_PAD) // 2)
        return max(200, half)

    def _sync_autofill_log_geometry(self, _event=None):
        frame = self._autofill_log_frame
        bubble = self._autofill_log_bubble
        try:
            if self._autofill_log_collapsed:
                if frame is not None:
                    frame.place_forget()
                if bubble is not None:
                    bubble.place(
                        relx=1.0,
                        rely=0.0,
                        x=-AUTOFILL_LOG_PAD,
                        y=AUTOFILL_LOG_PAD,
                        anchor="ne",
                    )
                    bubble.lift()
            elif frame is not None:
                if bubble is not None:
                    bubble.place_forget()
                frame.configure(
                    width=AUTOFILL_LOG_WIDTH,
                    height=self._autofill_log_height(),
                    corner_radius=TOAST_RADIUS,
                )
                frame.place(
                    relx=1.0,
                    rely=0.0,
                    x=-AUTOFILL_LOG_PAD,
                    y=AUTOFILL_LOG_PAD,
                    anchor="ne",
                )
                frame.lift()
        except Exception:  # noqa: BLE001
            pass
        self._place_toast_host()

    def _ensure_autofill_log_bubble(self) -> customtkinter.CTkButton:
        """Standalone circle — nested CTk widgets leave a square canvas halo."""
        bubble = self._autofill_log_bubble
        accent = getattr(self, "_autofill_log_accent", SUCCESS_BTN_FG)
        size = AUTOFILL_LOG_BUBBLE
        if bubble is None:
            bubble = customtkinter.CTkButton(
                self,
                text=">_",
                width=size,
                height=size,
                corner_radius=size // 2,
                fg_color=TOAST_BG,
                hover_color=TILE_BG_HOVER,
                bg_color="transparent",
                border_width=TOAST_BORDER_WIDTH,
                border_color=accent,
                text_color=SUCCESS_BTN_HOVER,
                font=customtkinter.CTkFont(
                    family="Menlo", size=FONT_META, weight="bold"
                ),
                command=self.expand_autofill_log,
            )
            self._autofill_log_bubble = bubble
        else:
            try:
                bubble.configure(border_color=accent, text=">_")
            except Exception:  # noqa: BLE001
                pass
        return bubble

    def collapse_autofill_log(self):
        """Shrink the autofill terminal into a top-right circle."""
        if self._autofill_log_frame is None or self._autofill_log_collapsed:
            return
        self._autofill_log_collapsed = True
        self._ensure_autofill_log_bubble()
        self._sync_autofill_log_geometry()

    def expand_autofill_log(self):
        """Restore the autofill terminal from the circle bubble."""
        if self._autofill_log_frame is None or not self._autofill_log_collapsed:
            return
        self._autofill_log_collapsed = False
        bubble = self._autofill_log_bubble
        if bubble is not None:
            try:
                bubble.configure(text=">_")
            except Exception:  # noqa: BLE001
                pass
        self._sync_autofill_log_geometry()

    def open_autofill_log(self, *, title: str = "自动填写"):
        """Show the top-half autofill terminal panel (clears any prior log)."""
        self.close_autofill_log()
        accent = SUCCESS_BTN_FG
        self._autofill_log_accent = accent
        h = self._autofill_log_height()
        frame = customtkinter.CTkFrame(
            self,
            width=AUTOFILL_LOG_WIDTH,
            height=h,
            corner_radius=TOAST_RADIUS,
            fg_color=TOAST_BG,
            border_width=TOAST_BORDER_WIDTH,
            border_color=accent,
        )
        frame.grid_propagate(False)
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)

        header = customtkinter.CTkFrame(frame, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=14, pady=(12, 6))
        header.grid_columnconfigure(0, weight=1)

        customtkinter.CTkLabel(
            header,
            text=title,
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_TITLE, weight="bold"),
            text_color=TOAST_TITLE_COLOR,
        ).grid(row=0, column=0, sticky="ew")

        status = customtkinter.CTkLabel(
            header,
            text="进行中",
            anchor="e",
            font=customtkinter.CTkFont(size=FONT_META, weight="bold"),
            text_color=SUCCESS_BTN_HOVER,
        )
        status.grid(row=0, column=1, sticky="e", padx=(8, 0))

        customtkinter.CTkButton(
            header,
            text="−",
            width=28,
            height=28,
            corner_radius=UI_RADIUS,
            fg_color="transparent",
            hover_color=TILE_BG_HOVER,
            border_width=0,
            text_color=TOAST_TITLE_COLOR,
            font=customtkinter.CTkFont(size=FONT_TITLE, weight="bold"),
            command=self.collapse_autofill_log,
        ).grid(row=0, column=2, sticky="e", padx=(6, 0))

        text = customtkinter.CTkTextbox(
            frame,
            width=AUTOFILL_LOG_WIDTH - 28,
            corner_radius=UI_RADIUS,
            fg_color=("#f4f6f8", "#121212"),
            text_color=TOAST_MESSAGE_COLOR,
            border_width=0,
            font=customtkinter.CTkFont(family="Menlo", size=FONT_META),
            activate_scrollbars=False,
            wrap="word",
        )
        text.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        text.configure(state="disabled")
        # Keep view pinned to latest lines; no manual scroll.
        text.bind("<MouseWheel>", lambda _e: "break")
        text.bind("<Button-4>", lambda _e: "break")
        text.bind("<Button-5>", lambda _e: "break")
        try:
            inner = getattr(text, "_textbox", None)
            if inner is not None:
                inner.bind("<MouseWheel>", lambda _e: "break")
                inner.bind("<Button-4>", lambda _e: "break")
                inner.bind("<Button-5>", lambda _e: "break")
        except Exception:  # noqa: BLE001
            pass

        self._autofill_log_frame = frame
        self._autofill_log_text = text
        self._autofill_log_status = status
        self._autofill_log_header = header
        self._autofill_log_bubble = None
        self._autofill_log_collapsed = False
        self._sync_autofill_log_geometry()
        # Shift any already-visible compact toasts left of the terminal column.
        self._place_toast_host()
        if not getattr(self, "_autofill_log_configure_bound", False):
            self.bind("<Configure>", self._sync_autofill_log_geometry, add="+")
            self._autofill_log_configure_bound = True

    def append_autofill_log(self, message: str, *, error: bool = False):
        """Append one substep line to the autofill terminal (opens panel if needed)."""
        if self._autofill_log_frame is None or self._autofill_log_text is None:
            self.open_autofill_log()
        text = self._autofill_log_text
        if text is None:
            return
        line = (message or "").rstrip()
        if not line:
            return
        prefix = "! " if error else "> "
        try:
            text.configure(state="normal")
            text.insert("end", prefix + line + "\n")
            text.see("end")
            text.configure(state="disabled")
        except Exception:  # noqa: BLE001
            pass
        self.set_status(message)
        if self._autofill_log_frame is not None and not self._autofill_log_collapsed:
            try:
                self._autofill_log_frame.lift()
            except Exception:  # noqa: BLE001
                pass

    def finish_autofill_log(self, *, ok: bool = True, auto_close_ms: int = AUTOFILL_LOG_FINISH_MS):
        """Mark the autofill terminal done/failed and show a close countdown."""
        done_label = "完成" if ok else "失败"
        accent = SUCCESS_BTN_HOVER if ok else DANGER_BTN_HOVER
        border = SUCCESS_BTN_FG if ok else DANGER_BTN_FG
        self._autofill_log_accent = border
        if self._autofill_log_status is not None:
            try:
                self._autofill_log_status.configure(
                    text=done_label,
                    text_color=accent,
                )
                if self._autofill_log_frame is not None:
                    self._autofill_log_frame.configure(border_color=border)
            except Exception:  # noqa: BLE001
                pass
        bubble = self._autofill_log_bubble
        if bubble is not None:
            try:
                bubble.configure(text_color=accent, border_color=border)
            except Exception:  # noqa: BLE001
                pass
        if self._autofill_log_finish_after_id is not None:
            try:
                self.after_cancel(self._autofill_log_finish_after_id)
            except Exception:  # noqa: BLE001
                pass
            self._autofill_log_finish_after_id = None
        if not auto_close_ms or auto_close_ms <= 0:
            return

        total_sec = max(1, int(round(auto_close_ms / 1000)))
        state = {"left": total_sec}

        def _tick():
            self._autofill_log_finish_after_id = None
            if self._autofill_log_frame is None:
                return
            left = int(state["left"])
            if left <= 0:
                self.close_autofill_log()
                return
            if self._autofill_log_status is not None:
                try:
                    self._autofill_log_status.configure(
                        text=f"{done_label} · {left}s",
                        text_color=accent,
                    )
                except Exception:  # noqa: BLE001
                    pass
            bubble = self._autofill_log_bubble
            if bubble is not None and self._autofill_log_collapsed:
                try:
                    bubble.configure(text=f"{left}")
                except Exception:  # noqa: BLE001
                    pass
            state["left"] = left - 1
            self._autofill_log_finish_after_id = self.after(1000, _tick)

        _tick()

    def close_autofill_log(self):
        if self._autofill_log_finish_after_id is not None:
            try:
                self.after_cancel(self._autofill_log_finish_after_id)
            except Exception:  # noqa: BLE001
                pass
            self._autofill_log_finish_after_id = None
        frame = self._autofill_log_frame
        bubble = self._autofill_log_bubble
        self._autofill_log_frame = None
        self._autofill_log_text = None
        self._autofill_log_status = None
        self._autofill_log_header = None
        self._autofill_log_bubble = None
        self._autofill_log_collapsed = False
        for widget in (bubble, frame):
            if widget is None:
                continue
            try:
                widget.place_forget()
                widget.destroy()
            except Exception:  # noqa: BLE001
                pass
        self._place_toast_host()

    def show_toast(
        self,
        message: str,
        *,
        title: str = "提醒",
        duration_ms: int = TOAST_DEFAULT_MS,
        action_text: str | None = "关闭",
        undo_text: str | None = None,
        on_undo=None,
        on_complete=None,
        complete_on_timeout: bool = True,
        style: str = "danger",
    ):
        """Show a top-right countdown toast; stacks below any already visible.

        Older toasts stay above; each new toast is appended at the bottom.
        Auto-closes when the countdown reaches 0 (runs on_complete if set).
        Primary action dismisses early and also runs on_complete.
        Optional grey undo button runs on_undo instead and skips on_complete.
        Pass action_text=None (and no undo_text) to hide action buttons.
        Set complete_on_timeout=False so expiry only dismisses (no on_complete).
        style: "danger" (red) or "success" (green).

        Repeating the same title/message/style/actions refreshes the existing
        toast countdown instead of stacking duplicates.
        """
        duration_ms = max(int(duration_ms or 0), TOAST_MIN_MS)
        identity = (
            str(title or ""),
            str(message or ""),
            str(style or "danger"),
            action_text,
            undo_text,
        )
        now_ms = int(self.winfo_toplevel().tk.call("clock", "milliseconds"))

        # Refresh an identical visible toast instead of spawning another copy.
        for entry in self._toasts:
            if entry.get("identity") != identity or entry.get("settled"):
                continue
            entry["duration_ms"] = duration_ms
            entry["deadline_ms"] = now_ms + duration_ms
            entry["on_complete"] = on_complete
            entry["on_undo"] = on_undo
            entry["complete_on_timeout"] = bool(complete_on_timeout)
            seconds = max(1, int(round(duration_ms / 1000)))
            try:
                entry["progress"].set(1.0)
            except Exception:  # noqa: BLE001
                pass
            try:
                entry["countdown"].configure(text=f"{seconds}s")
            except Exception:  # noqa: BLE001
                pass
            # Keep the refreshed toast at the bottom of the stack (most recent).
            try:
                self._toasts.remove(entry)
                self._toasts.append(entry)
            except ValueError:
                pass
            self._relayout_toast_stack()
            self._schedule_toast_tick()
            return

        while len(self._toasts) >= TOAST_STACK_MAX:
            self._dismiss_toast(self._toasts[0], run_complete=False)

        if style == "success":
            accent = SUCCESS_BTN_FG
            accent_hover = SUCCESS_BTN_HOVER
            accent_text = SUCCESS_BTN_TEXT
        else:
            accent = DANGER_BTN_FG
            accent_hover = DANGER_BTN_HOVER
            accent_text = DANGER_BTN_TEXT

        toast_w = self._toast_target_width()
        host = self._ensure_toast_host()
        toast = customtkinter.CTkFrame(
            host,
            width=toast_w,
            corner_radius=TOAST_RADIUS,
            fg_color=TOAST_BG,
            border_width=TOAST_BORDER_WIDTH,
            border_color=accent,
        )
        toast.grid_columnconfigure(0, weight=1)

        header = customtkinter.CTkFrame(toast, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 6))
        header.grid_columnconfigure(0, weight=1)

        customtkinter.CTkLabel(
            header,
            text=title,
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_TITLE, weight="bold"),
            text_color=TOAST_TITLE_COLOR,
        ).grid(row=0, column=0, sticky="ew")

        seconds = max(1, int(round(duration_ms / 1000)))
        countdown = customtkinter.CTkLabel(
            header,
            text=f"{seconds}s",
            anchor="e",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
            text_color=accent_hover,
        )
        countdown.grid(row=0, column=1, sticky="e", padx=(8, 0))

        customtkinter.CTkLabel(
            toast,
            text=message,
            anchor="w",
            justify="left",
            wraplength=max(120, toast_w - 36),
            font=customtkinter.CTkFont(size=FONT_ENTRY),
            text_color=TOAST_MESSAGE_COLOR,
        ).grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))

        show_actions = bool(undo_text) or action_text is not None
        progress = customtkinter.CTkProgressBar(
            toast,
            height=10,
            fg_color=TOAST_PROGRESS_TRACK,
            progress_color=accent,
        )
        progress.grid(
            row=2,
            column=0,
            sticky="ew",
            padx=18,
            pady=(0, 16 if not show_actions else 12),
        )
        progress.set(1.0)

        self._toast_seq += 1
        entry = {
            "id": self._toast_seq,
            "identity": identity,
            "frame": toast,
            "progress": progress,
            "countdown": countdown,
            "duration_ms": duration_ms,
            "deadline_ms": now_ms + duration_ms,
            "on_complete": on_complete,
            "on_undo": on_undo,
            "complete_on_timeout": bool(complete_on_timeout),
            "settled": False,
        }

        if show_actions:
            btn_row = customtkinter.CTkFrame(toast, fg_color="transparent")
            btn_row.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 16))
            if undo_text:
                btn_row.grid_columnconfigure(0, weight=1)
                btn_row.grid_columnconfigure(1, weight=1)
                customtkinter.CTkButton(
                    btn_row,
                    corner_radius=UI_RADIUS,
                    text=undo_text,
                    height=TOAST_BTN_HEIGHT,
                    font=self._button_font(FONT_SECTION),
                    fg_color=SECONDARY_BTN_FG,
                    hover_color=SECONDARY_BTN_HOVER,
                    text_color=SECONDARY_BTN_TEXT,
                    command=lambda e=entry: self._toast_undo(e),
                ).grid(row=0, column=0, sticky="ew", padx=(0, 6))
                customtkinter.CTkButton(
                    btn_row,
                    corner_radius=UI_RADIUS,
                    text=action_text or "关闭",
                    height=TOAST_BTN_HEIGHT,
                    font=self._button_font(FONT_SECTION),
                    fg_color=accent,
                    hover_color=accent_hover,
                    text_color=accent_text,
                    command=lambda e=entry: self._dismiss_toast(e, run_complete=True),
                ).grid(row=0, column=1, sticky="ew", padx=(6, 0))
            else:
                btn_row.grid_columnconfigure(0, weight=1)
                customtkinter.CTkButton(
                    btn_row,
                    corner_radius=UI_RADIUS,
                    text=action_text,
                    height=TOAST_BTN_HEIGHT,
                    font=self._button_font(FONT_SECTION),
                    fg_color=accent,
                    hover_color=accent_hover,
                    text_color=accent_text,
                    command=lambda e=entry: self._dismiss_toast(e, run_complete=True),
                ).grid(row=0, column=0, sticky="ew")

        # Size from content before packing into the stack host.
        toast.update_idletasks()
        scale = self._widget_scaling_factor()
        toast_h = max(1, int(round(toast.winfo_reqheight() / scale)))
        toast.configure(width=toast_w, height=toast_h)
        toast.grid_propagate(False)

        self._toasts.append(entry)
        self._relayout_toast_stack()
        self._schedule_toast_tick()

    def show_success_toast(self, message: str, *, title: str = "OCR提取"):
        self.show_toast(
            message,
            title=title,
            duration_ms=TOAST_SUCCESS_MS,
            action_text="关闭",
            style="success",
        )

    def _schedule_toast_tick(self):
        if self._toast_tick_after_id is not None:
            return
        if not self._toasts:
            return
        self._toast_tick_after_id = self.after(TOAST_TICK_MS, self._toast_tick)

    def _toast_tick(self):
        self._toast_tick_after_id = None
        if not self._toasts:
            return
        now = int(self.winfo_toplevel().tk.call("clock", "milliseconds"))
        expired = []
        for entry in self._toasts:
            remaining = max(0, entry["deadline_ms"] - now)
            duration = entry["duration_ms"] or 1
            ratio = remaining / duration
            try:
                entry["progress"].set(ratio)
            except Exception:  # noqa: BLE001
                pass
            try:
                secs = int((remaining + 999) // 1000)
                entry["countdown"].configure(text=f"{secs}s")
            except Exception:  # noqa: BLE001
                pass
            if remaining <= 0:
                expired.append(entry)
        for entry in expired:
            self._dismiss_toast(
                entry, run_complete=bool(entry.get("complete_on_timeout", True))
            )
        if self._toasts:
            self._toast_tick_after_id = self.after(TOAST_TICK_MS, self._toast_tick)

    def _toast_undo(self, entry: dict):
        if entry.get("settled"):
            return
        entry["settled"] = True
        cb = entry.get("on_undo")
        entry["on_complete"] = None
        entry["on_undo"] = None
        self._dismiss_toast(entry, run_complete=False)
        if cb is not None:
            cb()

    def _dismiss_toast(self, entry: dict, *, run_complete: bool = False):
        if entry not in self._toasts:
            return
        self._toasts.remove(entry)
        frame = entry.get("frame")
        if frame is not None:
            try:
                frame.pack_forget()
                frame.destroy()
            except Exception:  # noqa: BLE001
                pass

        complete_cb = None
        if run_complete and not entry.get("settled"):
            entry["settled"] = True
            complete_cb = entry.get("on_complete")
        entry["on_complete"] = None
        entry["on_undo"] = None

        if self._toasts:
            self._relayout_toast_stack()
        elif self._toast_host is not None:
            try:
                self._toast_host.place_forget()
                self._toast_host.destroy()
            except Exception:  # noqa: BLE001
                pass
            self._toast_host = None

        if complete_cb is not None:
            complete_cb()

    def hide_toast(self, *, run_complete: bool = False):
        """Dismiss all stacked toasts."""
        if self._toast_tick_after_id is not None:
            try:
                self.after_cancel(self._toast_tick_after_id)
            except Exception:  # noqa: BLE001
                pass
            self._toast_tick_after_id = None
        for entry in list(self._toasts):
            self._dismiss_toast(entry, run_complete=run_complete)

    def _apply_min_window_size(self):
        scale = self._widget_scaling_factor()
        if getattr(self, "_compact_min_window", False):
            min_w = WINDOW_MIN_WIDTH_COMPACT
            min_h = WINDOW_MIN_HEIGHT_COMPACT
        else:
            min_w = WINDOW_MIN_WIDTH
            min_h = (
                WINDOW_MIN_HEIGHT_ZOOMED
                if self._ui_zoomed
                else WINDOW_MIN_HEIGHT_NORMAL
            )
        self.minsize(
            int(min_w * scale),
            int(min_h * scale),
        )

    # ---------------------------------------------------------------- extract
    def _pack_certificate_fields(
        self,
        content: customtkinter.CTkFrame,
        *,
        start_row: int = 0,
    ) -> tuple[int, dict[str, customtkinter.CTkEntry | customtkinter.CTkTextbox]]:
        """Build the read/write certificate field stack; return next row + widgets."""
        entries: dict[str, customtkinter.CTkEntry | customtkinter.CTkTextbox] = {}
        row = start_row
        customtkinter.CTkLabel(
            content,
            text="比对字段",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
            anchor="w",
        ).grid(row=row, column=0, sticky="ew", pady=(0, 4))
        row += 1
        for key, label in MATCH_FIELDS:
            customtkinter.CTkLabel(
                content,
                text=label,
                anchor="w",
                width=96,
                font=customtkinter.CTkFont(size=FONT_LABEL),
                text_color="gray60",
            ).grid(row=row, column=0, sticky="w", pady=4)
            entry = self._make_field_entry(content, placeholder=f"请输入{label}")
            entry.grid(row=row, column=0, sticky="ew", padx=(104, 0), pady=4)
            entries[key] = entry
            row += 1

        customtkinter.CTkLabel(
            content,
            text="填写字段",
            font=customtkinter.CTkFont(size=FONT_SECTION, weight="bold"),
            anchor="w",
        ).grid(row=row, column=0, sticky="ew", pady=(14, 8))
        row += 1
        for key, label in METROLOGY_FIELDS:
            customtkinter.CTkLabel(
                content,
                text=label,
                anchor="w",
                width=96,
                font=customtkinter.CTkFont(size=FONT_LABEL),
                text_color="gray60",
            ).grid(row=row, column=0, sticky="w", pady=4)
            entry = self._make_field_entry(content, placeholder=f"请输入{label}")
            entry.grid(row=row, column=0, sticky="ew", padx=(104, 0), pady=4)
            entries[key] = entry
            row += 1

        result_label_wrap = customtkinter.CTkFrame(
            content,
            fg_color="transparent",
            width=96,
            height=ENTRY_HEIGHT,
        )
        result_label_wrap.grid(row=row, column=0, sticky="nw", pady=4)
        result_label_wrap.grid_propagate(False)
        result_label_wrap.grid_rowconfigure(0, weight=1)
        result_label_wrap.grid_columnconfigure(0, weight=1)
        customtkinter.CTkLabel(
            result_label_wrap,
            text="计量结果信息",
            anchor="w",
            font=customtkinter.CTkFont(size=FONT_LABEL),
            text_color="gray60",
        ).grid(row=0, column=0, sticky="w")
        result_box = customtkinter.CTkTextbox(
            content,
            height=RESULT_INFO_HEIGHT,
            corner_radius=UI_RADIUS,
            border_width=FIELD_BORDER_WIDTH,
            border_color=FIELD_FG_COLOR,
            fg_color=FIELD_FG_COLOR,
            text_color=FIELD_TEXT_COLOR,
            activate_scrollbars=False,
            font=customtkinter.CTkFont(size=FONT_ENTRY),
        )
        result_box.grid(row=row, column=0, sticky="ew", padx=(104, 0), pady=4)
        entries["result_info"] = result_box
        return row + 1, entries

    def _build_extract_controls(self, parent: customtkinter.CTkFrame):
        """提取核对: remove-failed pinned at top; editing locked until quarantine done."""
        header, content, footer = self._make_pinned_footer_layout(parent)

        self.remove_failed_button = customtkinter.CTkButton(
            header,
            corner_radius=UI_RADIUS,
            text="移出失败证书",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            font=self._button_font(FONT_SECTION),
            fg_color=DANGER_BTN_FG,
            hover_color=DANGER_BTN_HOVER,
            text_color=DANGER_BTN_TEXT,
            command=self._on_remove_failed_certificates,
        )
        self.remove_failed_button.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        self._style_primary_action_button(self.remove_failed_button)

        next_row, self.field_entries = self._pack_certificate_fields(
            content, start_row=0
        )

        self.extract_errors_label = self._track_content_wrap(
            customtkinter.CTkLabel(
                content,
                text="",
                anchor="w",
                justify="left",
                wraplength=CONTENT_WRAP,
                font=customtkinter.CTkFont(size=FONT_META),
                text_color="#c0392b",
            )
        )
        self.extract_errors_label.grid(
            row=next_row, column=0, sticky="ew", pady=(8, 0)
        )

        self.ocr_progress = customtkinter.CTkProgressBar(footer, height=10)
        self.ocr_progress.set(0)

        self.ocr_progress_label = customtkinter.CTkLabel(
            footer,
            text="OCR 进度 0/0",
            font=customtkinter.CTkFont(size=FONT_META),
            text_color="gray60",
            anchor="w",
        )

        self.ocr_extract_button = customtkinter.CTkButton(
            footer,
            corner_radius=UI_RADIUS,
            text="OCR提取",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            font=self._button_font(FONT_SECTION),
            fg_color=SUCCESS_BTN_FG,
            hover_color=SUCCESS_BTN_HOVER,
            text_color=SUCCESS_BTN_TEXT,
            command=self._on_ocr_extract,
        )
        self._style_primary_action_button(self.ocr_extract_button)

        actions = customtkinter.CTkFrame(footer, fg_color="transparent")
        actions.grid(row=3, column=0, sticky="ew")
        actions.grid_columnconfigure((0, 1), weight=1)

        self.approve_toggle_button = customtkinter.CTkButton(
            actions,
            corner_radius=UI_RADIUS,
            text="批准",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            fg_color=SUCCESS_BTN_FG,
            hover_color=SUCCESS_BTN_HOVER,
            text_color=SUCCESS_BTN_TEXT,
            font=self._button_font(FONT_SECTION),
            command=self._on_toggle_approve_entry,
        )
        self.approve_toggle_button.grid(row=0, column=0, padx=(0, 4), sticky="ew")
        self._style_primary_action_button(self.approve_toggle_button)

        self.remove_toggle_button = customtkinter.CTkButton(
            actions,
            corner_radius=UI_RADIUS,
            text="移除",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            fg_color=DANGER_BTN_FG,
            hover_color=DANGER_BTN_HOVER,
            text_color=DANGER_BTN_TEXT,
            font=self._button_font(FONT_SECTION),
            command=self._on_toggle_remove_entry,
        )
        self.remove_toggle_button.grid(row=0, column=1, padx=(4, 0), sticky="ew")
        self._style_primary_action_button(self.remove_toggle_button)

        self._update_extract_ocr_ui()
        self._update_remove_failed_button()
        self._update_review_fields_state()
        self._update_approve_toggle_button()
        self._update_remove_toggle_button()

    # ----------------------------------------------------------- automate
    def _build_automate_controls(self, parent: customtkinter.CTkFrame):
        """自动化: Excel export + autofill run controls."""
        header, content, footer = self._make_pinned_footer_layout(parent)
        self.automate_page_header = header

        self.custom_dir_autofill_button = customtkinter.CTkButton(
            header,
            corner_radius=UI_RADIUS,
            text="自定义 Excel 导入并自动填写",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            font=self._button_font(FONT_SECTION),
            fg_color=CUSTOM_AUTOFILL_BTN_FG,
            hover_color=CUSTOM_AUTOFILL_BTN_HOVER,
            text_color=CUSTOM_AUTOFILL_BTN_TEXT,
            command=self._on_custom_excel_autofill,
        )
        self._style_primary_action_button(self.custom_dir_autofill_button)
        self._update_custom_dir_autofill_button()

        _next_row, self.automate_field_entries = self._pack_certificate_fields(
            content, start_row=0
        )
        self._set_field_widgets_locked(self.automate_field_entries, True)

        self.export_excel_button = customtkinter.CTkButton(
            footer,
            corner_radius=UI_RADIUS,
            text="导出 Excel (0)",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            font=self._button_font(FONT_SECTION),
            fg_color=SECONDARY_BTN_FG,
            hover_color=SECONDARY_BTN_HOVER,
            text_color=SECONDARY_BTN_TEXT,
            command=self._on_export_excel,
        )
        self.export_excel_button.grid(row=0, column=0, sticky="ew")
        self._style_primary_action_button(self.export_excel_button)

        self.autofill_button = customtkinter.CTkButton(
            footer,
            corner_radius=UI_RADIUS,
            text="导出并自动填写 (0)",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            font=self._button_font(FONT_SECTION),
            fg_color=PRIMARY_BTN_FG,
            hover_color=PRIMARY_BTN_HOVER,
            text_color=PRIMARY_BTN_TEXT,
            command=self._on_master_autofill,
        )
        self.autofill_button.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        self._style_primary_action_button(self.autofill_button)

        self.autofill_controls_frame = customtkinter.CTkFrame(
            footer, fg_color="transparent"
        )
        self.autofill_controls_frame.grid_columnconfigure((0, 1), weight=1)
        self.autofill_pause_button = customtkinter.CTkButton(
            self.autofill_controls_frame,
            corner_radius=UI_RADIUS,
            text="暂停",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            font=self._button_font(FONT_SECTION),
            fg_color=SECONDARY_BTN_FG,
            hover_color=SECONDARY_BTN_HOVER,
            text_color=SECONDARY_BTN_TEXT,
            command=self._on_autofill_pause_toggle,
        )
        self.autofill_pause_button.grid(row=0, column=0, sticky="ew", padx=(0, 4))
        self._style_primary_action_button(self.autofill_pause_button)

        self.autofill_exit_button = customtkinter.CTkButton(
            self.autofill_controls_frame,
            corner_radius=UI_RADIUS,
            text="退出",
            height=PRIMARY_ACTION_BTN_HEIGHT,
            round_height_to_even_numbers=False,
            font=self._button_font(FONT_SECTION),
            fg_color=DANGER_BTN_FG,
            hover_color=DANGER_BTN_HOVER,
            text_color=DANGER_BTN_TEXT,
            command=self._on_autofill_exit,
        )
        self.autofill_exit_button.grid(row=0, column=1, sticky="ew", padx=(4, 0))
        self._style_primary_action_button(self.autofill_exit_button)

    def _schedule_doc_list_scrollbar_sync(self, _event=None):
        if getattr(self, "_doc_list_scroll_after", None) is not None:
            try:
                self.after_cancel(self._doc_list_scroll_after)
            except Exception:  # noqa: BLE001
                pass
        self._doc_list_scroll_after = self.after(20, self._update_doc_list_scrollbar)

    def _bind_scrollable_mousewheel(self, scrollable, widget):
        """Make trackpad / mouse-wheel scroll work over content, not only the bar."""
        handler = lambda e, sf=scrollable: self._on_scrollable_mousewheel(sf, e)
        if sys.platform == "darwin":
            widget.bind("<MouseWheel>", handler, add="+")
        elif sys.platform.startswith("win"):
            widget.bind("<MouseWheel>", handler, add="+")
        else:
            widget.bind("<Button-4>", handler, add="+")
            widget.bind("<Button-5>", handler, add="+")
        for child in widget.winfo_children():
            self._bind_scrollable_mousewheel(scrollable, child)

    def _bind_doc_list_mousewheel(self, widget):
        if hasattr(self, "doc_list_frame"):
            self._bind_scrollable_mousewheel(self.doc_list_frame, widget)

    def _mousewheel_scroll_steps(self, event) -> int:
        """Canvas yview units for one wheel/trackpad event (tuned for usable speed)."""
        # Default Tk/CTk is ~1 unit per notch — feels like a crawl on long lists.
        speed = 5
        if sys.platform == "darwin":
            delta = int(getattr(event, "delta", 0) or 0)
            if delta == 0:
                return 0
            # Trackpad often sends ±1; multiply. Larger deltas (mouse) scale less.
            if abs(delta) <= 1:
                return -delta * speed
            return int(-delta * max(2, speed // 2))
        if sys.platform.startswith("win"):
            # Windows path uses fraction-of-view scrolling in _on_scrollable_mousewheel.
            delta = int(getattr(event, "delta", 0) or 0)
            notches = int(delta / 120) if delta else 0
            if notches == 0 and delta:
                notches = 1 if delta > 0 else -1
            return -notches
        return (-speed) if getattr(event, "num", 0) == 4 else speed

    def _scrollable_yview_fractions(self, canvas) -> tuple[float, float]:
        try:
            first, last = canvas.yview()
            return float(first), float(last)
        except Exception:  # noqa: BLE001
            return 0.0, 1.0

    def _clamp_scrollable_yview(self, canvas, first: float) -> None:
        """Move canvas to ``first`` without allowing overscroll past ends."""
        cur_first, cur_last = self._scrollable_yview_fractions(canvas)
        view = max(0.0, min(1.0, cur_last - cur_first))
        # When content fits, stay locked at top.
        if view >= 1.0 - 1e-6:
            canvas.yview_moveto(0.0)
            return
        max_first = max(0.0, 1.0 - view)
        target = max(0.0, min(max_first, float(first)))
        canvas.yview_moveto(target)

    def _on_scrollable_mousewheel(self, scrollable, event):
        if (
            self._autofill_busy
            and hasattr(self, "doc_list_frame")
            and scrollable is self.doc_list_frame
        ):
            return "break"
        if (
            hasattr(self, "doc_list_frame")
            and scrollable is self.doc_list_frame
            and self._doc_name_tip is not None
        ):
            # Tip overlays the list and would block / drift while rows move.
            self._hide_doc_name_tip()
        try:
            canvas = scrollable._parent_canvas
        except Exception:  # noqa: BLE001
            return "break"

        first, last = self._scrollable_yview_fractions(canvas)
        view = max(0.0, last - first)
        # Content fits — no scroll, and kill bounce/flicker at the edges.
        if view >= 1.0 - 1e-6:
            self._clamp_scrollable_yview(canvas, 0.0)
            return "break"

        # Positive ``toward_top`` moves the viewport toward the start of content.
        toward_top = 0.0
        if sys.platform.startswith("win"):
            delta = int(getattr(event, "delta", 0) or 0)
            notches = int(delta / 120) if delta else 0
            if notches == 0 and delta:
                notches = 1 if delta > 0 else -1
            # Win: positive delta → scroll up (toward top).
            toward_top = float(notches) * max(view, 0.08) * 0.55
        elif sys.platform == "darwin":
            delta = int(getattr(event, "delta", 0) or 0)
            if delta == 0:
                return "break"
            # Darwin: positive delta → toward top. Keep speed similar to before.
            if abs(delta) <= 1:
                toward_top = float(delta) * 0.045
            else:
                toward_top = float(delta) * 0.02
        else:
            # X11 Button-4 = up, Button-5 = down.
            num = int(getattr(event, "num", 0) or 0)
            if num == 4:
                toward_top = max(view, 0.08) * 0.45
            elif num == 5:
                toward_top = -max(view, 0.08) * 0.45
            else:
                return "break"

        if toward_top == 0.0:
            return "break"

        eps = 1e-4
        at_top = first <= eps
        at_bottom = last >= 1.0 - eps
        if toward_top > 0 and at_top:
            self._clamp_scrollable_yview(canvas, 0.0)
            return "break"
        if toward_top < 0 and at_bottom:
            self._clamp_scrollable_yview(canvas, 1.0)
            return "break"

        self._clamp_scrollable_yview(canvas, first - toward_top)
        return "break"

    def _on_doc_list_mousewheel(self, event):
        if not hasattr(self, "doc_list_frame"):
            return
        # Overlay would otherwise stay pinned while rows move under it.
        if self._doc_name_tip is not None:
            self._hide_doc_name_tip()
        return self._on_scrollable_mousewheel(self.doc_list_frame, event)

    def _update_scrollable_scrollbar(
        self, scrollable, *, after_attr: str, vcenter_when_fits: bool = True
    ):
        """Doc-list scroll sync (auto-hide bar; optional vertical centering)."""
        setattr(self, after_attr, None)
        if scrollable is None:
            return
        canvas = scrollable._parent_canvas
        scrollbar = scrollable._scrollbar
        scrollable.update_idletasks()
        bbox = canvas.bbox("all")
        window_id = getattr(scrollable, "_create_window_id", None)
        if bbox is None:
            canvas.configure(scrollregion=(0, 0, 0, 0))
            scrollbar.grid_remove()
            if window_id is not None:
                canvas.coords(window_id, 0, 0)
            return

        _x1, y1, x2, y2 = bbox
        content_height = max(0, y2 - y1)
        canvas_height = max(canvas.winfo_height(), 1)

        if content_height > canvas_height + 1:
            if window_id is not None:
                canvas.coords(window_id, 0, 0)
            canvas.configure(scrollregion=(0, 0, max(0, x2), max(y2, content_height)))
            if getattr(self, "_hide_scrollbars", False):
                scrollbar.grid_remove()
            else:
                scrollable._create_grid()
            top, _bottom = canvas.yview()
            if top < 0 or top > 1:
                canvas.yview_moveto(0)
        else:
            scrollbar.grid_remove()
            y_off = (
                max(0, (canvas_height - content_height) // 2) if vcenter_when_fits else 0
            )
            if window_id is not None:
                canvas.coords(window_id, 0, y_off)
            canvas.configure(
                scrollregion=(0, 0, max(0, x2), max(canvas_height, y_off + content_height))
            )
            canvas.yview_moveto(0)

    def _update_doc_list_scrollbar(self, _event=None):
        if not hasattr(self, "doc_list_frame"):
            self._doc_list_scroll_after = None
            return
        self._update_scrollable_scrollbar(
            self.doc_list_frame,
            after_attr="_doc_list_scroll_after",
            vcenter_when_fits=False,
        )

    def _pick_folder(self):
        path = filedialog.askdirectory(parent=self, title="选择证书文件夹")
        if not path:
            self.set_status("未选择文件夹")
            return
        self.show_step("extract")
        self._reset_workflow_to_extract()
        self._load_folder(path)

    def _clear_extract(self):
        """Ask for confirmation before wiping the document list."""
        if self._autofill_busy:
            self.set_status("自动填写进行中，请先退出后再清空…")
            return
        if self._extract_busy:
            self.set_status("正在处理中，请稍候…")
            return
        n = len(self._imported_files)
        if n == 0 and not self._source_folder:
            self.set_status("列表已空")
            return
        self.show_toast(
            f"将清空文档列表中的 {n} 份证书及解析结果。\n此操作不可撤销。",
            title="清空文档列表",
            action_text="确认清空",
            undo_text="取消",
            on_complete=self._confirm_clear_extract,
            complete_on_timeout=False,
            duration_ms=TOAST_DEFAULT_MS,
            style="danger",
        )
        self.set_status("确认清空文档列表？")

    def _confirm_clear_extract(self):
        self.show_step("extract")
        self._extract_busy = False
        self._source_folder = None
        self._imported_files.clear()
        self._parse_results.clear()
        self._autofill_queue.clear()
        self._removed_paths.clear()
        self._current_cert_index = 0
        self._selected_path = None
        self._rebuild_doc_list()
        self.folder_label.configure(text="未选择文件夹")
        self._clear_extract_fields_display()
        self._reset_ocr_progress()
        if hasattr(self, "field_entries"):
            self._clear_approve_fields()
        self._update_autofill_button()
        self._sync_pdf_preview()
        self.set_status("已清空提取列表")
        self.show_success_toast("已清空文档列表。", title="清空")
        self._reset_workflow_to_extract()

    def _load_folder(self, folder: str):
        if self._extract_busy:
            self.set_status("正在处理中，请稍候…")
            return

        self._reset_workflow_to_extract()
        self.folder_label.configure(text=folder)
        self.set_status("正在扫描文件夹…")
        self.update_idletasks()

        try:
            pdfs = find_pdfs_in_folder(folder, recursive=False)
        except Exception as exc:  # noqa: BLE001
            self.set_status(f"文件夹加载失败：{exc}")
            return

        paths = [str(p) for p in pdfs]
        self._source_folder = folder
        self._imported_files = paths
        self._parse_results.clear()
        self._autofill_queue.clear()
        self._removed_paths.clear()
        self._current_cert_index = 0
        self._rebuild_doc_list()

        if not paths:
            self._selected_path = None
            self._sync_pdf_preview()
            self.set_status("文件夹中没有 PDF")
            return

        # Embedded-text parse only — OCR is manual via 「OCR提取」.
        self._extract_busy = True
        results: dict[str, ParseResult] = {}
        try:
            for i, path in enumerate(paths, start=1):
                self.set_status(f"解析进度 {i}/{len(paths)}")
                self.update_idletasks()
                results[path] = parse_certificate(
                    path,
                    use_ocr_fallback=False,
                    force_ocr=False,
                    extra_label_aliases=self._extra_label_aliases(),
                )
            self._on_folder_loaded(folder, paths, results)
        except Exception as exc:  # noqa: BLE001
            self._on_folder_fail(str(exc))
        finally:
            self._extract_busy = False

    def _on_folder_fail(self, message: str):
        self._extract_busy = False
        self.set_status(f"文件夹加载失败：{message}")

    def _on_folder_loaded(
        self,
        folder: str,
        paths: list[str],
        results: dict[str, ParseResult],
    ):
        self._extract_busy = False
        self._source_folder = folder
        self._imported_files = paths
        self._parse_results = results
        self._autofill_queue.clear()
        self._removed_paths.clear()
        self._sort_imported_files()
        self._current_cert_index = 0
        ok = sum(1 for r in results.values() if r.ok)
        pending = sum(1 for p in self._imported_files if self._cert_needs_ocr(p))
        self._rebuild_doc_list()
        self._reset_ocr_progress()
        if self._imported_files:
            self._select_document(self._imported_files[0])
        if pending:
            if self._ocr_enabled:
                self.set_status(f"文件夹加载完成：已解析 {ok} · 待 OCR {pending}")
            else:
                self.set_status(f"文件夹加载完成：已解析 {ok} · 未解析 {pending}")
        else:
            self.set_status(f"文件夹加载完成：{len(paths)} PDF")

    def _cert_is_parsed(self, path: str) -> bool:
        result = self._parse_results.get(path)
        return bool(result and result.ok)

    def _cert_needs_ocr(self, path: str) -> bool:
        """True when embedded-text parse left the cert incomplete / empty."""
        return not self._cert_is_parsed(path)

    def _sort_imported_files(self):
        """Parsed certificates first, then unparsed / pending OCR."""
        parsed = [p for p in self._imported_files if self._cert_is_parsed(p)]
        pending = [p for p in self._imported_files if not self._cert_is_parsed(p)]
        self._imported_files = parsed + pending
        if self._selected_path in self._imported_files:
            self._current_cert_index = self._imported_files.index(self._selected_path)
        elif self._imported_files:
            self._current_cert_index = 0
            self._selected_path = self._imported_files[0]
        else:
            self._current_cert_index = 0
            self._selected_path = None

    def _reset_ocr_progress(self):
        if hasattr(self, "ocr_progress"):
            self.ocr_progress.set(0)
        if hasattr(self, "ocr_progress_label"):
            self.ocr_progress_label.configure(text="OCR 进度 0/0")

    def _set_ocr_progress(self, current: int, total: int, name: str = ""):
        total = max(total, 1)
        if hasattr(self, "ocr_progress"):
            self.ocr_progress.set(current / total)
        label = f"OCR 进度 {current}/{total}"
        if name:
            label += f" · {name}"
        if hasattr(self, "ocr_progress_label"):
            self.ocr_progress_label.configure(text=label)

    def _failed_items_dir_for_import(self) -> Path | None:
        """Return ``{import_folder}/failed_items`` when an import folder is set."""
        folder = (self._source_folder or "").strip()
        if not folder:
            return None
        root = Path(folder)
        if not root.is_dir():
            return None
        return (root / FAILED_ITEMS_SUBDIR).resolve()

    def _import_folder_root(self) -> Path | None:
        """Return the selected import folder when it exists."""
        folder = (self._source_folder or "").strip()
        if not folder:
            return None
        root = Path(folder)
        return root.resolve() if root.is_dir() else None

    def _next_excel_export_path(self) -> Path:
        """Save batch Excel in the import folder root (fallback: project exports/)."""
        root = self._import_folder_root()
        return next_export_path(directory=root)

    def _excel_export_location_label(self, path: Path) -> str:
        """Short path label for toasts/status (prefer import-folder relative)."""
        root = self._import_folder_root()
        try:
            if root is not None:
                return str(path.relative_to(root))
        except Exception:  # noqa: BLE001
            pass
        return path.name

    def _failed_items_dir_short(self) -> str:
        out = self._failed_items_dir_for_import()
        if out is None:
            return f"…/{FAILED_ITEMS_SUBDIR}"
        try:
            parent = out.parent.name
            return f"{parent}/{FAILED_ITEMS_SUBDIR}"
        except Exception:  # noqa: BLE001
            return FAILED_ITEMS_SUBDIR

    def _unique_failed_path(self, src: Path, out_dir: Path) -> Path | None:
        """Destination path under ``out_dir``, or None if the same name already exists.

        Same filename in the failed folder counts as a duplicate — skip re-copy.
        """
        out_dir.mkdir(parents=True, exist_ok=True)
        dest = out_dir / src.name
        if dest.exists():
            return None
        return dest

    def _quarantine_failed_paths(self, paths: list[str]) -> tuple[int, int]:
        """Copy failed PDFs into the import folder's failed_items subdir and drop them.

        Returns ``(removed_from_queue, newly_copied)``. Same-name files already in
        the failed folder are not copied again, but are still removed from the queue.
        """
        out_dir = self._failed_items_dir_for_import()
        if out_dir is None:
            self.set_status("无法移出失败证书：未选择导入文件夹")
            self.show_toast(
                "请先选择导入文件夹，失败证书将保存到其中的 "
                f"「{FAILED_ITEMS_SUBDIR}/」子目录。",
                title="移出失败证书",
            )
            return 0, 0

        removed = 0
        copied = 0
        seen_names: set[str] = set()
        # Preserve order while dropping duplicate source paths in this batch.
        unique_paths: list[str] = []
        seen_paths: set[str] = set()
        for path in paths:
            if path in seen_paths:
                continue
            seen_paths.add(path)
            unique_paths.append(path)

        for path in unique_paths:
            src = Path(path)
            name_key = src.name.casefold()
            try:
                if name_key not in seen_names:
                    seen_names.add(name_key)
                    dest = self._unique_failed_path(src, out_dir)
                    if dest is None:
                        self.set_status(f"失败文件夹已有同名文件，跳过复制：{src.name}")
                    else:
                        shutil.copy2(src, dest)
                        copied += 1
            except Exception as exc:  # noqa: BLE001
                self.set_status(f"无法保存失败项 {src.name}：{exc}")
                continue

            if path in self._imported_files:
                self._imported_files.remove(path)
            self._parse_results.pop(path, None)
            if path in self._autofill_queue:
                self._autofill_queue.remove(path)
            self._removed_paths.discard(path)
            if self._selected_path == path:
                self._selected_path = None
            removed += 1

        self._sort_imported_files()
        return removed, copied

    def _on_ocr_extract(self):
        """Manually run PaddleOCR on certificates that still need it."""
        if not self._ocr_enabled:
            self._on_remove_failed_certificates()
            return
        if self._extract_busy or self._autofill_busy:
            self.set_status("正在处理中，请稍候…")
            return
        if not self._imported_files:
            self.set_status("请先选择文件夹")
            return

        targets = [p for p in self._imported_files if self._cert_needs_ocr(p)]
        if not targets:
            self.set_status("全部证书已有文本解析结果，无需 OCR")
            self.show_success_toast("没有需要 OCR 的证书。")
            return

        self._extract_busy = True
        if hasattr(self, "ocr_extract_button"):
            self.ocr_extract_button.configure(state="disabled")
        self._set_ocr_progress(0, len(targets))
        self.set_status(f"OCR提取开始：{len(targets)} 份")
        aliases = self._extra_label_aliases()

        def worker():
            done = 0
            failed_paths: list[str] = []
            try:
                for i, path in enumerate(targets, start=1):
                    name = Path(path).name
                    self.after(
                        0,
                        lambda i=i, name=name, n=len(targets): self._set_ocr_progress(i, n, name),
                    )
                    try:
                        result = parse_certificate(
                            path,
                            use_ocr_fallback=False,
                            force_ocr=True,
                            extra_label_aliases=aliases,
                        )
                        self._parse_results[path] = result
                        if result.ok:
                            done += 1
                        else:
                            failed_paths.append(path)
                    except Exception:  # noqa: BLE001
                        failed_paths.append(path)
                self.after(
                    0,
                    lambda: self._on_ocr_extract_done(done, failed_paths, len(targets)),
                )
            except Exception as exc:  # noqa: BLE001
                self.after(0, lambda: self._on_ocr_extract_fail(str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _on_remove_failed_certificates(self):
        """When OCR is off: quarantine unparsed / failed certificates."""
        if self._extract_busy or self._autofill_busy:
            self.set_status("正在处理中，请稍候…")
            return
        if not self._imported_files:
            self.set_status("请先选择文件夹")
            return

        targets = [p for p in self._imported_files if self._cert_needs_ocr(p)]
        if not targets:
            self.set_status("没有未解析或失败的证书 · 可开始核对")
            self.show_success_toast("没有需要移出的证书。", title="移出失败证书")
            self._advance_to_review()
            return

        self._pending_quarantine_paths = list(targets)
        msg = f"即将移出 {len(targets)} 份未解析/失败证书"
        self.set_status(msg)
        self.show_toast(
            f"{msg}\n倒计时结束后移至 {self._failed_items_dir_short()}/，可点撤销保留。",
            title="移出失败证书",
            action_text="立即移出",
            undo_text="撤销",
            on_undo=self._undo_pending_quarantine,
            on_complete=self._commit_pending_quarantine,
        )

    def _on_ocr_extract_done(self, done: int, failed_paths: list[str], total: int):
        self._extract_busy = False
        if hasattr(self, "ocr_extract_button"):
            self.ocr_extract_button.configure(state="normal")

        self._sort_imported_files()
        self._rebuild_doc_list()
        self._update_autofill_button()
        if self._imported_files:
            select = self._selected_path if self._selected_path in self._imported_files else self._imported_files[0]
            self._select_document(select)
        else:
            self._selected_path = None
            self._clear_extract_fields_display()
            self._sync_pdf_preview()

        if hasattr(self, "ocr_progress"):
            self.ocr_progress.set(1 if total else 0)
        if hasattr(self, "ocr_progress_label"):
            self.ocr_progress_label.configure(text=f"OCR 进度 {total}/{total}")

        msg = f"OCR 完成 · 成功 {done}/{total}"
        if failed_paths:
            pending = list(failed_paths)
            self._pending_quarantine_paths = pending
            msg += f" · 失败 {len(pending)} 份将移出队列"
            self.set_status(msg)
            self.show_toast(
                f"{msg}\n倒计时结束后移至 {self._failed_items_dir_short()}/，可点撤销保留。",
                title="OCR提取",
                action_text="立即移出",
                undo_text="撤销",
                on_undo=self._undo_pending_quarantine,
                on_complete=self._commit_pending_quarantine,
            )
            return

        self.set_status(msg)
        self.show_success_toast(msg)
        self._advance_to_review()

    def _undo_pending_quarantine(self):
        count = len(self._pending_quarantine_paths)
        self._pending_quarantine_paths = []
        msg = f"已撤销移出 · 失败 {count} 份仍留在队列"
        self.set_status(msg)

    def _cancel_pending_quarantine(self):
        """Drop a pending fail-folder countdown without copying files."""
        if self._pending_quarantine_paths:
            self._pending_quarantine_paths = []
        # Dismiss only the toast(s) that would quarantine on complete.
        for entry in list(self._toasts):
            if entry.get("on_complete") is self._commit_pending_quarantine:
                self._dismiss_toast(entry, run_complete=False)

    def _commit_pending_quarantine(self):
        paths = list(self._pending_quarantine_paths)
        self._pending_quarantine_paths = []
        if not paths:
            self.show_success_toast("没有需要移出的证书。")
            self._advance_to_review()
            return
        moved = self._quarantine_failed_paths(paths)
        self._sort_imported_files()
        self._rebuild_doc_list()
        self._update_autofill_button()
        if self._imported_files:
            select = self._selected_path if self._selected_path in self._imported_files else self._imported_files[0]
            self._select_document(select)
        else:
            self._selected_path = None
            self._clear_extract_fields_display()
            self._sync_pdf_preview()
        removed, copied = moved
        msg = (
            f"失败 {removed} 份已移出队列 · 新复制 {copied} 份到 "
            f"{self._failed_items_dir_short()}/"
        )
        self.set_status(msg)
        self.show_success_toast(msg)
        self._advance_to_review()

    def _on_ocr_extract_fail(self, message: str):
        self._extract_busy = False
        if hasattr(self, "ocr_extract_button"):
            self.ocr_extract_button.configure(state="normal")
        self._reset_ocr_progress()
        self.set_status(f"OCR 失败：{message}")
        self.show_toast(f"OCR 失败：{message}", title="OCR提取")

    def _doc_status_kind(self, path: str) -> str:
        """Return 'bad', 'ok', or '' for document list status indicators."""
        if path in self._removed_paths:
            return "bad"
        if path in self._autofill_queue:
            return "ok"
        return ""

    def _doc_status_mark(self, path: str) -> str:
        kind = self._doc_status_kind(path)
        if not kind:
            return ""
        if self._status_dots:
            return DOC_STATUS_DOT
        return "❌" if kind == "bad" else "✅"

    def _doc_leading_label(self, path: str, index: int) -> str:
        """Left column: index number; dots variant keeps the number beside the status dot."""
        num = str(index + 1)
        mark = self._doc_status_mark(path)
        if not mark:
            return num
        if self._status_dots:
            return f"{num}{DOC_STATUS_DOT}"
        return mark

    def _doc_name_label(self, path: str, index: int) -> str:
        return Path(path).name

    def _appearance_color(self, color) -> str:
        """Resolve a CTk (light, dark) color tuple for the current mode."""
        if isinstance(color, (tuple, list)) and len(color) >= 2:
            mode = customtkinter.get_appearance_mode()
            return color[1] if str(mode).lower() == "dark" else color[0]
        return str(color)

    def _doc_row_surface_bg(self, surface) -> str:
        """Solid bg for canvas cells — never leave system default (black/white flash)."""
        if surface is None or surface == "transparent":
            return self._appearance_color(EMBED_BG_COLOR)
        return self._appearance_color(surface)

    def _doc_list_font_scale(self) -> float:
        """Scale factor for doc-list canvas text (1.0 = pre-zoom-font-fix sizes)."""
        if self._doc_list_scale_fonts:
            return self._widget_scaling_factor()
        return 1.0

    def _doc_list_ctk_font_size(self, size: int) -> int:
        """CTk design-unit size so rendered text matches doc-list canvas fonts.

        CTk multiplies font size by widget scaling; doc-list canvas text applies
        scaling only when the list-text switch is on.
        """
        if self._doc_list_scale_fonts:
            return max(1, int(size))
        scale = self._widget_scaling_factor()
        return max(1, int(round(size / scale)))

    def _doc_tk_font(self, size: int) -> tkfont.Font:
        """Tk font for doc-list cells; optionally follows CTk UI zoom."""
        scale = self._doc_list_font_scale()
        scaled = max(1, int(round(size * scale)))
        return tkfont.Font(family="SF Pro Text", size=scaled)

    def _make_doc_text_canvas(
        self,
        parent,
        *,
        text: str,
        fill,
        size: int,
        width: int | None = None,
        anchor: str = "w",
    ) -> tuple[tk.Canvas, int]:
        """Plain Canvas text cell — no tk.Label, so macOS long-press has nothing to drag."""
        scale = self._doc_list_font_scale()
        # Leave a few px so the parent CTkFrame rounded corners stay visible.
        h = max(18, int(round((DOC_ROW_HEIGHT - 8) * scale)))
        canvas = tk.Canvas(
            parent,
            height=h,
            highlightthickness=0,
            bd=0,
            relief="flat",
            cursor="hand2",
            bg=self._doc_row_surface_bg("transparent"),
        )
        if width is not None:
            canvas.configure(width=max(1, int(round(width * scale))))
        fill_color = self._appearance_color(fill)
        font = self._doc_tk_font(size)
        pad_x = int(round(10 * scale)) if anchor != "center" else 0
        x = (int(round(width * scale)) // 2) if (width is not None and anchor == "center") else pad_x
        text_id = canvas.create_text(
            x,
            h // 2,
            text=text,
            fill=fill_color,
            font=font,
            anchor="center" if anchor == "center" else "w",
        )
        canvas._vincert_text_id = text_id
        canvas._vincert_anchor = anchor
        canvas._vincert_font = font
        canvas._vincert_design_size = size
        canvas._vincert_pad_x = pad_x

        def _on_configure(event, c=canvas, tid=text_id, anc=anchor):
            if event.width <= 1 or event.height <= 1:
                return
            if anc == "center":
                c.coords(tid, event.width / 2, event.height / 2)
            else:
                c.coords(tid, getattr(c, "_vincert_pad_x", 10), event.height / 2)

        canvas.bind("<Configure>", _on_configure, add="+")
        return canvas, text_id

    def _set_doc_canvas_text(
        self,
        canvas: tk.Canvas,
        text: str,
        *,
        fill,
        size: int | None = None,
    ) -> None:
        text_id = getattr(canvas, "_vincert_text_id", None)
        if text_id is None:
            return
        fill_color = self._appearance_color(fill)
        kwargs = {"text": text, "fill": fill_color}
        if size is not None:
            canvas._vincert_design_size = size
            font = self._doc_tk_font(size)
            canvas._vincert_font = font
            kwargs["font"] = font
        canvas.itemconfigure(text_id, **kwargs)

    def _refresh_doc_list_fonts(self):
        """Re-apply fonts/heights after UI zoom or list-text-scale toggle."""
        if not self._doc_rows:
            return
        scale = self._doc_list_font_scale()
        h = max(18, int(round((DOC_ROW_HEIGHT - 8) * scale)))
        pad_x = int(round(10 * scale))
        mark_w = max(1, int(round(DOC_MARK_COL_WIDTH * scale)))
        for path, row in self._doc_rows.items():
            mark = row.get("mark")
            name = row.get("name")
            if mark is not None:
                try:
                    mark.configure(height=h, width=mark_w)
                    mark._vincert_pad_x = 0
                except Exception:  # noqa: BLE001
                    pass
            if name is not None:
                try:
                    name.configure(height=h)
                    name._vincert_pad_x = pad_x
                except Exception:  # noqa: BLE001
                    pass
        self._highlight_selected_doc()
        self._schedule_doc_list_scrollbar_sync()
        self._refresh_doc_name_tip_font()

    def _refresh_doc_name_tip_font(self):
        """Rebuild an open name hover tip so it follows the list text-size switch."""
        path = self._doc_name_tip_path
        if path is None or self._doc_name_tip is None:
            return
        self._hide_doc_name_tip()
        self._show_doc_name_tip(path)

    def _set_doc_canvas_surface(self, canvas: tk.Canvas, surface) -> None:
        try:
            canvas.configure(bg=self._doc_row_surface_bg(surface))
        except Exception:  # noqa: BLE001
            pass

    def _is_doc_name_truncated(self, row: dict) -> bool:
        """True when the visible name canvas width crops the full filename."""
        try:
            canvas = row["name"]
            canvas.update_idletasks()
            visible = int(canvas.winfo_width())
            if visible <= 1:
                return False
            font = getattr(canvas, "_vincert_font", None)
            text = row.get("full_name") or ""
            if font is not None:
                needed = int(font.measure(text))
            else:
                bbox = canvas.bbox(row["name_id"])
                needed = int(bbox[2] - bbox[0]) if bbox else 0
            return needed > visible - 4
        except Exception:  # noqa: BLE001
            return False

    def _hide_doc_name_tip(self):
        tip = self._doc_name_tip
        self._doc_name_tip = None
        self._doc_name_tip_path = None
        if tip is None:
            return
        try:
            tip.place_forget()
            tip.destroy()
        except Exception:  # noqa: BLE001
            pass

    def _clear_doc_row_hovers(self, *, keep: str | None = None):
        """Clear sticky hover washes; optionally keep one path hovered."""
        for path, row in self._doc_rows.items():
            if path == keep or path == self._selected_path:
                continue
            try:
                row["frame"].configure(fg_color="transparent")
                self._sync_doc_row_surfaces(path, "transparent")
            except Exception:  # noqa: BLE001
                pass
        self._doc_hover_path = keep

    def _show_doc_name_tip(self, path: str):
        if self._autofill_busy:
            self._hide_doc_name_tip()
            return
        row = self._doc_rows.get(path)
        if row is None:
            self._hide_doc_name_tip()
            return
        name_canvas = row["name"]
        full_name = row.get("full_name") or self._doc_name_label(path, row.get("index", 0))
        if not self._is_doc_name_truncated(row):
            if self._doc_name_tip_path == path:
                self._hide_doc_name_tip()
            return
        if self._doc_name_tip_path == path and self._doc_name_tip is not None:
            return

        self._hide_doc_name_tip()
        selected = path == self._selected_path
        tip = customtkinter.CTkFrame(
            self,
            corner_radius=UI_RADIUS,
            fg_color=self._accent_doc_row_active() if selected else TOAST_BG,
            border_width=0 if selected else TOAST_BORDER_WIDTH,
            border_color=self._accent_outline(),
        )
        label = customtkinter.CTkLabel(
            tip,
            text=full_name,
            anchor="w",
            justify="left",
            font=customtkinter.CTkFont(size=self._doc_list_ctk_font_size(FONT_BODY)),
            text_color=DOC_ROW_ACTIVE_TEXT if selected else TOAST_TITLE_COLOR,
            fg_color="transparent",
        )
        label.pack(padx=DOC_NAME_TIP_PADX, pady=DOC_NAME_TIP_PADY)

        tip.bind("<Leave>", lambda _e, p=path: self._on_doc_name_tip_leave(p))
        label.bind("<Leave>", lambda _e, p=path: self._on_doc_name_tip_leave(p))
        tip.bind("<ButtonPress-1>", lambda _e, p=path: self._on_doc_row_press(p))
        label.bind("<ButtonPress-1>", lambda _e, p=path: self._on_doc_row_press(p))
        # Tip sits above the list — forward wheel so scrolling still works.
        self._bind_doc_list_mousewheel(tip)
        self._bind_doc_list_mousewheel(label)

        tip.update_idletasks()
        scale = self._widget_scaling_factor()
        tip_w = max(1, int(round(tip.winfo_reqwidth() / scale)))
        tip_h = max(1, int(round(tip.winfo_reqheight() / scale)))
        tip.configure(width=tip_w, height=tip_h)

        root_x = self.winfo_rootx()
        root_y = self.winfo_rooty()
        x = int(round((name_canvas.winfo_rootx() - root_x) / scale))
        y = int(round((name_canvas.winfo_rooty() - root_y) / scale))
        win_w = max(1, int(round(self.winfo_width() / scale)))
        win_h = max(1, int(round(self.winfo_height() / scale)))
        x = max(DOC_SIDEBAR_MARGIN, min(x, win_w - tip_w - DOC_SIDEBAR_MARGIN))
        y = max(DOC_SIDEBAR_MARGIN, min(y, win_h - tip_h - DOC_SIDEBAR_MARGIN))

        tip.place(x=x, y=y, anchor="nw")
        tip.lift()
        self._doc_name_tip = tip
        self._doc_name_tip_path = path
        self._set_doc_row_hover(path, False)

    def _on_doc_name_tip_leave(self, path: str):
        row = self._doc_rows.get(path)
        if row is not None:
            frame = row["frame"]
            try:
                under = frame.winfo_containing(
                    frame.winfo_pointerx(), frame.winfo_pointery()
                )
            except Exception:  # noqa: BLE001
                under = None
            if under is not None and self._is_descendant(under, frame):
                return
        tip = self._doc_name_tip
        if tip is not None:
            try:
                under = tip.winfo_containing(tip.winfo_pointerx(), tip.winfo_pointery())
            except Exception:  # noqa: BLE001
                under = None
            if under is not None and self._is_descendant(under, tip):
                return
        if self._doc_name_tip_path == path:
            self._hide_doc_name_tip()
        self._clear_doc_row_hovers()

    def _sync_doc_row_surfaces(self, path: str, surface) -> None:
        row = self._doc_rows.get(path)
        if row is None:
            return
        for key in ("mark", "name"):
            canvas = row.get(key)
            if canvas is not None:
                self._set_doc_canvas_surface(canvas, surface)

    def _on_doc_row_press(self, path: str, _event=None):
        if self._autofill_busy:
            return "break"
        self._select_document(path)
        return "break"

    def _bind_doc_row_interactions(self, widget, path: str):
        widget.bind("<ButtonPress-1>", lambda _e, p=path: self._on_doc_row_press(p))
        widget.bind("<B1-Motion>", lambda _e: "break")
        widget.bind("<ButtonRelease-1>", lambda _e: "break")
        widget.bind("<Enter>", lambda _e, p=path: self._hover_doc_row(p, True))
        widget.bind("<Leave>", lambda e, p=path: self._on_doc_row_leave(e, p))
        try:
            widget.configure(cursor="hand2")
        except Exception:  # noqa: BLE001
            pass
        for child in widget.winfo_children():
            self._bind_doc_row_interactions(child, path)

    def _on_doc_row_leave(self, event, path: str):
        row = self._doc_rows.get(path)
        if row is None:
            return
        frame = row["frame"]
        under = frame.winfo_containing(event.x_root, event.y_root)
        if under is not None and self._is_descendant(under, frame):
            return
        tip = self._doc_name_tip
        if tip is not None and self._doc_name_tip_path == path:
            tip_under = tip.winfo_containing(event.x_root, event.y_root)
            if tip_under is not None and self._is_descendant(tip_under, tip):
                self._set_doc_row_hover(path, False)
                return
        self._hover_doc_row(path, False)

    def _set_doc_row_hover(self, path: str, hovering: bool):
        if self._autofill_busy or self._autofill_ui_chrome_locked:
            return
        if path == self._selected_path:
            if not hovering and self._doc_hover_path == path:
                self._doc_hover_path = None
            return
        row = self._doc_rows.get(path)
        if row is None:
            return
        if hovering:
            self._clear_doc_row_hovers(keep=path)
            row["frame"].configure(fg_color=TILE_BG_HOVER)
            self._sync_doc_row_surfaces(path, TILE_BG_HOVER)
            self._doc_hover_path = path
            return
        row["frame"].configure(fg_color="transparent")
        self._sync_doc_row_surfaces(path, "transparent")
        if self._doc_hover_path == path:
            self._doc_hover_path = None

    def _hover_doc_row(self, path: str, entering: bool):
        if self._autofill_busy:
            return
        if entering:
            if self._doc_name_tip_path and self._doc_name_tip_path != path:
                self._hide_doc_name_tip()
            self._clear_doc_row_hovers(keep=path)
            self._show_doc_name_tip(path)
            if self._doc_name_tip_path == path:
                self._set_doc_row_hover(path, False)
            else:
                self._set_doc_row_hover(path, True)
            return
        if self._doc_name_tip_path == path:
            self._hide_doc_name_tip()
        self._set_doc_row_hover(path, False)
        if self._doc_hover_path is None:
            self._clear_doc_row_hovers()

    def _doc_mark_text_color(self, path: str, *, selected: bool):
        """Numbers are muted (~50% opacity); status marks stay full strength."""
        kind = self._doc_status_kind(path)
        if kind and self._status_dots:
            return DOC_STATUS_DOT_OK if kind == "ok" else DOC_STATUS_DOT_BAD
        if kind:
            return DOC_ROW_ACTIVE_TEXT if selected else ("gray10", "gray90")
        return self._accent_doc_mark_number_active() if selected else DOC_MARK_NUMBER_COLOR

    def _doc_mark_font_size(self, path: str) -> int:
        # Keep body size when number + dot share the cell so both stay readable.
        if self._status_dots and self._doc_status_kind(path):
            return DOC_STATUS_DOT_SIZE
        return FONT_BODY

    def _style_doc_row(self, path: str, *, selected: bool):
        row = self._doc_rows.get(path)
        if row is None:
            return
        mark_color = self._doc_mark_text_color(path, selected=selected)
        mark_size = self._doc_mark_font_size(path)
        mark_text = self._doc_leading_label(path, row.get("index", 0))
        name_text = row.get("full_name") or self._doc_name_label(path, row.get("index", 0))
        if self._autofill_busy or self._autofill_ui_chrome_locked:
            # Keep the live autofill target visible; mute the rest.
            if selected:
                row["frame"].configure(fg_color=self._accent_doc_row_active())
                self._set_doc_canvas_text(
                    row["mark"], mark_text, fill=mark_color, size=mark_size
                )
                self._set_doc_canvas_text(
                    row["name"], name_text, fill=DOC_ROW_ACTIVE_TEXT, size=FONT_BODY
                )
                self._sync_doc_row_surfaces(path, self._accent_doc_row_active())
            else:
                row["frame"].configure(fg_color="transparent")
                self._set_doc_canvas_text(
                    row["mark"], mark_text, fill=UI_LOCK_DOC_TEXT, size=mark_size
                )
                self._set_doc_canvas_text(
                    row["name"], name_text, fill=UI_LOCK_DOC_TEXT, size=FONT_BODY
                )
                self._sync_doc_row_surfaces(path, "transparent")
            return
        if selected:
            row["frame"].configure(fg_color=self._accent_doc_row_active())
            self._set_doc_canvas_text(row["mark"], mark_text, fill=mark_color, size=mark_size)
            self._set_doc_canvas_text(
                row["name"], name_text, fill=DOC_ROW_ACTIVE_TEXT, size=FONT_BODY
            )
            self._sync_doc_row_surfaces(path, self._accent_doc_row_active())
        else:
            row["frame"].configure(fg_color="transparent")
            self._set_doc_canvas_text(row["mark"], mark_text, fill=mark_color, size=mark_size)
            self._set_doc_canvas_text(
                row["name"], name_text, fill=("gray10", "gray90"), size=FONT_BODY
            )
            self._sync_doc_row_surfaces(path, "transparent")

    def _rebuild_doc_list(self):
        self._hide_doc_name_tip()
        self._doc_hover_path = None
        for child in self.doc_list_frame.winfo_children():
            child.destroy()
        self._doc_rows.clear()
        # Drop any leftover scroll offset from the previous list.
        self.doc_list_frame._parent_canvas.yview_moveto(0)

        if not self._imported_files:
            self.doc_list_empty = customtkinter.CTkLabel(
                self.doc_list_frame,
                text="选择文件夹后，\nPDF 将显示在这里",
                text_color="gray50",
                justify="center",
            )
            self.doc_list_empty.grid(row=0, column=0, sticky="ew", pady=20)
            self._bind_doc_list_mousewheel(self.doc_list_empty)
            self._schedule_doc_list_scrollbar_sync()
            return

        parsed = [p for p in self._imported_files if self._cert_is_parsed(p)]
        pending = [p for p in self._imported_files if not self._cert_is_parsed(p)]
        row = 0
        index = 0

        def add_section(title: str):
            nonlocal row
            label = customtkinter.CTkLabel(
                self.doc_list_frame,
                text=title,
                anchor="w",
                font=customtkinter.CTkFont(size=FONT_META, weight="bold"),
                text_color="gray60",
            )
            label.grid(row=row, column=0, sticky="ew", pady=(8 if row else 2, 4))
            self._bind_doc_list_mousewheel(label)
            row += 1

        def add_docs(paths: list[str]):
            nonlocal row, index
            for path in paths:
                frame = customtkinter.CTkFrame(
                    self.doc_list_frame,
                    height=DOC_ROW_HEIGHT,
                    corner_radius=UI_RADIUS,
                    fg_color="transparent",
                    border_width=0,
                )
                frame.grid(row=row, column=0, sticky="ew", pady=2)
                frame.grid_propagate(False)
                frame.grid_columnconfigure(0, weight=0, minsize=DOC_MARK_COL_WIDTH)
                frame.grid_columnconfigure(1, weight=1)
                frame.grid_rowconfigure(0, weight=1)

                full_name = self._doc_name_label(path, index)
                mark, mark_id = self._make_doc_text_canvas(
                    frame,
                    text=self._doc_leading_label(path, index),
                    fill=self._doc_mark_text_color(path, selected=False),
                    size=self._doc_mark_font_size(path),
                    width=DOC_MARK_COL_WIDTH,
                    anchor="center",
                )
                # Inset so CTkFrame corner_radius remains visible on hover/select.
                mark.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=4)

                name, name_id = self._make_doc_text_canvas(
                    frame,
                    text=full_name,
                    fill=("gray10", "gray90"),
                    size=FONT_BODY,
                    anchor="w",
                )
                name.grid(row=0, column=1, sticky="nsew", padx=(4, 10), pady=4)

                self._doc_rows[path] = {
                    "frame": frame,
                    "mark": mark,
                    "mark_id": mark_id,
                    "name": name,
                    "name_id": name_id,
                    "full_name": full_name,
                    "index": index,
                }
                self._bind_doc_row_interactions(frame, path)
                self._bind_doc_list_mousewheel(frame)
                row += 1
                index += 1

        if parsed:
            add_section(f"已解析（{len(parsed)}）")
            add_docs(parsed)
        if pending:
            title = (
                f"待OCR（{len(pending)}）"
                if self._ocr_enabled
                else f"解析失败（{len(pending)}）"
            )
            add_section(title)
            add_docs(pending)

        self._highlight_selected_doc()
        self._schedule_doc_list_scrollbar_sync()

    def _refresh_doc_list_marks(self):
        for i, path in enumerate(self._imported_files):
            row = self._doc_rows.get(path)
            if row is None:
                continue
            full_name = self._doc_name_label(path, i)
            row["full_name"] = full_name
            row["index"] = i
        self._highlight_selected_doc()

    def _highlight_selected_doc(self):
        for path in self._doc_rows:
            self._style_doc_row(path, selected=(path == self._selected_path))

    def _select_document(self, path: str, *, sync_preview: bool = True):
        if path not in self._imported_files:
            return
        if self._autofill_busy:
            return
        self._hide_doc_name_tip()
        self._clear_doc_row_hovers()
        # From settings, resume the step that was open before settings.
        self._leave_settings_if_open()
        if self._selected_path and self._selected_path != path:
            self._save_fields_before_navigate()
        self._selected_path = path
        self._current_cert_index = self._imported_files.index(path)
        self._sync_cert_index_to_list()
        self._show_parse_result(path)
        if hasattr(self, "field_entries"):
            self._load_approve_fields_for_current()
            self._update_review_cert_status()
        if not self._autofill_busy:
            self._load_automate_fields(path)
        if self._current_step == "parse_rules":
            self._refresh_parse_rules_panel()
        if sync_preview:
            self._sync_pdf_preview()

    @staticmethod
    def _reset_ctk_entry(entry: customtkinter.CTkEntry) -> None:
        """Clear a CTkEntry and restore its placeholder text.

        CTkEntry.delete() removes visible placeholder characters while leaving
        `_placeholder_text_active` True, so a second clear can leave the field
        blank with no “请输入…” hint. Always deactivate first, then re-activate.
        """
        if getattr(entry, "_placeholder_text_active", False):
            entry._deactivate_placeholder()
        else:
            entry._entry.delete(0, "end")
        # Force placeholder even if the widget still thinks it is focused.
        entry._is_focused = False
        entry._activate_placeholder()

    def _clear_extract_fields_display(self):
        # Blur entries first so placeholder restore is not fighting FocusIn state.
        self.focus_set()
        if hasattr(self, "field_entries"):
            self._clear_approve_fields()
        if hasattr(self, "extract_errors_label"):
            self.extract_errors_label.configure(text="")

    def _show_parse_result(self, path: str):
        result = self._parse_results.get(path)
        if not hasattr(self, "extract_errors_label"):
            return
        if result is None:
            self.extract_errors_label.configure(text="⚠ 尚未解析此文档")
            return

        if result.errors:
            # Some parse failures are actionable via OCR; hide that specific hint to reduce noise.
            filtered_errors = [
                e for e in result.errors if "封面无嵌入文本" not in str(e)
            ]
            if filtered_errors:
                self.extract_errors_label.configure(
                    text="⚠ " + "\n".join(filtered_errors)
                )
            else:
                self.extract_errors_label.configure(text="")
        else:
            self.extract_errors_label.configure(text="")

    def _get_field_widget_value(self, widget) -> str:
        if isinstance(widget, customtkinter.CTkTextbox):
            return widget.get("1.0", "end-1c").strip()
        return widget.get().strip()

    def _set_field_widget_value(self, widget, value: str):
        if isinstance(widget, customtkinter.CTkTextbox):
            widget.delete("1.0", "end")
            if value:
                widget.insert("1.0", value)
            return
        if value:
            if getattr(widget, "_placeholder_text_active", False):
                widget._deactivate_placeholder()
            else:
                widget.delete(0, "end")
            widget.insert(0, value)
        else:
            self._reset_ctk_entry(widget)

    def _clear_approve_fields(self):
        self.focus_set()
        if not hasattr(self, "field_entries"):
            return
        self._set_review_fields_locked(False)
        for widget in self.field_entries.values():
            self._set_field_widget_value(widget, "")
        self._update_review_fields_state()

    def _review_fields_locked(self) -> bool:
        # Locked until failed certs are cleared, then also when approved/removed.
        if getattr(self, "_workflow_phase", "extract") != "review":
            return True
        path = self._current_cert_path()
        if path is None:
            return False
        return path in self._autofill_queue or path in self._removed_paths

    def _set_field_widgets_locked(
        self,
        widgets: dict[str, customtkinter.CTkEntry | customtkinter.CTkTextbox] | None,
        locked: bool,
    ) -> None:
        if not widgets:
            return
        if locked:
            try:
                self.focus_set()
            except Exception:  # noqa: BLE001
                pass
        state = "disabled" if locked else "normal"
        fg_color = FIELD_FG_COLOR_DISABLED if locked else FIELD_FG_COLOR
        # Keep parsed values readable while locked (grey field, normal text).
        text_color = FIELD_TEXT_COLOR
        for widget in widgets.values():
            try:
                widget.configure(
                    state=state,
                    fg_color=fg_color,
                    text_color=text_color,
                    border_color=fg_color,
                )
            except Exception:  # noqa: BLE001
                pass

    def _load_fields_into_widgets(
        self,
        widgets: dict[str, customtkinter.CTkEntry | customtkinter.CTkTextbox] | None,
        *,
        path: str | None = None,
        fields: CertificateFields | None = None,
        keep_locked: bool = False,
    ) -> None:
        if not widgets:
            return
        self._set_field_widgets_locked(widgets, False)
        for widget in widgets.values():
            self._set_field_widget_value(widget, "")
        source = fields
        if source is None and path:
            result = self._parse_results.get(path)
            source = result.fields if result is not None else None
        if source is not None:
            for key, widget in widgets.items():
                value = getattr(source, key, "") or ""
                if not value and key == "result_info":
                    value = DEFAULT_RESULT_INFO
                self._set_field_widget_value(widget, value)
        self._set_field_widgets_locked(widgets, keep_locked)

    def _load_automate_fields(
        self,
        path: str | None = None,
        fields: CertificateFields | None = None,
    ) -> None:
        widgets = getattr(self, "automate_field_entries", None)
        self._load_fields_into_widgets(
            widgets, path=path, fields=fields, keep_locked=True
        )

    def _focus_certificate_for_autofill(
        self,
        path: str | None,
        fields: CertificateFields | None = None,
        *,
        index: int | None = None,
        total: int | None = None,
        pending: bool = False,
    ) -> None:
        """Show this certificate on 自动化 (fields + list + PDF) during a run."""
        del index, total, pending  # kept for call-site compatibility
        resolved = str(path) if path else ""
        if resolved and resolved in self._imported_files:
            self._selected_path = resolved
            self._current_cert_index = self._imported_files.index(resolved)
            self._highlight_selected_doc()
        self._load_automate_fields(resolved or None, fields)
        if resolved and self._pdf_preview_enabled:
            try:
                self._pdf_preview.update_pdf(resolved)
            except Exception:  # noqa: BLE001
                pass

    def _on_autofill_item(self, index: int, total: int, item: AutofillItem) -> None:
        path = str(item.pdf_path) if item.pdf_path else None
        self._focus_certificate_for_autofill(
            path, item.fields, index=index, total=total
        )

    def _set_review_fields_locked(self, locked: bool):
        self._set_field_widgets_locked(getattr(self, "field_entries", None), locked)

    def _update_review_fields_state(self):
        self._set_review_fields_locked(self._review_fields_locked())

    def _current_cert_path(self) -> str | None:
        if not self._imported_files:
            return None
        return self._imported_files[self._current_cert_index]

    def _read_fields_from_entries(self, base: CertificateFields | None = None) -> CertificateFields:
        fields = CertificateFields() if base is None else CertificateFields(
            name=base.name,
            serial_num=base.serial_num,
            model=base.model,
            measurement_unit=base.measurement_unit,
            measurement_date=base.measurement_date,
            measurement_type=base.measurement_type,
            certificate_no=base.certificate_no,
            client_name=base.client_name,
            manufacturer=base.manufacturer,
            due_date=base.due_date,
            issue_date=base.issue_date,
            result_info=base.result_info,
        )
        for key, _ in REVIEW_DISPLAY_FIELDS:
            if key in self.field_entries:
                setattr(fields, key, self._get_field_widget_value(self.field_entries[key]))
        return fields

    def _empty_review_field_labels(self) -> list[str]:
        if not hasattr(self, "field_entries"):
            return []
        empty: list[str] = []
        for key, label in REVIEW_DISPLAY_FIELDS:
            widget = self.field_entries.get(key)
            if widget is None:
                continue
            if not self._get_field_widget_value(widget):
                empty.append(label)
        return empty

    def _save_current_fields_to_result(self) -> ParseResult | None:
        path = self._current_cert_path()
        if path is None:
            return None
        existing = self._parse_results.get(path)
        base = existing.fields if existing else None
        fields = self._read_fields_from_entries(base)
        if existing is None:
            result = ParseResult(
                source_path=path,
                page_count=0,
                raw_text="",
                lines=[],
                fields=fields,
            )
        else:
            result = ParseResult(
                source_path=existing.source_path,
                page_count=existing.page_count,
                raw_text=existing.raw_text,
                lines=existing.lines,
                fields=fields,
                method=existing.method,
                errors=list(existing.errors),
            )
        self._parse_results[path] = result
        return result

    def _load_approve_fields_for_current(self):
        path = self._current_cert_path()
        self._load_fields_into_widgets(
            getattr(self, "field_entries", None),
            path=path,
            keep_locked=self._review_fields_locked(),
        )
        self._update_review_cert_status()

    def _update_review_cert_status(self):
        self._update_approve_toggle_button()
        self._update_remove_toggle_button()
        self._update_remove_failed_button()
        self._update_review_fields_state()

    def _update_remove_failed_button(self):
        """Grey out 移出失败证书 after it has been executed for this folder."""
        if not hasattr(self, "remove_failed_button"):
            return
        if self._autofill_busy or self._autofill_ui_chrome_locked:
            return
        done = getattr(self, "_workflow_phase", "extract") == "review"
        if done:
            self.remove_failed_button.configure(
                state="disabled",
                fg_color=UI_LOCK_BTN_FG,
                hover_color=UI_LOCK_BTN_FG,
                text_color=UI_LOCK_BTN_TEXT,
            )
        else:
            self.remove_failed_button.configure(
                state="normal",
                fg_color=DANGER_BTN_FG,
                hover_color=DANGER_BTN_HOVER,
                text_color=DANGER_BTN_TEXT,
            )

    def _update_approve_toggle_button(self):
        if not hasattr(self, "approve_toggle_button"):
            return
        if self._autofill_busy or self._autofill_ui_chrome_locked:
            return
        ready = getattr(self, "_workflow_phase", "extract") == "review"
        path = self._current_cert_path()
        if not ready:
            self.approve_toggle_button.configure(
                text="批准",
                state="disabled",
                fg_color=UI_LOCK_BTN_FG,
                hover_color=UI_LOCK_BTN_FG,
                text_color=UI_LOCK_BTN_TEXT,
            )
            return
        if path is not None and path in self._autofill_queue:
            self.approve_toggle_button.configure(
                text="撤销批准",
                fg_color=SECONDARY_BTN_FG,
                hover_color=SECONDARY_BTN_HOVER,
                text_color=SECONDARY_BTN_TEXT,
                state="normal",
            )
        else:
            self.approve_toggle_button.configure(
                text="批准",
                fg_color=SUCCESS_BTN_FG,
                hover_color=SUCCESS_BTN_HOVER,
                text_color=SUCCESS_BTN_TEXT,
                state="normal",
            )

    def _update_remove_toggle_button(self):
        if not hasattr(self, "remove_toggle_button"):
            return
        if self._autofill_busy or self._autofill_ui_chrome_locked:
            return
        ready = getattr(self, "_workflow_phase", "extract") == "review"
        path = self._current_cert_path()
        if not ready:
            self.remove_toggle_button.configure(
                text="移除",
                state="disabled",
                fg_color=UI_LOCK_BTN_FG,
                hover_color=UI_LOCK_BTN_FG,
                text_color=UI_LOCK_BTN_TEXT,
            )
            return
        if path is not None and path in self._removed_paths:
            self.remove_toggle_button.configure(
                text="撤销移除",
                fg_color=SECONDARY_BTN_FG,
                hover_color=SECONDARY_BTN_HOVER,
                text_color=SECONDARY_BTN_TEXT,
                state="normal",
            )
        else:
            self.remove_toggle_button.configure(
                text="移除",
                fg_color=DANGER_BTN_FG,
                hover_color=DANGER_BTN_HOVER,
                text_color=DANGER_BTN_TEXT,
                state="normal",
            )

    def _update_autofill_button(self):
        if hasattr(self, "autofill_button"):
            n = len(self._autofill_queue)
            if not self._autofill_busy:
                self.autofill_button.configure(text=f"导出并自动填写 ({n})")
            if hasattr(self, "export_excel_button"):
                self.export_excel_button.configure(text=f"导出 Excel ({n})")
        # Unlock/lock 自动化 when the approve queue gains or loses its first item.
        if (
            getattr(self, "_current_step", None) == "automate"
            and not self._automate_step_unlocked()
            and not self._autofill_busy
        ):
            self.show_step("extract")
        elif hasattr(self, "step_tiles"):
            self._update_step_tiles(self._current_step)

    def _show_autofill_run_controls(self):
        """Swap the autofill CTA for 暂停 / 退出 while a run is active."""
        if hasattr(self, "autofill_button"):
            self.autofill_button.grid_remove()
        self._update_custom_dir_autofill_button()
        if hasattr(self, "autofill_controls_frame"):
            self.autofill_pause_button.configure(
                text="暂停",
                state="normal",
                command=self._on_autofill_pause_toggle,
            )
            self.autofill_exit_button.configure(
                text="退出",
                state="normal",
                command=self._on_autofill_exit,
            )
            self.autofill_controls_frame.grid(
                row=1, column=0, sticky="ew", pady=(10, 0)
            )
        self._lock_ui_for_autofill()

    def _restore_autofill_button(self):
        """Restore the original autofill CTA after a run finishes or exits."""
        self._unlock_ui_after_autofill()
        self._autofill_control = None
        self._autofill_exit_confirming = False
        self._autofill_was_paused_before_exit = False
        if hasattr(self, "autofill_controls_frame"):
            self.autofill_controls_frame.grid_remove()
        if hasattr(self, "autofill_pause_button"):
            try:
                self.autofill_pause_button.configure(
                    text="暂停",
                    state="normal",
                    command=self._on_autofill_pause_toggle,
                )
            except Exception:  # noqa: BLE001
                pass
        if hasattr(self, "autofill_exit_button"):
            try:
                self.autofill_exit_button.configure(
                    text="退出",
                    state="normal",
                    command=self._on_autofill_exit,
                )
            except Exception:  # noqa: BLE001
                pass
        if hasattr(self, "autofill_button"):
            self.autofill_button.configure(state="normal")
            self.autofill_button.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        self._update_custom_dir_autofill_button()
        if hasattr(self, "export_excel_button"):
            self.export_excel_button.configure(state="normal")
        self._update_autofill_button()
        if hasattr(self, "field_entries"):
            self._load_approve_fields_for_current()

    def _autofill_interactive_types(self) -> tuple:
        types = [
            customtkinter.CTkButton,
            customtkinter.CTkEntry,
            customtkinter.CTkTextbox,
            customtkinter.CTkSwitch,
            customtkinter.CTkCheckBox,
            customtkinter.CTkSlider,
            customtkinter.CTkComboBox,
            customtkinter.CTkOptionMenu,
            customtkinter.CTkSegmentedButton,
        ]
        return tuple(types)

    def _widget_config_snapshot(self, widget, keys: tuple[str, ...]) -> dict:
        snap: dict = {}
        for key in keys:
            try:
                snap[key] = widget.cget(key)
            except Exception:  # noqa: BLE001
                pass
        return snap

    def _set_widget_tree_cursor(self, widget, cursor: str):
        try:
            widget.configure(cursor=cursor)
        except Exception:  # noqa: BLE001
            pass
        try:
            children = widget.winfo_children()
        except Exception:  # noqa: BLE001
            return
        for child in children:
            self._set_widget_tree_cursor(child, cursor)

    def _lock_step_tiles_chrome(self):
        """Mute module tiles (fill, outline, badge, label) while autofill runs."""
        for key, tile in getattr(self, "step_tiles", {}).items():
            try:
                tile.configure(
                    fg_color=UI_LOCK_TILE_FG,
                    border_color=UI_LOCK_TILE_FG,
                    border_width=0,
                )
            except Exception:  # noqa: BLE001
                pass
            for child in tile.winfo_children():
                try:
                    if isinstance(child, customtkinter.CTkFrame):
                        child.configure(fg_color=UI_LOCK_BADGE_FG)
                        for sub in child.winfo_children():
                            if isinstance(sub, customtkinter.CTkLabel):
                                sub.configure(text_color=("#ffffff", "#ffffff"))
                    elif isinstance(child, customtkinter.CTkLabel):
                        child.configure(text_color=UI_LOCK_LABEL)
                except Exception:  # noqa: BLE001
                    pass
            self._set_widget_tree_cursor(tile, "")

    def _unlock_step_tiles_chrome(self):
        for key, tile in getattr(self, "step_tiles", {}).items():
            for child in tile.winfo_children():
                try:
                    if isinstance(child, customtkinter.CTkFrame):
                        child.configure(fg_color=self._accent_outline())
                        for sub in child.winfo_children():
                            if isinstance(sub, customtkinter.CTkLabel):
                                sub.configure(text_color=("#ffffff", "#ffffff"))
                    elif isinstance(child, customtkinter.CTkLabel):
                        child.configure(text_color=SECONDARY_BTN_TEXT)
                except Exception:  # noqa: BLE001
                    pass
            self._set_widget_tree_cursor(tile, "")
        self._update_step_tiles(self._current_step)

    def _keep_autofill_run_controls_active(self):
        """Pause/Exit must stay full-color and clickable during the UI lock."""
        if hasattr(self, "autofill_pause_button"):
            try:
                self.autofill_pause_button.configure(
                    state="normal",
                    fg_color=SECONDARY_BTN_FG,
                    hover_color=SECONDARY_BTN_HOVER,
                    text_color=SECONDARY_BTN_TEXT,
                    border_width=0,
                )
                self._style_primary_action_button(self.autofill_pause_button)
            except Exception:  # noqa: BLE001
                pass
        if hasattr(self, "autofill_exit_button"):
            try:
                self.autofill_exit_button.configure(
                    state="normal",
                    fg_color=DANGER_BTN_FG,
                    hover_color=DANGER_BTN_HOVER,
                    text_color=DANGER_BTN_TEXT,
                    border_width=0,
                )
                self._style_primary_action_button(self.autofill_exit_button)
            except Exception:  # noqa: BLE001
                pass

    def _lock_ui_for_autofill(self):
        """Grey out / disable interactables; keep 暂停 / 退出 and toast actions enabled."""
        self._unlock_ui_after_autofill(reapply_intent=False)
        try:
            self.focus_set()
        except Exception:  # noqa: BLE001
            pass

        allow = set()
        for name in ("autofill_pause_button", "autofill_exit_button"):
            widget = getattr(self, name, None)
            if widget is not None:
                allow.add(widget)

        # Same locked look as validated / removed certificates.
        self._set_review_fields_locked(True)

        field_widgets = set()
        if hasattr(self, "field_entries"):
            field_widgets = set(self.field_entries.values())
        if hasattr(self, "automate_field_entries"):
            field_widgets |= set(self.automate_field_entries.values())

        toast_root = getattr(self, "_toast_host", None)
        log_frame = getattr(self, "_autofill_log_frame", None)
        log_bubble = getattr(self, "_autofill_log_bubble", None)

        def _under_protected(widget) -> bool:
            """Toast / terminal controls must stay dismissible during autofill."""
            cur = widget
            while cur is not None:
                if cur is toast_root or cur is log_frame or cur is log_bubble:
                    return True
                try:
                    cur = cur.master
                except Exception:  # noqa: BLE001
                    break
            return False

        disabled: list = []
        interactive = self._autofill_interactive_types()
        button_keys = (
            "state",
            "fg_color",
            "hover_color",
            "text_color",
            "border_color",
            "border_width",
        )
        plain_keys = ("state",)

        def walk(widget):
            for child in widget.winfo_children():
                walk(child)
            if widget in allow or widget in field_widgets:
                return
            if _under_protected(widget):
                return
            if not isinstance(widget, interactive):
                return
            try:
                current = str(widget.cget("state"))
            except Exception:  # noqa: BLE001
                current = "normal"
            if current == "disabled":
                # Still force-grey colored buttons that were already disabled
                # with a vivid fg_color (批准 / 移除 after style refresh).
                if not isinstance(widget, customtkinter.CTkButton):
                    return
            try:
                if isinstance(widget, customtkinter.CTkButton):
                    snap = self._widget_config_snapshot(widget, button_keys)
                    widget.configure(
                        state="disabled",
                        fg_color=UI_LOCK_BTN_FG,
                        hover_color=UI_LOCK_BTN_FG,
                        text_color=UI_LOCK_BTN_TEXT,
                        border_width=0,
                    )
                else:
                    snap = self._widget_config_snapshot(widget, plain_keys)
                    widget.configure(state="disabled")
                disabled.append((widget, snap))
            except Exception:  # noqa: BLE001
                pass

        walk(self)
        self._autofill_disabled_widgets = disabled

        self._autofill_ui_chrome_locked = True
        self._lock_step_tiles_chrome()
        self._hide_doc_name_tip()
        self._clear_doc_row_hovers()
        self._highlight_selected_doc()
        if hasattr(self, "doc_list_frame"):
            try:
                self.doc_list_frame._scrollbar.grid_remove()
            except Exception:  # noqa: BLE001
                pass
            self._set_widget_tree_cursor(self.doc_list_frame, "")

        self._keep_autofill_run_controls_active()

    def _unlock_ui_after_autofill(self, *, reapply_intent: bool = True):
        for item in self._autofill_disabled_widgets:
            if isinstance(item, tuple):
                widget, snap = item
            else:
                widget, snap = item, {"state": "normal"}
            try:
                if widget.winfo_exists() and snap:
                    widget.configure(**snap)
            except Exception:  # noqa: BLE001
                pass
        self._autofill_disabled_widgets = []

        was_chrome_locked = self._autofill_ui_chrome_locked
        self._autofill_ui_chrome_locked = False
        if was_chrome_locked:
            self._unlock_step_tiles_chrome()
            if hasattr(self, "doc_list_frame"):
                self._set_widget_tree_cursor(self.doc_list_frame, "hand2")
            self._highlight_selected_doc()
            self._schedule_doc_list_scrollbar_sync()

        if not reapply_intent:
            return

        # Re-apply intentional disabled states / colors.
        try:
            self._update_review_fields_state()
        except Exception:  # noqa: BLE001
            pass
        try:
            self._set_field_widgets_locked(
                getattr(self, "automate_field_entries", None), True
            )
        except Exception:  # noqa: BLE001
            pass
        try:
            self._update_extract_ocr_ui()
        except Exception:  # noqa: BLE001
            pass
        try:
            self._update_review_cert_status()
        except Exception:  # noqa: BLE001
            pass
        try:
            self._update_settings_button(self._current_step in SETTINGS_SUBPAGES)
        except Exception:  # noqa: BLE001
            pass
        try:
            self._restyle_primary_action_buttons()
        except Exception:  # noqa: BLE001
            pass

    def _on_autofill_pause_toggle(self):
        control = self._autofill_control
        if control is None or not self._autofill_busy:
            return
        if control.is_paused():
            control.resume()
            try:
                self.autofill_pause_button.configure(text="暂停", state="normal")
            except Exception:  # noqa: BLE001
                pass
            self.append_autofill_log("已继续自动填写")
            self.set_status("自动填写已继续")
        else:
            control.pause()
            try:
                self.autofill_pause_button.configure(text="继续", state="normal")
            except Exception:  # noqa: BLE001
                pass
            self.append_autofill_log("已暂停自动填写")
            self.set_status("自动填写已暂停")

    def _on_autofill_exit(self):
        """Pause first, then ask for confirmation; second click exits immediately."""
        control = self._autofill_control
        if control is None or not self._autofill_busy:
            return
        if control.cancelled():
            return
        # Second click (or click while toast is up) → exit now, no more waiting.
        if self._autofill_exit_confirming:
            self._confirm_autofill_exit()
            return

        self._autofill_was_paused_before_exit = control.is_paused()
        if not control.is_paused():
            control.pause()
            try:
                self.autofill_pause_button.configure(text="继续")
            except Exception:  # noqa: BLE001
                pass
            self.append_autofill_log("已暂停（等待确认退出）")

        self._autofill_exit_confirming = True
        try:
            self.autofill_exit_button.configure(text="立即退出", state="normal")
        except Exception:  # noqa: BLE001
            pass
        self.set_status("再点一次「立即退出」可马上结束；或等待 10 秒自动退出")
        self.append_autofill_log("退出确认 — 再点「立即退出」马上结束；点取消可继续")
        self.show_toast(
            "自动填写已暂停。\n再点「立即退出」马上结束。\n点「取消」继续填写；10 秒后也会自动退出。",
            title="退出自动填写",
            action_text="立即退出",
            undo_text="取消",
            on_undo=self._cancel_autofill_exit,
            on_complete=self._confirm_autofill_exit,
            duration_ms=AUTOFILL_EXIT_WARN_MS,
            style="danger",
        )

    def _cancel_autofill_exit(self):
        """User declined exit — resume unless they were already paused."""
        self._autofill_exit_confirming = False
        control = self._autofill_control
        if control is None or not self._autofill_busy or control.cancelled():
            return
        try:
            self.autofill_exit_button.configure(text="退出", state="normal")
        except Exception:  # noqa: BLE001
            pass
        if self._autofill_was_paused_before_exit:
            try:
                self.autofill_pause_button.configure(text="继续")
            except Exception:  # noqa: BLE001
                pass
            self.append_autofill_log("已取消退出，保持暂停")
            self.set_status("自动填写已暂停")
            return
        control.resume()
        try:
            self.autofill_pause_button.configure(text="暂停")
        except Exception:  # noqa: BLE001
            pass
        self.append_autofill_log("已取消退出，继续自动填写")
        self.set_status("自动填写已继续")

    def _confirm_autofill_exit(self):
        """Confirmed exit (button, second click, or toast countdown)."""
        if not self._autofill_busy:
            return
        self._autofill_exit_confirming = False
        control = self._autofill_control
        if control is None:
            return
        if control.cancelled():
            # Already exiting — still tear down UI/terminal.
            self.close_autofill_log()
            return
        control.request_exit()
        try:
            self.autofill_pause_button.configure(state="disabled")
            self.autofill_exit_button.configure(text="退出中…", state="disabled")
        except Exception:  # noqa: BLE001
            pass
        self.append_autofill_log("正在退出自动填写…", error=True)
        self.set_status("正在退出自动填写…")
        # Close the terminal immediately — browser stays open for manual review.
        self.close_autofill_log()
        # Safety net: if the worker never returns, restore controls anyway.
        self.after(2500, self._autofill_exit_watchdog)

    def _autofill_exit_watchdog(self):
        if not self._autofill_busy:
            return
        control = self._autofill_control
        if control is None or not control.cancelled():
            return
        self._autofill_busy = False
        self._autofill_exit_confirming = False
        self.close_autofill_log()
        self._restore_autofill_button()
        self.set_status("已停止自动填写（浏览器保持打开）")
        self.show_toast(
            "已停止自动填写。浏览器保持打开。",
            title="导出并自动填写",
            duration_ms=TOAST_SUCCESS_MS,
        )

    def _save_fields_before_navigate(self):
        if (
            hasattr(self, "field_entries")
            and not self._review_fields_locked()
        ):
            self._save_current_fields_to_result()

    def _sync_cert_index_to_list(self):
        path = self._current_cert_path()
        if path is None:
            return
        self._selected_path = path
        self._highlight_selected_doc()

    def _leave_settings_if_open(self, key: str | None = None):
        """If settings is showing, leave to `key` or the step that opened settings.

        Parse-rules and more-options stay open so nested settings are not dismissed.
        """
        if self._current_step in {"parse_rules", "more_options"}:
            return
        if self._current_step != "settings":
            return
        target = key or getattr(self, "_step_before_settings", None) or "extract"
        if target in SETTINGS_SUBPAGES:
            target = "extract"
        self.show_step(target)

    def _on_open_settings(self):
        if self._autofill_busy:
            self.set_status("自动填写进行中，请先暂停或退出…")
            return
        if self._current_step not in SETTINGS_SUBPAGES:
            self._step_before_settings = self._current_step
        self.show_step("settings")

    def _save_eams_login_info(self):
        username, password = self._eams_credentials()
        if not username:
            self.set_status("请先填写 EAMS 用户名")
            self.show_toast(
                "请先填写 EAMS 用户名（正式环境可无密码）。",
                title="EAMS 登录",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return
        try:
            save_credentials(username, password)
        except Exception as exc:  # noqa: BLE001
            self.set_status(f"保存登录信息失败：{exc}")
            self.show_toast(
                f"保存登录信息失败：{exc}",
                title="EAMS 登录",
                duration_ms=TOAST_DEFAULT_MS,
            )
            return
        self.set_status("EAMS 登录信息已保存")
        self.show_success_toast("登录信息已保存。", title="EAMS 登录")

    def _save_autofill_step_delay(self):
        raw = ""
        if hasattr(self, "autofill_step_delay_entry"):
            raw = self.autofill_step_delay_entry.get().strip()
        if not raw:
            raw = str(DEFAULT_AUTOFILL_STEP_DELAY_SEC)
        try:
            value = float(raw)
        except ValueError:
            self.set_status("步骤间隔必须是数字（秒）")
            self.show_toast(
                "请输入有效的秒数，例如 1 或 1.5。",
                title="自动填写",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return
        self._autofill_step_delay_sec = save_autofill_step_delay_sec(value)
        if hasattr(self, "autofill_step_delay_entry"):
            self.autofill_step_delay_entry.delete(0, "end")
            self.autofill_step_delay_entry.insert(0, f"{self._autofill_step_delay_sec:g}")
        msg = f"步骤间隔已设为 {self._autofill_step_delay_sec:g} 秒"
        self.set_status(msg)
        self.show_success_toast(msg, title="自动填写")

    def _advance_to_next_document_after_review_action(self):
        """After approve/remove, select the next item in the document list."""
        if not self._imported_files:
            self._update_review_cert_status()
            self._refresh_doc_list_marks()
            return
        nxt = self._current_cert_index + 1
        if nxt < len(self._imported_files):
            self._select_document(self._imported_files[nxt])
            return
        # Last item — refresh marks/status on the current selection.
        self._update_review_cert_status()
        self._refresh_doc_list_marks()

    def _on_toggle_approve_entry(self):
        if getattr(self, "_workflow_phase", "extract") != "review":
            self.set_status("请先移出失败证书后再核对批准")
            return
        path = self._current_cert_path()
        if path is None:
            self.set_status("没有可批准的证书")
            return

        if path in self._autofill_queue:
            self._autofill_queue.remove(path)
            self._update_autofill_button()
            name = Path(path).name
            msg = f"已撤销批准 · {name}"
            # Keep UI anchored on the revoked cert so fields match the current selection.
            if path in self._imported_files:
                self._current_cert_index = self._imported_files.index(path)
                self._sync_cert_index_to_list()
                self._load_approve_fields_for_current()
            else:
                self._update_review_cert_status()
            self._refresh_doc_list_marks()
            self.set_status(msg)
            return

        self._save_current_fields_to_result()
        empty_labels = self._empty_review_field_labels()
        if empty_labels:
            missing = "、".join(empty_labels)
            self.set_status(f"批准失败：{missing}")
            self.show_toast(
                f"以下字段不能为空：\n{missing}",
                title="提取核对",
                duration_ms=TOAST_DEFAULT_MS,
            )
            return

        self._removed_paths.discard(path)
        self._autofill_queue.append(path)
        self._update_autofill_button()
        self.set_status(f"已批准，队列 {len(self._autofill_queue)} 份")
        self._refresh_doc_list_marks()
        self._advance_to_next_document_after_review_action()

    def _on_toggle_remove_entry(self):
        if getattr(self, "_workflow_phase", "extract") != "review":
            self.set_status("请先移出失败证书后再核对移除")
            return
        path = self._current_cert_path()
        if path is None:
            self.set_status("没有可操作的证书")
            return

        if path in self._removed_paths:
            self._removed_paths.discard(path)
            self._update_autofill_button()
            self._load_approve_fields_for_current()
            self._refresh_doc_list_marks()
            name = Path(path).name
            msg = f"已撤销移除 · {name}"
            self.set_status(msg)
            return

        if path not in self._imported_files:
            self.set_status("没有需要移除的证书")
            return

        self._removed_paths.add(path)
        if path in self._autofill_queue:
            self._autofill_queue.remove(path)
        self._update_autofill_button()
        name = Path(path).name
        msg = f"已移出核对队列 · {name}"
        self.set_status(msg)
        self._refresh_doc_list_marks()
        self._advance_to_next_document_after_review_action()

    def _on_master_autofill(self):
        """Approve-queue → Excel in import folder → Playwright MAS batch import + fill."""
        # Cancel any fail-folder countdown so export doesn't archive leftovers
        # (including review-removed certs that were still in a pending toast).
        self._cancel_pending_quarantine()
        self._save_fields_before_navigate()
        n = len(self._autofill_queue)
        if n == 0:
            self.set_status("自动填写：队列为空")
            self.show_toast(
                "请先批准至少一份证书。",
                title="导出并自动填写",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return
        if self._autofill_busy:
            self.set_status("自动填写进行中，请稍候…")
            return

        # Untouched certs (neither approved nor removed) → red-dot removed
        # before automation starts, so the list matches what will be filled.
        auto_removed = 0
        approved = set(self._autofill_queue)
        for path in self._imported_files:
            if path in approved or path in self._removed_paths:
                continue
            self._removed_paths.add(path)
            auto_removed += 1
        if auto_removed:
            self._refresh_doc_list_marks()
            self._update_autofill_button()
            self._update_approve_toggle_button()
            self._update_remove_toggle_button()
            self._update_review_fields_state()
            self.set_status(
                f"未操作 {auto_removed} 份已标为移出 · 批准 {n} 份开始自动填写"
            )

        items: list[AutofillItem] = []
        for path in self._autofill_queue:
            result = self._parse_results.get(path)
            if result is None:
                continue
            items.append(AutofillItem(fields=result.fields, pdf_path=path))
        if not items:
            self.set_status("队列中没有可填写的解析结果")
            self.show_toast(
                "队列中没有可填写的解析结果。",
                title="导出并自动填写",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return

        excel_rows = self._export_rows()
        excel_headers = [label for _key, label in EXPORT_COLUMNS]
        excel_path = self._next_excel_export_path()

        try:
            write_batch_excel(excel_rows, excel_headers, excel_path)
        except Exception as exc:  # noqa: BLE001
            self.set_status(f"自动填写失败：{exc}")
            self.show_toast(
                f"生成 Excel 失败：{exc}",
                title="导出并自动填写",
            )
            return

        where = self._excel_export_location_label(excel_path)
        self.show_success_toast(
            f"已导出 {len(excel_rows)} 份到导入文件夹 · {where}",
            title="导出 Excel",
        )
        self._start_autofill_run(
            items,
            excel_path=excel_path,
            excel_headers=excel_headers,
            excel_rows=excel_rows,
            title="导出并自动填写",
            log_note=f"队列导出 · {where}",
        )

    def _on_custom_excel_autofill(self):
        """Pick one generated-format Excel, import to EAMS, upload PDFs by serial."""
        if self._autofill_busy:
            self.set_status("自动填写进行中，请稍候…")
            return
        excel_pick = filedialog.askopenfilename(
            parent=self,
            title="选择自定义导入 Excel",
            filetypes=[
                ("Excel 工作簿", "*.xlsx *.xlsm"),
                ("所有文件", "*.*"),
            ],
        )
        if not excel_pick:
            return
        try:
            excel_path, items, excel_headers, excel_rows = (
                self._load_custom_excel_import(Path(excel_pick))
            )
        except Exception as exc:  # noqa: BLE001
            self.set_status(f"自定义 Excel 无效：{exc}")
            self.show_toast(str(exc), title="自定义 Excel 导入")
            return

        self._start_autofill_run(
            items,
            excel_path=excel_path,
            excel_headers=excel_headers,
            excel_rows=excel_rows,
            title="自定义 Excel 导入",
            log_note=f"自定义 Excel · {excel_path.name}",
        )

    @staticmethod
    def _excel_cell_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _custom_excel_pdf_candidates(self, excel_path: Path) -> list[Path]:
        """PDFs to match against Excel 编号: loaded files, import folder, then Excel folder."""
        seen: set[str] = set()
        out: list[Path] = []

        def add_path(path: Path) -> None:
            if not path.is_file() or path.suffix.lower() != ".pdf":
                return
            key = str(path.resolve())
            if key in seen:
                return
            seen.add(key)
            out.append(path)

        def add_folder(folder: Path | None) -> None:
            if folder is None or not folder.is_dir():
                return
            for pattern in ("*.pdf", "*.PDF"):
                for path in sorted(folder.glob(pattern)):
                    add_path(path)

        for raw in getattr(self, "_imported_files", []) or []:
            add_path(Path(raw))
        add_folder(self._import_folder_root())
        if not out:
            add_folder(excel_path.parent)
        return out

    def _load_custom_excel_import(
        self, excel_path: Path
    ) -> tuple[Path, list[AutofillItem], list[str], list[list[str]]]:
        """Read one VinCert-format Excel and pair each 编号 with a matching PDF."""
        excel_path = Path(excel_path)
        if not excel_path.is_file():
            raise FileNotFoundError("请选择有效的 Excel 文件")
        if excel_path.name.startswith("~$"):
            raise ValueError("请选择正式 Excel 文件，不要选择临时锁文件")
        if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError("请选择 .xlsx / .xlsm 文件")

        from openpyxl import load_workbook  # pyright: ignore[reportMissingModuleSource]

        workbook = load_workbook(excel_path, data_only=True, read_only=True)
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration as exc:
                raise ValueError("Excel 为空") from exc
            headers = [self._excel_cell_text(v) for v in header_row]
            expected = [label for _key, label in EXPORT_COLUMNS]
            missing = [h for h in expected if h not in headers]
            if missing:
                raise ValueError(
                    "Excel 表头不正确，需要包含："
                    + "、".join(expected)
                    + f"。缺少：{'、'.join(missing)}"
                )
            col_index = {label: headers.index(label) for _key, label in EXPORT_COLUMNS}

            data_rows: list[list[str]] = []
            field_rows: list[CertificateFields] = []
            for raw in rows_iter:
                if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                    continue
                values = [self._excel_cell_text(v) for v in raw]
                fields = CertificateFields()
                export_vals: list[str] = []
                for key, label in EXPORT_COLUMNS:
                    idx = col_index[label]
                    cell = values[idx] if idx < len(values) else ""
                    setattr(fields, key, cell)
                    export_vals.append(cell)
                if not any(export_vals):
                    continue
                serial = (fields.serial_num or "").strip()
                if not serial:
                    raise ValueError("Excel 行缺少「编号」，无法匹配 PDF")
                field_rows.append(fields)
                data_rows.append(export_vals)
        finally:
            workbook.close()

        if not field_rows:
            raise ValueError("Excel 没有可导入的数据行")

        pdfs = self._custom_excel_pdf_candidates(excel_path)
        if not pdfs:
            raise FileNotFoundError(
                "未找到可匹配的 PDF。请先加载证书文件夹，或将 PDF 与 Excel 放在同一目录。"
            )

        used: set[str] = set()
        items: list[AutofillItem] = []
        for fields in field_rows:
            serial = (fields.serial_num or "").strip()
            pdf = self._match_pdf_for_serial(serial, pdfs, used=used)
            if pdf is None:
                raise FileNotFoundError(f"找不到编号「{serial}」对应的 PDF")
            used.add(str(pdf))
            items.append(AutofillItem(fields=fields, pdf_path=str(pdf)))

        return excel_path, items, expected, data_rows

    def _match_pdf_for_serial(
        self,
        serial: str,
        pdfs: list[Path],
        *,
        used: set[str],
    ) -> Path | None:
        """Match a PDF to Excel 编号: filename, cached parse, then digital text."""
        serial = (serial or "").strip()
        if not serial:
            return None
        serial_l = serial.casefold()
        serial_compact = re.sub(r"\s+", "", serial_l)
        available = [p for p in pdfs if str(p) not in used]
        if not available:
            return None

        def filename_score(path: Path) -> int:
            stem = path.stem
            stem_l = stem.casefold()
            stem_compact = re.sub(r"\s+", "", stem_l)
            if stem_l == serial_l or stem_compact == serial_compact:
                return 100
            if serial_l in stem_l or serial_compact in stem_compact:
                return 80
            return 0

        ranked = sorted(
            ((filename_score(p), len(p.stem), p) for p in available),
            key=lambda t: (-t[0], t[1], t[2].name.casefold()),
        )
        best_score, _len, best = ranked[0]
        if best_score > 0:
            return best

        parsed_map = getattr(self, "_parse_results", {}) or {}
        for path in available:
            result = parsed_map.get(str(path)) or parsed_map.get(str(path.resolve()))
            if result is None:
                continue
            parsed = re.sub(r"\s+", "", (result.fields.serial_num or "").strip())
            if parsed.casefold() == serial_compact:
                return path

        for path in available:
            try:
                result = parse_certificate(
                    str(path),
                    use_ocr_fallback=False,
                    force_ocr=False,
                    extra_label_aliases=self._extra_label_aliases(),
                )
            except Exception:  # noqa: BLE001
                continue
            parsed = re.sub(r"\s+", "", (result.fields.serial_num or "").strip())
            if parsed.casefold() == serial_compact:
                return path
        return None

    def _start_autofill_run(
        self,
        items: list[AutofillItem],
        *,
        excel_path: Path,
        excel_headers: list[str],
        excel_rows: list[list[str]],
        title: str,
        log_note: str,
    ) -> None:
        """Shared Playwright autofill launcher for queue / custom Excel runs."""
        self._autofill_busy = True
        self._autofill_control = AutofillControl()
        self._show_autofill_run_controls()
        browser_bounds = self._prepare_side_browser_layout()
        step_delay = float(self._autofill_step_delay_sec)
        start_msg = f"开始自动填写 {len(items)} 份…"
        self.set_status(f"{title}：{len(items)} 份")
        self.show_success_toast(start_msg, title=title)
        self.open_autofill_log(title=title)
        self.append_autofill_log(
            f"{log_note} · 步骤间隔 {step_delay:g}s · "
            f"{'测试环境' if self._testing_mode else '正式环境'}"
        )
        if items:
            first = items[0]
            self._focus_certificate_for_autofill(
                first.pdf_path,
                first.fields,
                index=1,
                total=len(items),
                pending=True,
            )

        username, password = self._eams_credentials()
        if not username or not password:
            username, password = load_credentials()
        if username and password:
            try:
                save_credentials(username, password)
            except Exception:  # noqa: BLE001
                pass

        control = self._autofill_control

        def worker():
            profile = self._browser_profile_dir()

            def _run(context, page):
                return run_mas_autofill(
                    items,
                    excel_path=excel_path,
                    excel_headers=excel_headers,
                    excel_rows=excel_rows,
                    username=username,
                    password=password,
                    batch_import=True,
                    fill_details=True,
                    upload_pdf=True,
                    submit_workflow=False,
                    status=lambda msg: self.after(
                        0, lambda m=msg: self._on_autofill_progress(m)
                    ),
                    control=control,
                    testing=bool(self._testing_mode),
                        fill_due_date=not bool(self._valid_till_from_system),
                    window_bounds=browser_bounds,
                    step_delay_sec=step_delay,
                    user_data_dir=profile,
                    shared_context=context,
                    shared_page=page,
                    on_item=lambda i, n, item: self.after(
                        0,
                        lambda i=i, n=n, item=item: self._on_autofill_item(
                            i, n, item
                        ),
                    ),
                )

            try:
                report = self._pdf_preview.run_on_browser(
                    _run,
                    bounds=browser_bounds,
                    profile_dir=profile,
                    accept_downloads=True,
                    slow_mo=100,
                )
                self.after(0, lambda: self._on_autofill_done(report, excel_path))
            except Exception as exc:  # noqa: BLE001
                self.after(0, lambda: self._on_autofill_fail(str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _on_autofill_progress(self, message: str, *, duration_ms: int | None = None):
        del duration_ms  # Kept for call-site compat; lines live in the terminal panel.
        if message.startswith("已保存下载："):
            body = message.removeprefix("已保存下载：").strip()
            if " → " in body:
                name, folder = body.split(" → ", 1)
                toast_body = f"{name}\n保存到：{folder}"
            else:
                toast_body = body or message
            self.show_success_toast(toast_body, title="下载已保存")
            return
        self.append_autofill_log(
            message,
            error=message.startswith("失败"),
        )

    def _restore_layout_after_browser_session(self):
        """Split if Chromium is still open; otherwise fullscreen.

        Do not reopen the PDF tab — leave the EAMS page as the last view.
        """
        self._sync_window_layout_to_browser()
        self._raise_workspace_windows()

    def _refresh_automate_fields_idle(self) -> None:
        path = self._selected_path or self._current_cert_path()
        self._load_automate_fields(path)

    def _on_autofill_done(self, report, excel_path: Path | None = None):
        self._autofill_busy = False
        self._restore_autofill_button()
        self._restore_layout_after_browser_session()

        quarantine = [
            p
            for p in (report.quarantine_paths or [])
            if p in self._imported_files or p in self._autofill_queue
        ]
        # Unique while preserving order.
        seen: set[str] = set()
        quarantine_unique = []
        for path in quarantine:
            if path in seen:
                continue
            seen.add(path)
            quarantine_unique.append(path)
        moved = 0
        copied = 0
        if quarantine_unique:
            moved, copied = self._quarantine_failed_paths(quarantine_unique)
            self._rebuild_doc_list()
            self._update_autofill_button()
            if self._imported_files:
                select = (
                    self._selected_path
                    if self._selected_path in self._imported_files
                    else self._imported_files[0]
                )
                self._select_document(select, sync_preview=False)
            else:
                self._selected_path = None
                if hasattr(self, "field_entries"):
                    self._clear_approve_fields()

        err_n = len(report.errors or [])
        cancelled = bool(getattr(report, "cancelled", False))
        excel_note = f" · {excel_path.name}" if excel_path else ""
        summary = (
            f"自动填写完成{excel_note}"
            f" · 导入Excel {'是' if report.imported_excel else '否'}"
            f" · 填写 {report.filled} · 附件 {report.uploaded}"
        )
        if cancelled:
            summary = (
                f"自动填写已退出{excel_note}"
                f" · 填写 {report.filled} · 附件 {report.uploaded}"
            )
        if moved:
            summary += (
                f" · 移出失败 {moved}（新复制 {copied}）→ "
                f"{self._failed_items_dir_short()}/"
            )
        if err_n:
            summary += f" · 失败 {err_n}"
            if report.errors:
                summary += f"：{report.errors[0]}"
        self.set_status(summary)
        n_ok = len(self._autofill_queue)
        work_ok = (
            not cancelled
            and moved == 0
            and bool(getattr(report, "imported_excel", False))
            and int(report.filled or 0) >= n_ok
            and int(report.uploaded or 0) >= n_ok
            and n_ok > 0
        )
        all_ok = (err_n == 0 and moved == 0 and not cancelled) or work_ok
        if cancelled:
            # Exit already closed the terminal; don't reopen via append.
            self.close_autofill_log()
            self.show_toast(
                summary,
                title="导出并自动填写",
                duration_ms=TOAST_SUCCESS_MS,
            )
        else:
            self.append_autofill_log(summary, error=not all_ok)
            self.finish_autofill_log(ok=all_ok)
            if all_ok:
                self.show_success_toast("Successfully done", title="自动填写")
            else:
                self.show_toast(summary, title="导出并自动填写")
        if hasattr(self, "field_entries"):
            self._load_approve_fields_for_current()
        self._refresh_automate_fields_idle()

    def _on_autofill_fail(self, message: str):
        self._autofill_busy = False
        self._restore_autofill_button()
        self._restore_layout_after_browser_session()
        self.set_status(f"自动填写失败：{message}")
        self.append_autofill_log(f"自动填写失败：{message}", error=True)
        self.finish_autofill_log(ok=False, auto_close_ms=4000)
        self.show_toast(f"自动填写失败：{message}", title="导出并自动填写")
        if hasattr(self, "field_entries"):
            self._load_approve_fields_for_current()
        self._refresh_automate_fields_idle()

    def _export_rows(self) -> list[list[str]]:
        rows: list[list[str]] = []
        for path in self._autofill_queue:
            result = self._parse_results.get(path)
            if result is None:
                continue
            fields = result.fields
            rows.append([getattr(fields, key, "") or "" for key, _label in EXPORT_COLUMNS])
        return rows

    def _on_export_excel(self):
        """Save approved queue Excel into the import folder root (no autofill)."""
        self._cancel_pending_quarantine()
        self._save_fields_before_navigate()
        if self._autofill_busy:
            self.set_status("自动填写进行中，请稍候再导出…")
            return
        rows = self._export_rows()
        if not rows:
            self.set_status("导出 Excel：队列为空")
            self.show_toast(
                "请先批准至少一份证书。",
                title="导出 Excel",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return

        target = self._next_excel_export_path()
        try:
            write_batch_excel(
                rows,
                [label for _key, label in EXPORT_COLUMNS],
                target,
            )
        except Exception as exc:  # noqa: BLE001
            self.set_status(f"导出 Excel 失败：{exc}")
            self.show_toast(f"导出失败：{exc}", title="导出 Excel")
            return

        where = self._excel_export_location_label(target)
        msg = f"已导出 {len(rows)} 份到导入文件夹 · {where}"
        self.set_status(msg)
        self.show_success_toast(msg, title="导出 Excel")

    def _on_close(self):
        if self._autofill_busy:
            self.show_toast(
                "自动填写进行中，请先退出后再关闭窗口。",
                title="自动填写",
                duration_ms=TOAST_SUCCESS_MS,
            )
            return
        try:
            self._pdf_preview.shutdown()
        except Exception:  # noqa: BLE001
            pass
        try:
            close_keepalive_browser()
        except Exception:  # noqa: BLE001
            pass
        self.destroy()


if __name__ == "__main__":
    app = App()
    app.mainloop()
